"""
Enhanced Bulk Upload Module for GHG Emissions Data - Scope 3
Multi-sheet Excel template with cascading dropdowns per category.
Supports all 15 Scope 3 categories (C1-C15) with dynamic validation.
"""

import io
import uuid
import re
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple
from fastapi import APIRouter, HTTPException, UploadFile, File, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.protection import SheetProtection
from rapidfuzz import fuzz, process


# Scope 3 Category mapping (C1-C15) - Updated to match emission_categories table
SCOPE3_CATEGORIES = {
    "C1": "C1 - Purchased Goods and Services",
    "C2": "C2 - Capital Goods",
    "C3": "C3 - Fuel and Energy Related Activities Not Included in Scope 1 or Scope 2",
    "C4": "C4 - Upstream Transportation and Distribution",
    "C5": "C5 - Waste Generated in Operations",
    "C6": "C6 - Business Travel",
    "C7": "C7 - Employee Commuting",
    "C8": "C8 - Upstream Leased Assets",
    "C9": "C9 - Downstream Transportation and Distribution",
    "C10": "C10 - Processing of Sold Products",
    "C11": "C11 - Use of Sold Products",
    "C12": "C12 - End-of-Life Treatment of Sold Products",
    "C13": "C13 - Downstream Leased Assets",
    "C14": "C14 - Franchises",
    "C15": "C15 - Investments",
}

# Reverse mapping
CATEGORY_TO_CODE = {v: k for k, v in SCOPE3_CATEGORIES.items()}


def normalize_string(s: str) -> str:
    """Normalize string for case-insensitive matching."""
    if not s:
        return ""
    normalized = s.lower().strip()
    normalized = re.sub(r'\s+', ' ', normalized)
    return normalized


def find_best_match(value: str, candidates: List[str], threshold: int = 80, use_token_set: bool = False) -> Tuple[Optional[str], int]:
    """Find the best fuzzy match for a value in candidates."""
    if not value or not candidates:
        return None, 0
    
    normalized_value = normalize_string(value)
    normalized_candidates = {normalize_string(c): c for c in candidates}
    
    if normalized_value in normalized_candidates:
        return normalized_candidates[normalized_value], 100
    
    scorer = fuzz.token_set_ratio if use_token_set else fuzz.ratio
    
    result = process.extractOne(
        normalized_value, 
        list(normalized_candidates.keys()),
        scorer=scorer
    )
    
    if result and result[1] >= threshold:
        return normalized_candidates[result[0]], result[1]
    
    return None, 0


def get_suggestions(value: str, candidates: List[str], limit: int = 3) -> List[str]:
    """Get top suggestions for a misspelled value."""
    if not value or not candidates:
        return []
    
    normalized_value = normalize_string(value)
    normalized_candidates = {normalize_string(c): c for c in candidates}
    
    results = process.extract(
        normalized_value,
        list(normalized_candidates.keys()),
        scorer=fuzz.ratio,
        limit=limit
    )
    
    return [normalized_candidates[r[0]] for r in results if r[1] >= 50]


def create_enhanced_bulk_upload_router(db, get_current_user, get_admin_user):
    """Create the enhanced bulk upload router with multi-sheet support."""
    
    router = APIRouter(tags=["Bulk Upload"])
    
    # Style definitions
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    CATEGORY_FILL = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    EXAMPLE_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    INSTRUCTION_HEADER = Font(bold=True, size=14, color="1F4E79")
    INSTRUCTION_SUBHEADER = Font(bold=True, size=12, color="2E7D32")
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    async def get_reference_data(org_id: str) -> Dict[str, Any]:
        """Fetch all reference data needed for template generation and validation."""
        
        facilities = await db.facilities.find(
            {"organization_id": org_id, "is_active": {"$ne": False}},
            {"_id": 0, "id": 1, "name": 1}
        ).to_list(1000)
        
        scopes = await db.scopes.find(
            {"is_active": {"$ne": False}},
            {"_id": 0, "id": 1, "name": 1, "code": 1}
        ).to_list(100)
        
        categories = await db.emission_categories.find(
            {"is_active": {"$ne": False}},
            {"_id": 0, "id": 1, "name": 1, "scope_id": 1}
        ).to_list(500)
        
        units = await db.units.find(
            {"is_active": {"$ne": False}},
            {"_id": 0, "symbol": 1, "name": 1, "unit_type": 1}
        ).to_list(500)
        
        scope3_ef = await db.scope3_ef.find(
            {},
            {"_id": 0, "id": 1, "activity": 1, "method": 1, "category": 1, "scope": 1, 
             "emission_factor": 1, "unit": 1, "allowed_units": 1, "default_unit": 1}
        ).to_list(5000)
        
        # Get Scope 3 ID
        scope3_id = next((s["id"] for s in scopes if "3" in s.get("name", "")), None)
        
        # Build category -> method -> activities mapping
        cat_method_activities = {}
        cat_methods = {}
        activity_units = {}
        
        for ef in scope3_ef:
            cat = ef.get("category", "")
            method = ef.get("method", "")
            activity = ef.get("activity", "")
            allowed_units = ef.get("allowed_units", [])
            
            if not cat or not method:
                continue
            
            if cat not in cat_method_activities:
                cat_method_activities[cat] = {}
                cat_methods[cat] = set()
            
            cat_methods[cat].add(method)
            
            if method not in cat_method_activities[cat]:
                cat_method_activities[cat][method] = set()
            
            if activity:
                cat_method_activities[cat][method].add(activity)
                
                # Store allowed units per activity
                if activity not in activity_units:
                    activity_units[activity] = set()
                activity_units[activity].update(allowed_units)
        
        # Convert sets to sorted lists
        for cat in cat_method_activities:
            for method in cat_method_activities[cat]:
                cat_method_activities[cat][method] = sorted(cat_method_activities[cat][method])
        
        for cat in cat_methods:
            cat_methods[cat] = sorted(cat_methods[cat])
        
        for activity in activity_units:
            activity_units[activity] = sorted(activity_units[activity])
        
        # Physical units (exclude CO2 composite units)
        excluded_patterns = ['co2', 'ch4', 'n2o', '/', 'per']
        physical_unit_symbols = [
            u["symbol"] for u in units 
            if not any(p in u["symbol"].lower() for p in excluded_patterns)
        ]
        
        # EF units with tCO2e ONLY in numerator (for supplier-based method)
        # Filter to only tCO2e units (not kgCO2e) as per user request
        common_ef_units = [
            "tCO2e/kg", "tCO2e/t", "tCO2e/g", "tCO2e/lb",
            "tCO2e/L", "tCO2e/kL", "tCO2e/m3", "tCO2e/gal", "tCO2e/ml",
            "tCO2e/kWh", "tCO2e/MWh", "tCO2e/GJ", "tCO2e/TJ", "tCO2e/MMBtu",
            "tCO2e/km", "tCO2e/mi", "tCO2e/m",
            "tCO2e/USD", "tCO2e/INR", "tCO2e/EUR", "tCO2e/GBP",
            "tCO2e/passenger.km", "tCO2e/t.km", "tCO2e/vehicle.km",
            "tCO2e/Room*night", "tCO2e/working_hour", "tCO2e/unit",
        ]
        ef_units_list = sorted(common_ef_units)
        
        return {
            "facilities": facilities,
            "facility_names": [f["name"] for f in facilities],
            "scopes": scopes,
            "categories": categories,
            "units": units,
            "scope3_ef": scope3_ef,
            "scope3_id": scope3_id,
            "cat_method_activities": cat_method_activities,
            "cat_methods": cat_methods,
            "activity_units": activity_units,
            "physical_unit_symbols": physical_unit_symbols,
            "unit_symbols": [u["symbol"] for u in units],
            "ef_units": ef_units_list,  # EF units with tCO2e numerator only
        }
    
    def get_category_columns(category_code: str) -> List[Tuple[str, str, int]]:
        """Get column definitions for a category. Returns (key, label, width)."""
        base_columns = [
            ("facility", "Facility Name *", 25),
            ("calculation_method", "Calculation Method *", 20),
            ("activity", "Activity *", 35),
            ("supplier_name", "Supplier Name", 25),
            ("supplier_code", "Supplier ID/Code", 18),
            ("process_name", "Process Name", 25),
            ("process_description", "Process Description", 30),
            ("responsible_name", "Responsible Person Name", 25),
            ("responsible_designation", "Responsible Person Designation", 25),
            ("responsible_contact", "Responsible Person Contact", 25),
            ("reporting_period", "Reporting Period (YYYY-MM) *", 22),
            ("activity_value", "Activity Value *", 15),
            ("activity_unit", "Activity Value Unit *", 20),
            ("ef_supplier", "Emission Factor (Supplier Based)", 25),
            ("ef_unit_supplier", "EF Unit (Supplier Based)", 22),
            ("notes", "Notes", 30),
        ]
        
        # Add Employee fields for C7
        if category_code == "C7":
            # Insert after supplier_code
            idx = next(i for i, (k, _, _) in enumerate(base_columns) if k == "supplier_code") + 1
            base_columns.insert(idx, ("employee_name", "Employee Name", 25))
            base_columns.insert(idx + 1, ("employee_id", "Employee ID", 18))
        
        return base_columns
    
    @router.get("/bulk-upload/template")
    async def download_scope3_template(current_user: dict = Depends(get_admin_user)):
        """Generate enhanced Excel template for Scope 3 emissions with multi-sheet support."""
        
        org_id = current_user.get("organization_id")
        ref_data = await get_reference_data(org_id)
        
        wb = Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # ========== CREATE INSTRUCTIONS SHEET ==========
        instr_ws = wb.create_sheet("Instructions", 0)
        
        instructions = [
            ("GHG Scope 3 Emissions - Bulk Upload Template", INSTRUCTION_HEADER),
            ("", None),
            ("IMPORTANT: Sheet names and structure are locked. Do not rename or reorder sheets.", Font(bold=True, color="FF0000")),
            ("", None),
            ("Template Structure:", INSTRUCTION_SUBHEADER),
            ("• Each Scope 3 category has its own sheet (C1-C15)", None),
            ("• Only fill sheets for categories relevant to your organization", None),
            ("• Delete example rows (blue highlighted) before uploading", None),
            ("• The '_Data' sheet contains reference data - do not modify", None),
            ("", None),
            ("Column Descriptions:", INSTRUCTION_SUBHEADER),
            ("• Facility Name * - Select from dropdown (your organization's facilities)", None),
            ("• Calculation Method * - Select based on available data:", None),
            ("    - activity_basis: When you have activity data (e.g., 5000 km travel)", None),
            ("    - spend_basis: When you have monetary spend data (e.g., ₹50,000)", None),
            ("    - supplier_basis: When supplier provides emission factor directly", None),
            ("• Activity * - Select from dropdown based on category and method", None),
            ("• Supplier Name/Code - Optional supplier information", None),
            ("• Process Name/Description - Business process details", None),
            ("• Responsible Person - Contact for this data", None),
            ("• Reporting Period * - Format: YYYY-MM (e.g., 2024-01)", None),
            ("• Activity Value * - Numeric value (quantity or spend amount)", None),
            ("• Activity Value Unit * - Unit matching the method:", None),
            ("    - For activity_basis: Physical units (kg, t, km, kWh, L)", None),
            ("    - For spend_basis: Currency (INR, USD, EUR)", None),
            ("• Emission Factor (Supplier Based) - Only for supplier_basis method", None),
            ("• EF Unit (Supplier Based) - Required if EF is provided (dropdown with tCO2e/kgCO2e units)", None),
            ("", None),
            ("Supplier-Based Method Notes:", INSTRUCTION_SUBHEADER),
            ("• Use supplier_basis when your supplier provides their own emission factor", None),
            ("• Activity dropdown shows all activities from activity_basis and spend_basis", None),
            ("• EF Unit must have tCO2e or kgCO2e in the numerator (e.g., kgCO2e/kg, tCO2e/L)", None),
            ("", None),
            ("Category-Specific Notes:", INSTRUCTION_SUBHEADER),
            ("• C7 (Employee Commuting) - Has additional Employee Name/ID columns", None),
            ("• Only categories with emission factor data have populated Activity dropdowns", None),
            ("", None),
            ("Validation Rules:", INSTRUCTION_SUBHEADER),
            ("• All fields marked with * are required", None),
            ("• Dropdowns enforce valid selections", None),
            ("• System will fuzzy-match values to handle minor typos", None),
            ("• Errors will be highlighted with suggestions after upload", None),
            ("", None),
            ("Tips:", INSTRUCTION_SUBHEADER),
            ("• Copy-paste from your existing data sources", None),
            ("• Use the Reference sheet '_Data' to see all valid values", None),
            ("• Upload will validate before saving - you can fix errors and re-upload", None),
        ]
        
        for row_idx, (text, font) in enumerate(instructions, 1):
            cell = instr_ws.cell(row=row_idx, column=1, value=text)
            if font:
                cell.font = font
        
        instr_ws.column_dimensions['A'].width = 100
        instr_ws.protection = SheetProtection(sheet=True, objects=True, scenarios=True)
        
        # ========== CREATE DATA REFERENCE SHEET (HIDDEN) ==========
        data_ws = wb.create_sheet("_Data")
        
        # Column A: Facilities
        data_ws.cell(row=1, column=1, value="Facilities").font = Font(bold=True)
        for idx, f in enumerate(ref_data["facility_names"], 2):
            data_ws.cell(row=idx, column=1, value=f)
        
        # Column B: All Methods
        data_ws.cell(row=1, column=2, value="Methods").font = Font(bold=True)
        all_methods = sorted(set(m for methods in ref_data["cat_methods"].values() for m in methods))
        for idx, m in enumerate(all_methods, 2):
            data_ws.cell(row=idx, column=2, value=m)
        
        # Column C: Currency Units
        data_ws.cell(row=1, column=3, value="Currencies").font = Font(bold=True)
        currencies = ["INR", "USD", "EUR", "GBP", "JPY", "AUD", "CAD", "CNY", "CHF", "SGD"]
        for idx, c in enumerate(currencies, 2):
            data_ws.cell(row=idx, column=3, value=c)
        
        # Column D: Physical Units
        data_ws.cell(row=1, column=4, value="Physical Units").font = Font(bold=True)
        for idx, u in enumerate(ref_data["physical_unit_symbols"][:50], 2):
            data_ws.cell(row=idx, column=4, value=u)
        
        # Column E: EF Units (tCO2e/kgCO2e based)
        data_ws.cell(row=1, column=5, value="EF Units (Supplier Based)").font = Font(bold=True)
        ef_units = ref_data.get("ef_units", [])
        for idx, u in enumerate(ef_units, 2):
            data_ws.cell(row=idx, column=5, value=u)
        
        # Columns F onwards: Category-specific activities
        col_offset = 6
        category_activity_ranges = {}  # Store ranges for INDIRECT formulas
        
        for cat_code, cat_name in SCOPE3_CATEGORIES.items():
            # Find matching category in scope3_ef
            cat_data = ref_data["cat_method_activities"].get(cat_name, {})
            methods = ref_data["cat_methods"].get(cat_name, [])
            
            if not methods:
                continue
            
            # Write methods for this category
            method_col = col_offset
            data_ws.cell(row=1, column=method_col, value=f"{cat_code}_Methods").font = Font(bold=True)
            for idx, m in enumerate(methods, 2):
                data_ws.cell(row=idx, column=method_col, value=m)
            
            # Create named range for methods (for reference)
            # method_range_name = f"{cat_code}_Methods"
            # method_range = f"'_Data'!${get_column_letter(method_col)}$2:${get_column_letter(method_col)}${len(methods)+1}"
            
            col_offset += 1
            
            # Write activities for each method
            for method in methods:
                activities = cat_data.get(method, [])
                if not activities:
                    continue
                
                act_col = col_offset
                col_name = f"{cat_code}_{method}"
                data_ws.cell(row=1, column=act_col, value=col_name).font = Font(bold=True)
                for idx, a in enumerate(activities, 2):
                    data_ws.cell(row=idx, column=act_col, value=a)
                
                # Store range for this category/method combination
                category_activity_ranges[f"{cat_code}_{method}"] = {
                    "col": get_column_letter(act_col),
                    "count": len(activities)
                }
                
                col_offset += 1
        
        # Adjust column widths
        for col in range(1, col_offset):
            data_ws.column_dimensions[get_column_letter(col)].width = 30
        
        # Protect and hide data sheet
        data_ws.protection = SheetProtection(sheet=True, objects=True, scenarios=True)
        data_ws.sheet_state = 'hidden'
        
        # ========== CREATE CATEGORY SHEETS (C1-C15) ==========
        for cat_code, cat_name in SCOPE3_CATEGORIES.items():
            ws = wb.create_sheet(cat_code)
            
            # Get columns for this category
            columns = get_category_columns(cat_code)
            
            # Category header row
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
            header_cell = ws.cell(row=1, column=1, value=f"{cat_code}: {cat_name}")
            header_cell.fill = CATEGORY_FILL
            header_cell.font = Font(bold=True, color="FFFFFF", size=12)
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 25
            
            # Column headers row
            for col_idx, (key, label, width) in enumerate(columns, 1):
                cell = ws.cell(row=2, column=col_idx, value=label)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = THIN_BORDER
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            
            ws.row_dimensions[2].height = 35
            
            # Get column indices
            col_map = {key: idx for idx, (key, _, _) in enumerate(columns, 1)}
            
            # ===== DATA VALIDATIONS =====
            
            # Facility dropdown
            if ref_data["facility_names"]:
                fac_count = len(ref_data["facility_names"])
                fac_dv = DataValidation(
                    type="list", 
                    formula1=f"'_Data'!$A$2:$A${fac_count+1}",
                    allow_blank=False
                )
                fac_dv.error = "Please select a valid facility"
                fac_dv.errorTitle = "Invalid Facility"
                ws.add_data_validation(fac_dv)
                fac_dv.add(f"{get_column_letter(col_map['facility'])}3:{get_column_letter(col_map['facility'])}1000")
            
            # Calculation Method dropdown (category-specific + always include supplier_basis)
            methods = ref_data["cat_methods"].get(cat_name, [])
            # Always ensure supplier_basis is available for all categories
            methods_set = set(methods)
            methods_set.add("supplier_basis")
            methods = sorted(methods_set)
            
            method_list = ",".join(methods)
            method_dv = DataValidation(
                type="list",
                formula1=f'"{method_list}"',
                allow_blank=False
            )
            method_dv.error = f"Valid methods for {cat_code}: {method_list}"
            method_dv.errorTitle = "Invalid Method"
            ws.add_data_validation(method_dv)
            method_dv.add(f"{get_column_letter(col_map['calculation_method'])}3:{get_column_letter(col_map['calculation_method'])}1000")
            
            # Activity dropdown - use all activities for this category (deduplicated)
            # For supplier_basis, activities from all methods are available
            cat_data = ref_data["cat_method_activities"].get(cat_name, {})
            all_activities = set()
            for method_activities in cat_data.values():
                all_activities.update(method_activities)
            
            if all_activities:
                # Limit to 100 for dropdown (Excel limitation)
                activities_list = sorted(all_activities)[:100]
                act_str = ",".join(activities_list)
                
                # For long lists, use a range reference
                if len(act_str) > 200:
                    # Find the column in _Data that has activities for any method of this category
                    for method in methods:
                        range_key = f"{cat_code}_{method}"
                        if range_key in category_activity_ranges:
                            range_info = category_activity_ranges[range_key]
                            act_dv = DataValidation(
                                type="list",
                                formula1=f"'_Data'!${range_info['col']}$2:${range_info['col']}${range_info['count']+1}",
                                allow_blank=False
                            )
                            break
                    else:
                        act_dv = DataValidation(type="list", formula1=f'"{",".join(activities_list[:50])}"', allow_blank=False)
                else:
                    act_dv = DataValidation(type="list", formula1=f'"{act_str}"', allow_blank=False)
                
                act_dv.error = "Please select a valid activity"
                act_dv.errorTitle = "Invalid Activity"
                ws.add_data_validation(act_dv)
                act_dv.add(f"{get_column_letter(col_map['activity'])}3:{get_column_letter(col_map['activity'])}1000")
            
            # Activity Value Unit dropdown - Hybrid approach
            # Common physical units + currencies (validation on upload enforces activity-specific rules)
            common_physical_units = [
                "kg", "t", "g", "lb", "oz",  # Mass
                "L", "kL", "m3", "gal", "ml",  # Volume
                "km", "mi", "m", "ft",  # Distance
                "kWh", "MWh", "GJ", "TJ", "MMBtu",  # Energy
                "passenger.km", "t.km", "vehicle.km",  # Transport
                "Room*night", "working_hour", "unit", "piece",  # Other
            ]
            all_units = currencies + common_physical_units
            unit_str = ",".join(all_units)
            unit_dv = DataValidation(type="list", formula1=f'"{unit_str}"', allow_blank=False)
            unit_dv.error = "Select a valid unit (will be validated against activity's allowed units)"
            unit_dv.errorTitle = "Select Unit"
            ws.add_data_validation(unit_dv)
            unit_dv.add(f"{get_column_letter(col_map['activity_unit'])}3:{get_column_letter(col_map['activity_unit'])}1000")
            
            # EF Unit (Supplier Based) dropdown - tCO2e based units only
            ef_units = ref_data.get("ef_units", [])
            if ef_units:
                # Limit to fit Excel dropdown
                ef_units_str = ",".join(ef_units[:40])
                ef_unit_dv = DataValidation(type="list", formula1=f'"{ef_units_str}"', allow_blank=True)
                ef_unit_dv.error = "Select a valid emission factor unit (must have tCO2e in numerator)"
                ef_unit_dv.errorTitle = "Invalid EF Unit"
                ws.add_data_validation(ef_unit_dv)
                ef_unit_dv.add(f"{get_column_letter(col_map['ef_unit_supplier'])}3:{get_column_letter(col_map['ef_unit_supplier'])}1000")
            
            # Activity Value validation - must be greater than 0
            value_dv = DataValidation(
                type="decimal",
                operator="greaterThan",
                formula1="0",
                allow_blank=False
            )
            value_dv.error = "Activity value must be a number greater than 0"
            value_dv.errorTitle = "Invalid Value"
            ws.add_data_validation(value_dv)
            value_dv.add(f"{get_column_letter(col_map['activity_value'])}3:{get_column_letter(col_map['activity_value'])}1000")
            
            # Example row
            example_row = 3
            example_data = {
                "facility": ref_data["facility_names"][0] if ref_data["facility_names"] else "[Your Facility]",
                "calculation_method": methods[0] if methods else "activity_basis",
                "activity": sorted(all_activities)[0] if all_activities else "[Select Activity]",
                "supplier_name": "Example Supplier",
                "supplier_code": "SUP001",
                "process_name": "Example Process",
                "process_description": "Description of process",
                "responsible_name": "John Doe",
                "responsible_designation": "Sustainability Manager",
                "responsible_contact": "+91-9876543210",
                "reporting_period": "2024-01",
                "activity_value": 100,
                "activity_unit": "kg",
                "ef_supplier": "",
                "ef_unit_supplier": "",
                "notes": "Example entry - delete before upload",
            }
            
            if cat_code == "C7":
                example_data["employee_name"] = "Jane Smith"
                example_data["employee_id"] = "EMP001"
            
            for key, _, _ in columns:
                col_idx = col_map[key]
                cell = ws.cell(row=example_row, column=col_idx, value=example_data.get(key, ""))
                cell.fill = EXAMPLE_FILL
                cell.border = THIN_BORDER
            
            # Protect sheet structure (but allow data entry)
            ws.protection = SheetProtection(
                sheet=True,
                objects=True,
                scenarios=True,
                formatCells=False,
                formatColumns=False,
                formatRows=False,
                insertColumns=False,
                insertRows=True,
                insertHyperlinks=False,
                deleteColumns=False,
                deleteRows=True,
                selectLockedCells=False,
                sort=True,
                autoFilter=True,
                pivotTables=False,
                selectUnlockedCells=False
            )
            
            # Unlock data cells
            for row in range(3, 1001):
                for col in range(1, len(columns) + 1):
                    ws.cell(row=row, column=col).protection = Protection(locked=False)
        
        # Save workbook
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"GHG_Scope3_BulkUpload_Template_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    @router.post("/bulk-upload/validate")
    async def validate_enhanced_upload(
        file: UploadFile = File(...),
        current_user: dict = Depends(get_admin_user)
    ):
        """Parse and validate uploaded multi-sheet Scope 3 Excel file."""
        
        if not file.filename.endswith('.xlsx'):
            raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
        
        org_id = current_user.get("organization_id")
        upload_id = str(uuid.uuid4())
        
        contents = await file.read()
        
        try:
            wb = load_workbook(io.BytesIO(contents))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")
        
        ref_data = await get_reference_data(org_id)
        
        # Build lookup maps
        facility_map = {normalize_string(f["name"]): f for f in ref_data["facilities"]}
        
        all_rows = []
        total_valid = 0
        total_invalid = 0
        category_summaries = {}
        
        # Process each category sheet
        for cat_code, cat_name in SCOPE3_CATEGORIES.items():
            if cat_code not in wb.sheetnames:
                continue
            
            ws = wb[cat_code]
            columns = get_category_columns(cat_code)
            key_map = {key: idx-1 for idx, (key, _, _) in enumerate(columns, 1)}
            
            cat_methods = ref_data["cat_methods"].get(cat_name, [])
            cat_activities = set()
            for method_acts in ref_data["cat_method_activities"].get(cat_name, {}).values():
                cat_activities.update(method_acts)
            
            cat_valid = 0
            cat_invalid = 0
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
                if not any(row):
                    continue
                
                # Skip example rows (check facility, notes, and any cell containing "example")
                row_lower = [str(cell).lower() if cell else "" for cell in row]
                is_example = any(
                    "example" in cell or "[your facility]" in cell or "delete before upload" in cell
                    for cell in row_lower
                )
                if is_example:
                    continue
                
                errors = []
                matched_data = {
                    "scope": "Scope 3",
                    "scope_id": ref_data["scope3_id"],
                    "category_code": cat_code,
                    "category": cat_name,
                }
                
                # Extract row data
                row_data = {}
                for key, _, _ in columns:
                    idx = key_map.get(key)
                    row_data[key] = row[idx] if idx is not None and idx < len(row) else None
                
                # ===== REQUIRED FIELDS =====
                required = ["facility", "calculation_method", "activity", "reporting_period", "activity_value", "activity_unit"]
                for field in required:
                    if not row_data.get(field):
                        errors.append({
                            "column": field,
                            "message": f"Required field '{field}' is missing",
                            "suggestion": "Please provide a value"
                        })
                
                # ===== FACILITY VALIDATION =====
                if row_data.get("facility"):
                    facility_match, _ = find_best_match(str(row_data["facility"]), ref_data["facility_names"])
                    if facility_match:
                        matched_data["facility"] = facility_match
                        matched_data["facility_id"] = facility_map[normalize_string(facility_match)]["id"]
                    else:
                        suggestions = get_suggestions(str(row_data["facility"]), ref_data["facility_names"])
                        errors.append({
                            "column": "facility",
                            "message": f"Facility '{row_data['facility']}' not found",
                            "suggestion": f"Did you mean: {', '.join(suggestions)}" if suggestions else "Check your facilities"
                        })
                
                # ===== METHOD VALIDATION =====
                if row_data.get("calculation_method"):
                    method_str = normalize_string(str(row_data["calculation_method"]))
                    method_aliases = {
                        "spend_basis": ["spend_basis", "spend", "spend-basis", "spendbasis", "spend_based", "spend based"],
                        "activity_basis": ["activity_basis", "activity", "activity-basis", "activitybasis", "activity_based", "activity based"],
                        "supplier_basis": ["supplier_basis", "supplier", "supplier-basis", "supplierbasis", "supplier_based", "supplier based"],
                    }
                    
                    matched_method = None
                    for canonical, aliases in method_aliases.items():
                        if method_str in [normalize_string(a) for a in aliases]:
                            matched_method = canonical
                            break
                    
                    if matched_method:
                        # supplier_basis is always allowed for all categories
                        if matched_method == "supplier_basis":
                            matched_data["calculation_method"] = matched_method
                        elif matched_method in cat_methods or matched_method.replace("_basis", "_based") in cat_methods:
                            matched_data["calculation_method"] = matched_method
                        else:
                            errors.append({
                                "column": "calculation_method",
                                "message": f"Method '{matched_method}' not available for {cat_code}",
                                "suggestion": f"Available methods: {', '.join(cat_methods + ['supplier_basis'])}"
                            })
                    else:
                        errors.append({
                            "column": "calculation_method",
                            "message": f"Invalid method '{row_data['calculation_method']}'",
                            "suggestion": f"Use: {', '.join(cat_methods)}"
                        })
                
                # ===== ACTIVITY VALIDATION =====
                if row_data.get("activity"):
                    activity_match, _ = find_best_match(str(row_data["activity"]), list(cat_activities))
                    if activity_match:
                        matched_data["activity"] = activity_match
                    else:
                        suggestions = get_suggestions(str(row_data["activity"]), list(cat_activities))
                        errors.append({
                            "column": "activity",
                            "message": f"Activity '{row_data['activity']}' not found for {cat_code}",
                            "suggestion": f"Did you mean: {', '.join(suggestions[:3])}" if suggestions else "Check Scope 3 EF data"
                        })
                
                # ===== REPORTING PERIOD FORMAT =====
                if row_data.get("reporting_period"):
                    period_str = str(row_data["reporting_period"])
                    if not re.match(r'^\d{4}-\d{2}$', period_str):
                        errors.append({
                            "column": "reporting_period",
                            "message": "Invalid date format",
                            "suggestion": "Use YYYY-MM format (e.g., 2024-01)"
                        })
                    else:
                        matched_data["reporting_period"] = period_str
                
                # ===== ACTIVITY VALUE VALIDATION =====
                if row_data.get("activity_value"):
                    try:
                        matched_data["activity_value"] = float(row_data["activity_value"])
                    except (ValueError, TypeError):
                        errors.append({
                            "column": "activity_value",
                            "message": "Activity value must be a number",
                            "suggestion": f"Got '{row_data['activity_value']}'"
                        })
                
                # ===== UNIT VALIDATION =====
                if row_data.get("activity_unit") and matched_data.get("calculation_method"):
                    unit_str = str(row_data["activity_unit"]).strip()
                    currency_units = ["INR", "USD", "EUR", "GBP", "JPY", "AUD", "CAD", "CNY", "CHF", "SGD"]
                    
                    # Get allowed units for the matched activity (if available)
                    activity_allowed_units = []
                    if matched_data.get("activity"):
                        activity_allowed_units = ref_data["activity_units"].get(matched_data["activity"], [])
                    
                    if matched_data["calculation_method"] == "spend_basis":
                        if unit_str.upper() not in currency_units:
                            errors.append({
                                "column": "activity_unit",
                                "message": f"Unit '{unit_str}' invalid for spend_basis",
                                "suggestion": f"Use currency: {', '.join(currency_units[:5])}"
                            })
                        else:
                            matched_data["activity_unit"] = unit_str.upper()
                    else:
                        if unit_str.upper() in currency_units:
                            # Show specific allowed units for the activity
                            if activity_allowed_units:
                                suggestion = f"Allowed units for '{matched_data.get('activity', 'this activity')}': {', '.join(activity_allowed_units[:8])}"
                                if len(activity_allowed_units) > 8:
                                    suggestion += f" (+{len(activity_allowed_units) - 8} more)"
                            else:
                                suggestion = "Use physical unit (kg, t, km, kWh, L)"
                            errors.append({
                                "column": "activity_unit",
                                "message": f"Currency '{unit_str}' invalid for {matched_data['calculation_method']}",
                                "suggestion": suggestion
                            })
                        else:
                            # Validate against activity-specific allowed units if available
                            if activity_allowed_units:
                                # Check if unit matches any allowed unit (case-insensitive)
                                unit_lower = unit_str.lower()
                                matched_unit = None
                                for allowed in activity_allowed_units:
                                    if unit_lower == allowed.lower():
                                        matched_unit = allowed
                                        break
                                
                                if matched_unit:
                                    matched_data["activity_unit"] = matched_unit
                                else:
                                    # Fuzzy match against allowed units
                                    unit_match, score = find_best_match(unit_str, activity_allowed_units)
                                    if unit_match and score >= 80:
                                        matched_data["activity_unit"] = unit_match
                                    else:
                                        # Show specific allowed units in error
                                        suggestion = f"Allowed units: {', '.join(activity_allowed_units[:8])}"
                                        if len(activity_allowed_units) > 8:
                                            suggestion += f" (+{len(activity_allowed_units) - 8} more)"
                                        errors.append({
                                            "column": "activity_unit",
                                            "message": f"Unit '{unit_str}' not in allowed units for '{matched_data.get('activity', 'this activity')}'",
                                            "suggestion": suggestion
                                        })
                            else:
                                # Fallback: match against physical unit symbols
                                unit_match, _ = find_best_match(unit_str, ref_data["physical_unit_symbols"])
                                if unit_match:
                                    matched_data["activity_unit"] = unit_match
                                else:
                                    matched_data["activity_unit"] = unit_str  # Accept as-is, may be valid
                
                # ===== SUPPLIER-BASED EMISSION FACTOR =====
                if row_data.get("ef_supplier"):
                    try:
                        matched_data["ef_supplier"] = float(row_data["ef_supplier"])
                        if not row_data.get("ef_unit_supplier"):
                            errors.append({
                                "column": "ef_unit_supplier",
                                "message": "EF unit required when emission factor is provided",
                                "suggestion": "Specify unit (e.g., tCO2e/t)"
                            })
                        else:
                            matched_data["ef_unit_supplier"] = str(row_data["ef_unit_supplier"]).strip()
                    except (ValueError, TypeError):
                        errors.append({
                            "column": "ef_supplier",
                            "message": "Emission factor must be a number",
                            "suggestion": f"Got '{row_data['ef_supplier']}'"
                        })
                
                # ===== OPTIONAL FIELDS =====
                optional_fields = ["supplier_name", "supplier_code", "process_name", "process_description", 
                                   "responsible_name", "responsible_designation", "responsible_contact", "notes"]
                if cat_code == "C7":
                    optional_fields.extend(["employee_name", "employee_id"])
                
                for field in optional_fields:
                    if row_data.get(field):
                        matched_data[field] = str(row_data[field])
                
                # Record result
                status = "valid" if not errors else "invalid"
                if status == "valid":
                    cat_valid += 1
                    total_valid += 1
                else:
                    cat_invalid += 1
                    total_invalid += 1
                
                all_rows.append({
                    "sheet": cat_code,
                    "row_number": row_idx,
                    "status": status,
                    "original_data": {k: str(v) if v is not None else "" for k, v in row_data.items()},
                    "matched_data": matched_data,
                    "errors": errors
                })
            
            if cat_valid + cat_invalid > 0:
                category_summaries[cat_code] = {
                    "category_name": cat_name,
                    "valid_rows": cat_valid,
                    "invalid_rows": cat_invalid
                }
        
        # Store upload session
        upload_session = {
            "id": upload_id,
            "organization_id": org_id,
            "uploaded_by": current_user.get("id"),
            "uploaded_by_email": current_user.get("email"),
            "filename": file.filename,
            "template_type": "scope3_enhanced",
            "total_rows": len(all_rows),
            "valid_rows": total_valid,
            "invalid_rows": total_invalid,
            "category_summaries": category_summaries,
            "rows": all_rows,
            "status": "pending",
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        
        await db.bulk_upload_sessions.insert_one(upload_session)
        
        return {
            "upload_id": upload_id,
            "template_type": "scope3_enhanced",
            "summary": {
                "total_rows": len(all_rows),
                "valid_rows": total_valid,
                "invalid_rows": total_invalid,
                "categories": category_summaries
            },
            "rows": all_rows
        }
    
    @router.post("/bulk-upload/{upload_id}/save")
    async def save_valid_rows(
        upload_id: str,
        save_mode: str = Query("valid_only", description="valid_only or all_or_nothing"),
        current_user: dict = Depends(get_admin_user)
    ):
        """Save valid rows from enhanced upload session."""
        
        org_id = current_user.get("organization_id")
        
        session = await db.bulk_upload_sessions.find_one(
            {"id": upload_id, "organization_id": org_id},
            {"_id": 0}
        )
        
        if not session:
            raise HTTPException(status_code=404, detail="Upload session not found")
        
        if session["status"] == "completed":
            raise HTTPException(status_code=400, detail="This upload has already been processed")
        
        rows = session.get("rows", [])
        valid_rows = [r for r in rows if r["status"] == "valid"]
        
        if save_mode == "all_or_nothing" and session["invalid_rows"] > 0:
            raise HTTPException(
                status_code=400, 
                detail=f"Cannot save: {session['invalid_rows']} invalid rows. Fix errors and re-upload."
            )
        
        if not valid_rows:
            raise HTTPException(status_code=400, detail="No valid rows to save")
        
        # Fetch EF data for calculations
        scope3_ef_map = {}
        ef_entries = await db.scope3_ef.find({}, {"_id": 0}).to_list(5000)
        for ef in ef_entries:
            key = f"{ef.get('category')}|{ef.get('activity')}|{ef.get('method')}"
            scope3_ef_map[key] = ef
        
        saved_count = 0
        emissions_to_insert = []
        
        for row in valid_rows:
            matched = row["matched_data"]
            
            # Find emission factor
            ef_key = f"{matched.get('category')}|{matched.get('activity')}|{matched.get('calculation_method')}"
            ef_data = scope3_ef_map.get(ef_key)
            
            # Use supplier EF if provided, otherwise use database EF
            emission_factor = matched.get("ef_supplier")
            ef_unit = matched.get("ef_unit_supplier")
            
            if not emission_factor and ef_data:
                emission_factor = ef_data.get("emission_factor")
                ef_unit = ef_data.get("unit")
            
            # Calculate emissions if EF available
            activity_value = matched.get("activity_value", 0)
            calculated_emissions = None
            if emission_factor and activity_value:
                calculated_emissions = float(activity_value) * float(emission_factor)
            
            emission_record = {
                "id": str(uuid.uuid4()),
                "facility_id": matched.get("facility_id"),
                "organization_id": org_id,
                "reporting_period": f"{matched.get('reporting_period')}-15",
                "scope": "scope3",
                "category": matched.get("category"),
                "sub_category": matched.get("activity"),
                "scope3_activity": matched.get("activity"),
                "calculation_method_scope3": matched.get("calculation_method"),
                "supplier_name": matched.get("supplier_name"),
                "supplier_code": matched.get("supplier_code"),
                "employee_name": matched.get("employee_name"),
                "employee_id": matched.get("employee_id"),
                "process_names": [matched.get("process_name")] if matched.get("process_name") else [],
                "process_descriptions": [{"name": matched.get("process_name", ""), "description": matched.get("process_description", "")}] if matched.get("process_name") else [],
                "responsible_person": matched.get("responsible_name"),
                "responsible_person_designation": matched.get("responsible_designation"),
                "responsible_person_contact": matched.get("responsible_contact"),
                "notes": matched.get("notes"),
                "dynamic_field_values": {
                    "activity_value": {"value": activity_value, "unit": matched.get("activity_unit")},
                },
                "outputs": {
                    "total": {
                        "value": calculated_emissions,
                        "unit": "tCO2e"
                    }
                } if calculated_emissions else {},
                "emission_factor_used": emission_factor,
                "emission_factor_unit": ef_unit,
                "source": "bulk_upload",
                "bulk_upload_id": upload_id,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": current_user.get("id"),
                "created_by_email": current_user.get("email"),
                "created_by_name": current_user.get("name", current_user.get("email")),
            }
            
            emissions_to_insert.append(emission_record)
            saved_count += 1
        
        if emissions_to_insert:
            await db.emission_records.insert_many(emissions_to_insert)
        
        # Update session status
        await db.bulk_upload_sessions.update_one(
            {"id": upload_id},
            {"$set": {
                "status": "completed",
                "saved_rows": saved_count,
                "completed_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        # Create audit log
        audit_entry = {
            "id": str(uuid.uuid4()),
            "action": "bulk_upload_save",
            "entity_type": "emissions",
            "entity_id": upload_id,
            "organization_id": org_id,
            "user_id": current_user.get("id"),
            "user_email": current_user.get("email"),
            "details": {
                "filename": session.get("filename"),
                "total_rows": session.get("total_rows"),
                "saved_rows": saved_count,
                "categories": list(session.get("category_summaries", {}).keys())
            },
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.audit_logs.insert_one(audit_entry)
        
        return {
            "message": f"Successfully saved {saved_count} emission records",
            "saved_count": saved_count,
            "upload_id": upload_id
        }
    
    @router.get("/bulk-upload/sessions")
    async def list_upload_sessions(
        current_user: dict = Depends(get_admin_user),
        limit: int = Query(20, le=100)
    ):
        """List recent upload sessions for the organization."""
        
        org_id = current_user.get("organization_id")
        
        sessions = await db.bulk_upload_sessions.find(
            {"organization_id": org_id},
            {"_id": 0, "rows": 0}  # Exclude large rows array
        ).sort("created_at", -1).limit(limit).to_list(limit)
        
        return sessions
    
    @router.get("/bulk-upload/{upload_id}/errors")
    async def download_errors_excel(
        upload_id: str,
        current_user: dict = Depends(get_admin_user)
    ):
        """Download Excel file with error annotations."""
        
        org_id = current_user.get("organization_id")
        
        session = await db.bulk_upload_sessions.find_one(
            {"id": upload_id, "organization_id": org_id},
            {"_id": 0}
        )
        
        if not session:
            raise HTTPException(status_code=404, detail="Upload session not found")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Validation Errors"
        
        # Headers
        headers = ["Sheet", "Row", "Status", "Column", "Error", "Suggestion", "Original Value"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
        
        # Error rows
        row_idx = 2
        for row in session.get("rows", []):
            if row["status"] == "invalid":
                for error in row.get("errors", []):
                    ws.cell(row=row_idx, column=1, value=row["sheet"])
                    ws.cell(row=row_idx, column=2, value=row["row_number"])
                    ws.cell(row=row_idx, column=3, value="INVALID")
                    ws.cell(row=row_idx, column=4, value=error.get("column", ""))
                    ws.cell(row=row_idx, column=5, value=error.get("message", ""))
                    ws.cell(row=row_idx, column=6, value=error.get("suggestion", ""))
                    ws.cell(row=row_idx, column=7, value=row["original_data"].get(error.get("column", ""), ""))
                    
                    for col in range(1, 8):
                        ws.cell(row=row_idx, column=col).border = THIN_BORDER
                    
                    row_idx += 1
        
        # Adjust column widths
        widths = [10, 8, 10, 20, 40, 40, 30]
        for col_idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"BulkUpload_Errors_{upload_id[:8]}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    return router
