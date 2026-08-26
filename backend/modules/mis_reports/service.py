"""Shared aggregation, export, history, and delivery helpers for MIS Reports."""
import io
import uuid
import re
import hashlib
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional

from dateutil.relativedelta import relativedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from shared.database.mongo import db
from shared.helpers.email import send_email_with_attachments
from modules.esg_records.services.dashboard.dashboard_metrics_service import get_dashboard_metrics_service
from modules.dashboards.environment_detail_service import get_environment_detail
from modules.esg_records.services.dashboard.unit_utils import to_kilolitres
from modules.mis_reports.reporting_period_service import ReportingPeriodService
from modules.supplier_assessment.service import SupplierAssessmentService
from shared.utils.emission_records import eligible_ghg_record_filter


ALL_SCOPES = ["scope1", "scope2", "scope3", "biogenic"]


def _governance_period_filter(period_str: str) -> dict:
    """Convert a 'YYYY-MM' string into a MongoDB filter matching the nested reporting_period object."""
    year, month = int(period_str[:4]), int(period_str[5:7])
    return {"reporting_period.year": year, "reporting_period.month": {"$in": [str(month), month]}}


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def numeric_emissions(record: Dict[str, Any]) -> float:
    for field in ("total_emissions", "calculated_co2e", "co2e_emissions"):
        try:
            return float(record.get(field) or 0)
        except (TypeError, ValueError):
            continue
    return 0.0


async def organization_facility_ids(current_user: dict) -> List[str]:
    if current_user.get("role") == "super_admin":
        facilities = await db.facilities.find({}, {"_id": 0, "id": 1}).to_list(5000)
    else:
        facilities = await db.facilities.find(
            {"organization_id": current_user.get("organization_id")},
            {"_id": 0, "id": 1},
        ).to_list(5000)
    return [facility["id"] for facility in facilities]


async def aggregate_emissions(filters: Dict[str, Any], current_user: dict) -> Dict[str, Any]:
    from shared.utils.period_utils import period_variants

    allowed_facilities = await organization_facility_ids(current_user)
    requested_facilities = filters.get("facility_ids") or allowed_facilities
    facility_ids = [facility_id for facility_id in requested_facilities if facility_id in allowed_facilities]
    scopes = filters.get("scopes") or ALL_SCOPES
    categories = filters.get("categories") or []
    report_year = int(filters["reporting_period_start"][:4])
    fiscal_periods = period_variants(report_year, "FY")
    query: Dict[str, Any] = {
        "facility_id": {"$in": facility_ids},
        "scope": {"$in": scopes},
    }
    if filters.get("strict_period"):
        if filters.get("period_frequency") in {"daily", "weekly"}:
            # Daily/weekly reporting is exact only for records stored as ISO dates; monthly records are intentionally not stretched across shorter periods.
            query["reporting_period"] = {"$gte": filters["period_start_date"], "$lte": filters["period_end_date"]}
        else:
            query["reporting_period"] = {"$gte": filters["reporting_period_start"], "$lte": filters["reporting_period_end"]}
    else:
        query["$or"] = [
            {"reporting_period": {"$gte": filters["reporting_period_start"], "$lte": filters["reporting_period_end"]}},
            {"reporting_period": {"$in": fiscal_periods}},
        ]
    if categories:
        query["category"] = {"$in": categories}
    query.update(eligible_ghg_record_filter())
    records = await db.emission_records.find(query, {"_id": 0}).to_list(10000)

    facility_names = {
        facility["id"]: facility.get("name", facility["id"])
        for facility in await db.facilities.find({"id": {"$in": facility_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(5000)
    }
    grouped = {"scope": {}, "category": {}, "facility": {}, "period": {}}
    total = 0.0
    for record in records:
        value = numeric_emissions(record)
        total += value
        values = {
            "scope": record.get("scope", "unknown"),
            "category": record.get("category", "Uncategorized"),
            "facility": facility_names.get(record.get("facility_id"), "Unknown facility"),
            "period": record.get("reporting_period", "Unknown period"),
        }
        for key, name in values.items():
            grouped[key][name] = grouped[key].get(name, 0) + value

    def breakdown(items: Dict[str, float], field: str) -> List[Dict[str, Any]]:
        return [{field: name, "emissions": round(value, 4)} for name, value in sorted(items.items(), key=lambda item: item[1], reverse=True)]

    return {
        "total_emissions": round(total, 4),
        "unit": "tCO2e",
        "record_count": len(records),
        "scope_breakdown": breakdown(grouped["scope"], "scope"),
        "category_breakdown": breakdown(grouped["category"], "category"),
        "facility_breakdown": breakdown(grouped["facility"], "facility"),
        "period_breakdown": breakdown(grouped["period"], "period"),
    }


def previous_period_filters(filters: Dict[str, Any]) -> Dict[str, Any]:
    start = datetime.strptime(filters["reporting_period_start"], "%Y-%m")
    end = datetime.strptime(filters["reporting_period_end"], "%Y-%m")
    month_count = (end.year - start.year) * 12 + end.month - start.month + 1
    previous_start = start - relativedelta(months=month_count)
    previous_end = end - relativedelta(months=month_count)
    return {**filters, "reporting_period_start": previous_start.strftime("%Y-%m"), "reporting_period_end": previous_end.strftime("%Y-%m")}


def percentage_change(current: float, previous: float) -> float | None:
    if not previous:
        return None
    return round(((current - previous) / previous) * 100, 2)


def target_direction_and_status(actual: Optional[float], target: Optional[float], direction: Optional[str]) -> tuple[str, str]:
    """Classify targets using their stored goal direction—never a one-size-fits-all ratio."""
    if direction not in {"decrease", "increase", "maintain"}:
        return "Not configured", "Direction required"
    if actual is None or target is None:
        return direction, "No Data"
    if direction == "decrease":
        return direction, "On Track" if actual <= target else ("At Risk" if actual <= target * 1.1 else "Behind")
    if direction == "increase":
        return direction, "Achieved" if actual >= target else ("At Risk" if actual >= target * 0.9 else "Behind")
    return direction, "Achieved" if actual == target else "At Risk"


def inferred_legacy_target_direction(name: str) -> str:
    """Temporary fallback for legacy targets until their explicit direction is configured."""
    value = (name or "").lower()
    if any(term in value for term in ("renewable", "recycle", "recovery", "recycling")):
        return "increase"
    if any(term in value for term in ("ghg", "emission", "water consumption", "energy", "waste generation", "waste generated")):
        return "decrease"
    return "maintain"


def comparison_status(current: float, previous: float, direction: str) -> tuple[Optional[float], str]:
    """Return meaningful management status without dividing by zero or treating every decrease as good."""
    if previous == 0:
        return None, "New activity / No comparable baseline" if current else "No activity in either period"
    if current == 0:
        return -100.0, "No current-period activity"
    change = percentage_change(current, previous)
    if change is None or abs(change) < 0.05:
        return change, "No material change"
    if direction == "lower":
        if change > 100:
            return change, "Large period-over-period change — review recommended"
        return change, "Improving" if change < 0 else "Needs attention"
    if direction == "higher":
        return change, "Improving" if change > 0 else "Needs attention"
    return change, "Informational"



async def _enrich_targets_with_progress(targets_raw: List[Dict], organization_id: str) -> List[Dict]:
    """Compute actual progress for each target using kpi_calculator — mirrors /targets/with-progress."""
    from modules.kpi_engine import kpi_calculator
    from modules.esg_targets.router import (
        _calculate_progress,
        _get_denominator_for_intensity,
        _get_period_for_target,
        _resolve_target_value,
    )

    enriched = []
    for t in targets_raw:
        entry: Dict[str, Any] = {
            "name": t.get("target_name", "Unnamed"),
            "target_value": t.get("target_value"),
            "baseline_value": (t.get("baseline") or {}).get("value"),
            "baseline_period": (t.get("baseline") or {}).get("period"),
            "unit": t.get("unit", ""),
            "category": t.get("category", ""),
            "section": t.get("section", "environment"),
            "tracking_mode": t.get("tracking_mode", "static"),
            "kpi_name": t.get("kpi_name", ""),
            "reporting_period": t.get("reporting_period", ""),
            "actual_value": None,
            "progress_pct": None,
            "gap": None,
            "target_type": t.get("target_type", "absolute"),
            "goal_type": t.get("goal_type", "upper_limit"),
            "target_direction": t.get("target_direction") or t.get("percentage_direction") or inferred_legacy_target_direction(t.get("name") or t.get("target_name", "")),
            "status": "No Data",
            "tracking_values": t.get("tracking_values"),
        }
        kpi_id = t.get("kpi_id")
        if kpi_id:
            try:
                period = _get_period_for_target(t)
                facility_ids = t.get("facility_ids") if t.get("scope_type") == "facility" else None
                calculation = await kpi_calculator.calculate(
                    kpi_id=kpi_id,
                    org_id=organization_id,
                    scope_type=t.get("scope_type", "organization"),
                    facility_ids=facility_ids,
                    period=period,
                )
                actual_value = calculation.get("value")
                target_value = _resolve_target_value(t, period)
                goal_type = t.get("goal_type", "upper_limit")

                # Intensity targets: divide actual by revenue/production denominator
                if t.get("target_type") in ("intensity_revenue", "intensity_production"):
                    denom = await _get_denominator_for_intensity(t, organization_id, period)
                    if denom.get("error") or not denom.get("value"):
                        actual_value = None
                    elif actual_value is not None:
                        actual_value = round(actual_value / denom["value"], 6)

                # Percentage targets: recalculate absolute target from baseline × percentage
                if t.get("target_type") == "percentage" and t.get("percentage_amount"):
                    baseline = t.get("baseline") or {}
                    bv = float(baseline.get("value", 0)) if baseline else 0
                    pct = float(t.get("percentage_amount", 0))
                    if bv and pct:
                        direction = t.get("percentage_direction", "decrease")
                        new_tv = bv * (1 + pct / 100) if direction == "increase" else bv * (1 - pct / 100)
                        if target_value is None or abs((target_value or 0) - new_tv) > 0.01:
                            target_value = new_tv

                progress = _calculate_progress(actual_value, target_value, goal_type, t)
                entry["actual_value"] = actual_value
                entry["progress_pct"] = round(progress["percentage"], 1) if progress["percentage"] is not None else None
                if target_value is not None:
                    entry["target_value"] = target_value
                if actual_value is not None and target_value is not None:
                    entry["gap"] = round(actual_value - target_value, 4)
                direction, status = target_direction_and_status(actual_value, target_value, entry["target_direction"])
                entry["target_direction"] = direction
                entry["status"] = status
            except Exception:
                pass
        enriched.append(entry)
    return enriched


async def _enrich_sbti_targets_with_progress(targets_raw: List[Dict], organization_id: str) -> List[Dict]:
    """Map separately stored SBTi targets into the MIS target presentation model."""
    from modules.kpi_engine import kpi_calculator
    from modules.esg_targets.router import _get_denominator_for_intensity, _get_org_reporting_type
    from modules.sbti_targets.service import compute_achievement

    reporting_type = await _get_org_reporting_type(organization_id)
    current_year = datetime.now(timezone.utc).year
    enriched = []
    for target in targets_raw:
        target_type = target.get("target_type", "percentage")
        is_intensity = target_type in {"intensity_revenue", "intensity_production"}
        target_value = target.get("target_intensity") if is_intensity else target.get("target_value")
        baseline_value = target.get("base_year_intensity") if is_intensity else target.get("base_year_value")
        actual_value = None
        kpi_id = target.get("kpi_id")
        if kpi_id:
            try:
                calculation = await kpi_calculator.calculate(
                    kpi_id=kpi_id,
                    org_id=organization_id,
                    scope_type="organization",
                    facility_ids=None,
                    period={"year": current_year},
                )
                actual_value = calculation.get("value")
                if is_intensity and actual_value is not None:
                    denominator = await _get_denominator_for_intensity(
                        {"target_type": target_type, "scope_type": "organization", "facility_ids": [], "_reporting_type": reporting_type},
                        organization_id,
                        {"year": current_year},
                    )
                    actual_value = round(actual_value / denominator["value"], 6) if denominator.get("value") else None
            except Exception:
                actual_value = None

        progress_pct = compute_achievement(
            "intensity" if is_intensity else "percentage",
            target_intensity=target_value,
            current_intensity=actual_value,
            base_intensity=baseline_value,
            target_value=target_value,
            current_value=actual_value,
            base_value=baseline_value,
        )
        direction, status = target_direction_and_status(actual_value, target_value, "decrease")
        enriched.append({
            "name": target.get("target_name", "SBTi Target"),
            "target_value": target_value,
            "baseline_value": baseline_value,
            "baseline_period": target.get("base_year", ""),
            "unit": target.get("unit", ""),
            "category": "SBTi",
            "section": "sbti",
            "tracking_mode": "static",
            "kpi_name": target.get("kpi_name", ""),
            "reporting_period": target.get("target_year", ""),
            "actual_value": actual_value,
            "progress_pct": progress_pct,
            "gap": round(actual_value - target_value, 4) if actual_value is not None and target_value is not None else None,
            "target_type": target_type,
            "goal_type": "upper_limit",
            "target_direction": direction,
            "status": status,
            "tracking_values": None,
            "target_source": "sbti",
            "term_type": target.get("term_type", ""),
        })
    return enriched


async def build_executive_mis_report(filters: Dict[str, Any], current_user: dict, reporting_context: Optional[Dict[str, Any]] = None, selected_sections: Optional[List[str]] = None) -> Dict[str, Any]:
    """Build a factual ESG MIS data pack; no generative metrics are invented."""
    current = await aggregate_emissions(filters, current_user)
    if reporting_context:
        frequency = reporting_context["frequency"]
        previous_filters = ReportingPeriodService.filters_for(filters, reporting_context["comparison_period"], frequency)
        ytd_filters = ReportingPeriodService.filters_for(filters, reporting_context["ytd_period"], frequency)
        previous_ytd_filters = ReportingPeriodService.filters_for(filters, reporting_context["previous_ytd_period"], frequency)
    else:
        previous_filters = previous_period_filters(filters)
        ytd_filters = filters
        previous_ytd_filters = previous_period_filters(filters)
    previous = await aggregate_emissions(previous_filters, current_user)
    ytd = await aggregate_emissions(ytd_filters, current_user)
    previous_ytd = await aggregate_emissions(previous_ytd_filters, current_user)
    previous_scopes = {row["scope"]: row["emissions"] for row in previous["scope_breakdown"]}
    current_scopes = {row["scope"]: row["emissions"] for row in current["scope_breakdown"]}
    previous_ytd_scopes = {row["scope"]: row["emissions"] for row in previous_ytd["scope_breakdown"]}
    ytd_scopes = {row["scope"]: row["emissions"] for row in ytd["scope_breakdown"]}
    kpis = [{"label": "Total Emissions", "value": current["total_emissions"], "previous": previous["total_emissions"], "ytd": ytd["total_emissions"], "previous_ytd": previous_ytd["total_emissions"], "unit": current["unit"], "change_pct": percentage_change(current["total_emissions"], previous["total_emissions"]), "ytd_change_pct": percentage_change(ytd["total_emissions"], previous_ytd["total_emissions"])}]
    for scope in ALL_SCOPES:
        kpis.append({"label": scope.replace("scope", "Scope ").title(), "value": current_scopes.get(scope, 0), "previous": previous_scopes.get(scope, 0), "ytd": ytd_scopes.get(scope, 0), "previous_ytd": previous_ytd_scopes.get(scope, 0), "unit": current["unit"], "change_pct": percentage_change(current_scopes.get(scope, 0), previous_scopes.get(scope, 0)), "ytd_change_pct": percentage_change(ytd_scopes.get(scope, 0), previous_ytd_scopes.get(scope, 0))})

    organization_id = current_user.get("organization_id")
    facility_ids = filters.get("facility_ids") or await organization_facility_ids(current_user)
    # Use the selected reporting period with the dashboard metric services.
    all_facilities = await organization_facility_ids(current_user)
    is_all_facilities = set(facility_ids) == set(all_facilities)
    esg_facility_filter = None if is_all_facilities else facility_ids
    current_resources = await build_resource_snapshot(organization_id, esg_facility_filter, filters)
    previous_resources = await build_resource_snapshot(organization_id, esg_facility_filter, previous_filters)
    relationships = await db.supplier_relationships.find({"customer_org_id": organization_id, "is_active": True, "is_deleted": {"$ne": True}}, {"_id": 0, "supplier_name": 1, "supplier_org_name": 1, "company_name": 1, "overall_score": 1, "esg_score": 1, "ghg_score": 1, "overall_completion_percent": 1, "esg_completion_percent": 1, "ghg_completion_percent": 1, "invitation_status": 1}).to_list(1000)
    # Get dynamically computed supplier rankings (same as dashboard)
    supplier_svc = SupplierAssessmentService()
    supplier_rankings_data = await supplier_svc.get_supplier_rankings(organization_id)
    supplier_rankings = supplier_rankings_data.get("rankings", [])
    targets_raw = await db.esg_targets.find(
        {"organization_id": organization_id, "is_deleted": {"$ne": True}, "status": "active"},
        {"_id": 0},
    ).to_list(100)
    sbti_targets_raw = await db.sbti_targets.find(
        {"organization_id": organization_id, "status": "active"},
        {"_id": 0},
    ).to_list(100)
    targets = await _enrich_targets_with_progress(targets_raw, organization_id)
    targets.extend(await _enrich_sbti_targets_with_progress(sbti_targets_raw, organization_id))
    frameworks = await db.organization_esg_responses.find({"organization_id": organization_id}, {"_id": 0, "framework": 1, "approval_status": 1}).to_list(10000)
    compliance: Dict[str, Dict[str, int]] = {}
    for item in frameworks:
        framework = item.get("framework") or "Unclassified"
        compliance.setdefault(framework, {"total": 0, "complete": 0})["total"] += 1
        if item.get("approval_status") in {"approved", "not_required"}: compliance[framework]["complete"] += 1
    compliance_rows = [{"framework": name, "completion_pct": round((values["complete"] / values["total"] * 100) if values["total"] else 0, 1)} for name, values in compliance.items()]

    insights, actions = [], []
    energy, water, waste = current_resources["energy"], current_resources["water"], current_resources["waste"]
    energy_total = energy.get("total", 0) or 0
    incidents = await db.governance_records.count_documents({
        "org_id": organization_id,
        **_governance_period_filter(filters["reporting_period_start"]),
        "approval_status": {"$in": ["approved", "not_required", None]},
        "$or": [
            {"subcategory": "Health & Safety Incidents"},
            {"category": {"$regex": "data breach", "$options": "i"}},
            {"subcategory": {"$regex": "data breach", "$options": "i"}},
            {"category": {"$regex": "violation", "$options": "i"}},
            {"subcategory": {"$regex": "violation", "$options": "i"}},
        ],
    })
    operational = await get_operational_kpis(organization_id, current_scopes.get("scope1", 0) + current_scopes.get("scope2", 0), energy_total, incidents, filters)
    total_change = kpis[0]["change_pct"]
    if total_change is not None:
        insights.append(f"Emissions {'decreased' if total_change < 0 else 'increased'} {abs(total_change):.1f}% versus the comparable previous period.")
    previous_facilities = {row["facility"]: row["emissions"] for row in previous["facility_breakdown"]}
    facility_comparisons = []
    for row in current["facility_breakdown"]:
        prior_value = previous_facilities.get(row["facility"], 0)
        change_pct, status = comparison_status(row["emissions"], prior_value, "lower")
        facility_comparisons.append({"facility": row["facility"], "current": row["emissions"], "previous": prior_value, "change_pct": change_pct, "status": status})
    if current["facility_breakdown"] and current["total_emissions"]:
        top = current["facility_breakdown"][0]
        concentration = round(top["emissions"] / current["total_emissions"] * 100, 1)
        insights.append(f"{top['facility']} contributed {concentration:.1f}% of current-period emissions.")
    energy_change, energy_status = comparison_status(energy_total, previous_resources["energy"].get("total", 0) or 0, "lower")
    water_change, water_status = comparison_status(water.get("consumption", 0) or 0, previous_resources["water"].get("consumption", 0) or 0, "lower")
    renewable_change, renewable_status = comparison_status(energy.get("renewable_pct", 0) or 0, previous_resources["energy"].get("renewable_pct", 0) or 0, "higher")
    if energy_change is not None and energy_change > 20:
        insights.append(f"Energy consumption increased {energy_change:.1f}% and requires investigation.")
    if energy_total and not energy.get("renewable_pct"):
        insights.append("Renewable energy contribution remains at 0%, creating a clear improvement opportunity.")
    if water_change is not None and water_change > 5:
        insights.append(f"Water consumption increased {water_change:.1f}% versus the comparable previous period and requires review.")
    if supplier_rankings:
        high_risk = sum(1 for r in supplier_rankings if r.get("overall_score") is not None and r["overall_score"] < 50)
        insights.append(f"{len(supplier_rankings)} suppliers have been assessed, with {high_risk} currently high-risk.")
    target_counts = {status: sum(1 for target in targets if target.get("status") == status) for status in ("On Track", "At Risk", "Behind", "Achieved")}
    active_targets = len(targets)
    if active_targets:
        insights.append(f"{active_targets} active targets: {target_counts['On Track'] + target_counts['Achieved']} on track or achieved, {target_counts['At Risk']} at risk, and {target_counts['Behind']} behind.")
    if target_counts["Behind"]:
        actions.append({"priority": "High", "area": "Targets", "action": f"Review {target_counts['Behind']} target{'s' if target_counts['Behind'] != 1 else ''} currently behind plan"})
    if energy_status == "Large period-over-period change — review recommended":
        actions.append({"priority": "High", "area": "Energy", "action": f"Investigate {energy_change:.1f}% increase in consumption"})
    if energy_total and not energy.get("renewable_pct"):
        actions.append({"priority": "Medium", "area": "Renewable Energy", "action": "Develop a renewable-energy improvement plan"})
    if water_status in {"Needs attention", "Large period-over-period change — review recommended"}:
        actions.append({"priority": "Medium", "area": "Water", "action": "Review current consumption against target and prior period"})
    availability = {"current": "available" if current["record_count"] else "No data available for this reporting period.", "comparison": "available" if previous["record_count"] else "Previous-period comparison unavailable.", "previous_ytd": "available" if previous_ytd["record_count"] else "Previous FY/CY YTD comparison unavailable."}
    resource_status = {"energy": energy_status, "water": water_status, "renewable": renewable_status, "waste_recovery": comparison_status(waste.get("recovery_pct", 0) or 0, previous_resources["waste"].get("recovery_pct", 0) or 0, "higher")[1]}
    high_priority = sum(1 for action in actions if action["priority"] == "High")
    overall_management_status = "Attention Required" if high_priority or target_counts["Behind"] else ("Monitor" if actions or target_counts["At Risk"] else "On Track")
    twelve_month_emissions_trend = await build_twelve_month_emissions_trend(filters, current_user, reporting_context)
    twelve_month_resource_trends = await build_twelve_month_resource_trends(filters, current_user, reporting_context)
    operational_trends = await build_twelve_month_operational_trends(filters, current_user, reporting_context)
    result = {"filters": filters, "reporting_context": reporting_context, "selected_sections": selected_sections or [], "current": current, "previous": previous, "ytd": ytd, "previous_ytd": previous_ytd, "kpis": kpis, "energy": energy, "water": water, "waste": waste, "previous_resources": previous_resources, "resource_status": resource_status, "operational_kpis": operational, "compliance": compliance_rows, "supplier_assessment": {"suppliers_assessed": len(supplier_rankings), "high_risk_suppliers": sum(1 for r in supplier_rankings if r.get("overall_score") is not None and r["overall_score"] < 50), "pending_assessments": sum(1 for row in relationships if row.get("invitation_status") not in {"completed", "accepted"})}, "supplier_scores": supplier_rankings, "targets": targets, "target_summary": {"active": active_targets, **target_counts}, "insights": insights[:7], "actions": actions, "overall_management_status": {"status": overall_management_status, "high_priority_count": high_priority}, "facility_comparisons": facility_comparisons, "monthly_trend": current["period_breakdown"], "twelve_month_emissions_trend": twelve_month_emissions_trend, "twelve_month_resource_trends": twelve_month_resource_trends, "twelve_month_operational_trends": operational_trends, "availability": availability}
    result["executive_summary"] = await build_executive_summary_data(result, filters, current_user)
    result["emissions_deep"] = await build_emissions_deep_data(result, filters, current_user)
    result["resources_deep"] = await _build_resources_deep(result, current_user)
    result["facility_deep"] = await build_facility_deep_data(result, filters, current_user)
    _attach_monthly_target_history(result)
    return result


async def dashboard_recycled_water(organization_id: str, facility_ids: Optional[List[str]], filters: Optional[Dict[str, Any]] = None) -> float:
    """Apply the Environment dashboard's approved-record scope to Water Recycle."""
    query: Dict[str, Any] = {"org_id": organization_id, "category": "Water", "subcategory": "Recycle", "approval_status": {"$in": ["approved", "not_required", None]}}
    if facility_ids:
        query["facility_id"] = {"$in": facility_ids}
    if filters:
        query["reporting_period"] = {"$gte": filters["reporting_period_start"], "$lte": filters["reporting_period_end"]}
    records = await db.environment_records.find(query, {"_id": 0, "field_values": 1}).to_list(10000)
    total = 0.0
    for record in records:
        values = record.get("field_values") or {}
        try:
            total += to_kilolitres(float(values.get("total_quantity_of_water_recycled") or 0), values.get("unit"))
        except (TypeError, ValueError):
            continue
    return round(total, 2)


async def build_resource_snapshot(organization_id: str, facility_ids: Optional[List[str]], filters: Dict[str, Any]) -> Dict[str, Dict[str, float]]:
    """Use the dashboard metric service as the single source of truth for MIS resources."""
    dashboard = await get_dashboard_metrics_service(db).get_dashboard_metrics(
        organization_id, facility_ids, start_date=filters["reporting_period_start"], end_date=filters["reporting_period_end"]
    )
    energy = dict(dashboard.get("energy", {}))
    energy_total = energy.get("total", 0) or 0
    energy["renewable_pct"] = round(((energy.get("renewable_total", 0) or 0) / energy_total * 100) if energy_total else 0, 2)
    water = dict(dashboard.get("water", {}))
    water["recycle_pct"] = round((water.get("recycled", 0) / water["totalinput"] * 100) if water["totalinput"] else 0, 2)
    waste = dict(dashboard.get("waste", {}))
    return {"energy": energy, "water": water, "waste": waste}


async def build_twelve_month_emissions_trend(filters: Dict[str, Any], current_user: dict, context: Optional[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Map each month of current and prior FY/CY YTD for the management trend chart."""
    if not context:
        return []
    end = datetime.fromisoformat(context["reporting_period"]["start_date"]).date().replace(day=1)
    months = [(end - relativedelta(months=index)).strftime("%Y-%m") for index in range(11, -1, -1)]
    facility_ids = filters.get("facility_ids") or await organization_facility_ids(current_user)
    records = await db.emission_records.find({"facility_id": {"$in": facility_ids}, "scope": {"$in": filters.get("scopes") or ALL_SCOPES}, "reporting_period": {"$in": months}, **eligible_ghg_record_filter()}, {"_id": 0, "reporting_period": 1, "co2e_emissions": 1}).to_list(100000)
    totals, seen = {key: 0.0 for key in months}, set()
    for row in records:
        try: totals[row.get("reporting_period", "")] += float(row.get("co2e_emissions") or 0); seen.add(row.get("reporting_period", ""))
        except (TypeError, ValueError): continue
    return [{"period": key, "value": round(totals[key], 2) if key in seen else None} for key in months]


async def build_twelve_month_resource_trends(filters: Dict[str, Any], current_user: dict, context: Optional[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    if context:
        end = datetime.fromisoformat(context["reporting_period"]["start_date"]).date().replace(day=1)
    elif filters.get("reporting_period_start"):
        end = datetime.strptime(filters["reporting_period_start"], "%Y-%m").date().replace(day=1)
    else:
        return {}
    months = [(end - relativedelta(months=index)).strftime("%Y-%m") for index in range(11, -1, -1)]
    facilities = filters.get("facility_ids") or await organization_facility_ids(current_user)
    all_facilities = await organization_facility_ids(current_user)
    facility_filter = None if set(facilities) == set(all_facilities) else facilities
    metrics = {"energy": ("energy", "total"), "water_consumption": ("water", "consumption"), "water_withdrawal": ("water", "withdrawal"), "water_discharge": ("water", "discharge"), "water_recycle": ("water", "recycled"), "waste_generated": ("waste", "generated"), "waste_recovery": ("waste", "recovery_pct"), "renewable_energy": ("energy", "renewable_pct"), "energy_renewable_total": ("energy", "renewable_total"), "energy_non_renewable_total": ("energy", "non_renewable_total"), "waste_recovered": ("waste", "recovered"), "waste_disposal": ("waste", "disposal"), "waste_haz_generated": ("waste", "hazardous_generated"), "waste_haz_recovered": ("waste", "hazardous_recovered"), "waste_haz_disposed": ("waste", "hazardous_disposed"), "waste_nonhaz_generated": ("waste", "non_hazardous_generated"), "waste_nonhaz_recovered": ("waste", "non_hazardous_recovered"), "waste_nonhaz_disposed": ("waste", "non_hazardous_disposed")}
    result = {key: [] for key in metrics}
    for period in months:
        snapshot = await build_resource_snapshot(current_user.get("organization_id"), facility_filter, {**filters, "reporting_period_start": period, "reporting_period_end": period, "strict_period": True})
        for key, (section, field) in metrics.items(): result[key].append({"period": period, "value": snapshot[section].get(field)})
    return result


async def build_twelve_month_operational_trends(filters: Dict[str, Any], current_user: dict, context: Optional[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    if context:
        end = datetime.fromisoformat(context["reporting_period"]["start_date"]).date().replace(day=1)
    elif filters.get("reporting_period_start"):
        end = datetime.strptime(filters["reporting_period_start"], "%Y-%m").date().replace(day=1)
    else:
        return {}
    months = [(end - relativedelta(months=index)).strftime("%Y-%m") for index in range(11, -1, -1)]
    org_id = current_user.get("organization_id"); result = {"incidents": [], "ltifr": [], "account_payable_days": []}
    incident_filter_base = {"org_id": org_id, "approval_status": {"$in": ["approved", "not_required", None]}, "$or": [{"subcategory": "Health & Safety Incidents"}, {"category": {"$regex": "data breach", "$options": "i"}}, {"subcategory": {"$regex": "data breach", "$options": "i"}}, {"category": {"$regex": "violation", "$options": "i"}}, {"subcategory": {"$regex": "violation", "$options": "i"}}]}
    for period in months:
        period_filters = {**filters, "reporting_period_start": period, "reporting_period_end": period, "strict_period": True}
        emissions = await aggregate_emissions(period_filters, current_user)
        resources = await build_resource_snapshot(org_id, None, period_filters)
        incidents = await db.governance_records.count_documents({**incident_filter_base, **_governance_period_filter(period)})
        ops = await get_operational_kpis(org_id, emissions["total_emissions"], resources["energy"].get("total", 0) or 0, incidents, period_filters)
        result["incidents"].append({"period": period, "value": incidents})
        result["ltifr"].append({"period": period, "value": ops.get("ltifr")})
        result["account_payable_days"].append({"period": period, "value": ops.get("account_payable_days")})
    return result


async def build_fy_resource_trends(filters: Dict[str, Any], current_user: dict, context: Optional[Dict[str, Any]]) -> Dict[str, Dict[str, List[Dict[str, Any]]]]:
    """Aligned monthly FY/CY curves for management resource comparisons."""
    if not context:
        return {}
    start = datetime.fromisoformat(context["ytd_period"]["start_date"]).date().replace(day=1)
    prior_start = datetime.fromisoformat(context["previous_ytd_period"]["start_date"]).date().replace(day=1)
    end = datetime.fromisoformat(context["ytd_period"]["end_date"]).date().replace(day=1)
    months = []
    cursor = start
    while cursor <= end:
        months.append(cursor.strftime("%Y-%m")); cursor = (cursor + relativedelta(months=1)).replace(day=1)
    prior_months = [(prior_start + relativedelta(months=index)).strftime("%Y-%m") for index in range(len(months))]
    facilities = filters.get("facility_ids") or await organization_facility_ids(current_user)
    all_facilities = await organization_facility_ids(current_user)
    facility_filter = None if set(facilities) == set(all_facilities) else facilities
    metrics = {"energy": ("energy", "total"), "water_recycle": ("water", "recycled"), "waste_recovery": ("waste", "recovery_pct"), "renewable_energy": ("energy", "renewable_pct")}
    result = {key: {"current": [], "previous": []} for key in metrics}
    for series, period_keys in (("current", months), ("previous", prior_months)):
        for period in period_keys:
            month_filters = {**filters, "reporting_period_start": period, "reporting_period_end": period, "strict_period": True}
            snapshot = await build_resource_snapshot(current_user.get("organization_id"), facility_filter, month_filters)
            for key, (section, field) in metrics.items():
                result[key][series].append({"period": period, "value": round(float(snapshot[section].get(field, 0) or 0), 2)})
    return result


async def get_operational_kpis(organization_id: str, scope12_emissions: float, energy_total: float, incidents: int, filters: Dict[str, Any]) -> Dict[str, Any]:
    """Mirror the dashboard KPI formulas and approved-record filters exactly."""
    from modules.esg_records.services.dashboard.date_utils import build_date_filter
    from shared.utils.period_utils import period_variants

    period_conditions = build_date_filter(filters["reporting_period_start"], filters["reporting_period_end"])

    async def latest_value(collection, field_key):
        base = {"org_id": organization_id, "is_current": {"$ne": False}, "status": {"$ne": "draft"}, "approval_status": {"$in": ["approved", "not_required", None]}, f"field_values.{field_key}": {"$exists": True, "$ne": None}}
        query = {"$and": [base, {"$or": period_conditions}]} if period_conditions else base
        record = await collection.find_one(query, {"_id": 0, "field_values": 1}, sort=[("created_at", -1)])
        try:
            return float(record["field_values"][field_key]) if record else None
        except (KeyError, TypeError, ValueError):
            return None

    lost_time_injuries = await latest_value(db.social_records, "no_of_loss_time_injuries")
    total_hours_worked = await latest_value(db.social_records, "total_hours_worked")
    accounts_payable = await latest_value(db.governance_records, "accounts_payable")
    cogs = await latest_value(db.governance_records, "cost_of_goods_services_procured")
    report_year = int(filters["reporting_period_start"][:4])
    production = None
    for variant in period_variants(report_year, "FY"):
        record = await db.production_quantities.find_one({"organization_id": organization_id, "facility_id": None, "reporting_period": variant, "is_deleted": {"$ne": True}}, {"_id": 0, "quantity": 1})
        if record:
            production = record.get("quantity")
            break
    return {"ltifr": round((lost_time_injuries * 1000000) / total_hours_worked, 2) if lost_time_injuries and total_hours_worked else None, "account_payable_days": round((accounts_payable * 365) / cogs, 1) if accounts_payable and cogs else None, "incident_count": incidents, "ghg_intensity": round(scope12_emissions / production, 4) if production else None, "energy_intensity": round(energy_total / production, 4) if production else None}


async def save_report_run(filters: Dict[str, Any], summary: Dict[str, Any], current_user: dict, status: str = "generated") -> Dict[str, Any]:
    run = {
        "id": str(uuid.uuid4()), "template_id": "emissions_summary", "template_name": "Emissions Summary",
        "organization_id": current_user.get("organization_id"), "generated_by_email": current_user.get("email"),
        "generated_at": now_iso(), "status": status, "filters": filters, "summary": summary,
    }
    await db.mis_report_history.insert_one(run.copy())
    return run


def next_run_at(frequency: str, run_time: str = "09:00", run_day: Optional[int] = None) -> str:
    """Return a predictable UTC next run without changing existing schedules' defaults."""
    days = {"daily": 1, "weekly": 7, "monthly": 30, "quarterly": 90, "yearly": 365}
    hour, minute = (int(value) for value in (run_time or "09:00").split(":", 1))
    candidate = datetime.now(timezone.utc).replace(hour=hour, minute=minute, second=0, microsecond=0)
    if candidate <= datetime.now(timezone.utc):
        candidate += timedelta(days=days[frequency])
    elif frequency != "daily":
        candidate += timedelta(days=days[frequency] - 1)
    return candidate.isoformat()


def schedule_recipients(schedule: Dict[str, Any]) -> List[Dict[str, str]]:
    recipients = schedule.get("recipients") or []
    if recipients:
        return [{"id": str(item.get("id") or uuid.uuid4()), "name": item.get("name") or item.get("email", ""), "email": item.get("email", "")} for item in recipients]
    return [{"id": str(uuid.uuid4()), "name": email.split("@", 1)[0], "email": email} for email in schedule.get("recipient_emails", [])]


def safe_report_filename(name: str, extension: str) -> str:
    base = re.sub(r"[^A-Za-z0-9_-]+", "_", name).strip("_") or "ESG_MIS_Report"
    return f"{base}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.{extension}"


def archive_org_name(organization_name: str) -> str:
    """Keep organization names readable in archive paths without allowing path traversal."""
    value = re.sub(r"[\\/\x00-\x1f]+", "-", (organization_name or "Organization")).strip()
    return value.strip(".") or "Organization"


def build_report_configuration_snapshot(schedule: Dict[str, Any], resolved_filters: Dict[str, Any], facility_names: List[str]) -> Dict[str, Any]:
    content = dict(schedule.get("content") or {})
    return {
        "schedule_name": schedule.get("name"),
        "frequency": schedule.get("frequency"),
        "timezone": schedule.get("timezone"),
        "facility_selection": {
            "mode": schedule.get("facility_mode", "all"),
            "facility_ids": list(resolved_filters.get("facility_ids") or []),
            "facility_names": list(facility_names),
        },
        "sections": list(content.get("sections") or []),
        "filters": dict(resolved_filters),
        "content": content,
    }


def build_report_period_snapshot(reporting_context: Dict[str, Any]) -> Dict[str, Any]:
    reporting_period = reporting_context.get("reporting_period") or {}
    return {
        "label": reporting_period.get("label"),
        "start_date": reporting_period.get("start_date"),
        "end_date": reporting_period.get("end_date"),
        "frequency": reporting_context.get("frequency"),
    }


async def facility_names_for_filters(filters: Dict[str, Any], current_user: dict) -> List[str]:
    facility_ids = filters.get("facility_ids") or await organization_facility_ids(current_user)
    facilities = await db.facilities.find({"id": {"$in": facility_ids}}, {"_id": 0, "name": 1}).to_list(1000)
    return [facility.get("name", "Facility") for facility in facilities]


async def store_delivery_artifact(content: bytes, filename: str, content_type: str, organization_name: str, archive_date: str, delivery_id: str) -> Dict[str, Any]:
    """Persist a generated artifact once so historical downloads never regenerate a report."""
    from r2_storage import get_r2_storage

    extension = filename.rsplit(".", 1)[-1].lower()
    checksum = hashlib.sha256(content).hexdigest()
    object_key = f"{archive_org_name(organization_name)}/{archive_date}/{delivery_id}/report.{extension}"
    result = await get_r2_storage().upload_file(
        content,
        filename,
        "mis_reports",
        content_type,
        metadata={"delivery_run_id": delivery_id, "organization_name": organization_name, "sha256": checksum},
        object_key=object_key,
    )
    if not result.get("success"):
        raise RuntimeError(result.get("error", "Unable to store report artifact"))
    return {"format": extension, "filename": filename, "content_type": content_type, "file_size": result["file_size"], "storage_key": result["key"], "bucket_type": "mis_reports", "sha256": checksum}


def build_summary_email(schedule_name: str, summary: Dict[str, Any], filters: Dict[str, Any]) -> str:
    scopes = {row["scope"]: row["emissions"] for row in summary["scope_breakdown"]}
    return f"<div style='font-family:Arial,sans-serif;color:#17211d;max-width:640px'><p>Hello,</p><p>Your ESG MIS Report for <strong>{filters['reporting_period_start']} to {filters['reporting_period_end']}</strong> has been generated successfully.</p><h3>Quick Summary</h3><table style='border-collapse:collapse;width:100%'><tr><td>Total Emissions</td><td style='text-align:right'>{summary['total_emissions']:,.2f} {summary['unit']}</td></tr><tr><td>Scope 1</td><td style='text-align:right'>{scopes.get('scope1', 0):,.2f} {summary['unit']}</td></tr><tr><td>Scope 2</td><td style='text-align:right'>{scopes.get('scope2', 0):,.2f} {summary['unit']}</td></tr><tr><td>Scope 3</td><td style='text-align:right'>{scopes.get('scope3', 0):,.2f} {summary['unit']}</td></tr><tr><td>Biogenic</td><td style='text-align:right'>{scopes.get('biogenic', 0):,.2f} {summary['unit']}</td></tr></table><p>Please find the detailed PDF and Excel reports attached.</p><p>Regards,<br/>SustainRepo</p></div>"


async def send_schedule(schedule: Dict[str, Any], current_user: dict) -> Dict[str, Any]:
    requested_at = now_iso()
    delivery_id = str(uuid.uuid4())
    recipients = schedule_recipients(schedule)
    organization = await db.organizations.find_one({"id": schedule.get("organization_id")}, {"_id": 0, "name": 1, "reporting_year_type": 1, "financial_year_start_month": 1})
    organization_name = (organization or {}).get("name", "SustainRepo Organization")
    reporting_context = ReportingPeriodService(organization, datetime.now(timezone.utc)).resolve(schedule["frequency"])
    resolved_filters = ReportingPeriodService.filters_for(schedule["filters"], reporting_context["reporting_period"], schedule["frequency"])
    summary = await aggregate_emissions(resolved_filters, current_user)
    facility_names = await facility_names_for_filters(resolved_filters, current_user)
    report_configuration_snapshot = build_report_configuration_snapshot(schedule, resolved_filters, facility_names)
    report_period = build_report_period_snapshot(reporting_context)
    executive = await build_executive_mis_report(resolved_filters, current_user, reporting_context, (schedule.get("content") or {}).get("sections", []))
    pdf_name = safe_report_filename(schedule["name"], "pdf")
    xlsx_name = safe_report_filename(schedule["name"], "xlsx")
    archive_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    def delivery_run_record(status: str, generated_at: str, artifacts: List[Dict[str, Any]], failure_reason: Optional[str] = None) -> Dict[str, Any]:
        by_format = {item.get("format"): item for item in artifacts}
        pdf_artifact = by_format.get("pdf", {})
        excel_artifact = by_format.get("xlsx", {})
        return {
            "id": delivery_id,
            "delivery_run_id": delivery_id,
            "schedule_id": schedule.get("id"),
            "schedule_name": schedule["name"],
            "organization_id": schedule.get("organization_id"),
            "organization_name": organization_name,
            "status": status,
            "requested_at": requested_at,
            "generated_at": generated_at,
            "reporting_period_label": report_period.get("label"),
            "report_period": report_period,
            "report_configuration_snapshot": report_configuration_snapshot,
            "recipient_snapshot": list(recipients),
            "filters": resolved_filters,
            "reporting_context": reporting_context,
            "recipients": list(recipients),
            "content": schedule.get("content") or {},
            "facility_mode": schedule.get("facility_mode", "all"),
            "facility_names": facility_names,
            "artifacts": artifacts,
            "pdf_storage_key": pdf_artifact.get("storage_key"),
            "excel_storage_key": excel_artifact.get("storage_key"),
            "pdf_file_size": pdf_artifact.get("file_size"),
            "excel_file_size": excel_artifact.get("file_size"),
            "pdf_checksum": pdf_artifact.get("sha256"),
            "excel_checksum": excel_artifact.get("sha256"),
            "archive_bucket_type": "mis_reports" if artifacts else None,
            "failure_reason": failure_reason,
            "schedule_snapshot": dict(schedule),
        }

    try:
        pdf_content = build_executive_pdf(executive, organization_name, current_user.get("email"))
        xlsx_content = build_executive_excel(executive)
        artifacts = []
        artifacts.append(await store_delivery_artifact(pdf_content, pdf_name, "application/pdf", organization_name, archive_date, delivery_id))
        artifacts.append(await store_delivery_artifact(xlsx_content, xlsx_name, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", organization_name, archive_date, delivery_id))
    except Exception as error:
        from r2_storage import get_r2_storage
        for artifact in locals().get("artifacts", []):
            try:
                await get_r2_storage().delete_file(artifact["bucket_type"], artifact["storage_key"])
            except Exception:
                pass
        delivery_run = delivery_run_record("failed", now_iso(), [], str(error))
        await db.mis_report_delivery_runs.insert_one(delivery_run.copy())
        return delivery_run

    attachments = [(xlsx_name, xlsx_content), (pdf_name, pdf_content)]
    failures = []
    for recipient in recipients:
        success = await send_email_with_attachments(recipient["email"], f"ESG MIS Report – {reporting_context['reporting_period']['label']}", build_summary_email(schedule["name"], summary, resolved_filters), attachments)
        error = None if success else "Resend delivery failed"
        if error:
            failures.append(recipient["email"])
        delivery = {"id": str(uuid.uuid4()), "delivery_run_id": delivery_id, "schedule_id": schedule["id"], "organization_id": schedule.get("organization_id"), "recipient_email": recipient["email"], "recipient_name": recipient["name"], "status": "sent" if success else "failed", "sent_at": now_iso(), "error": error}
        await db.mis_report_deliveries.insert_one(delivery.copy())
    status = "sent" if not failures else ("failed" if len(failures) == len(recipients) else "partial")
    delivery_run = delivery_run_record(status, now_iso(), artifacts, "Resend delivery failed for: " + ", ".join(failures) if failures else None)
    delivery_run["delivered_at"] = now_iso()
    await db.mis_report_delivery_runs.insert_one(delivery_run.copy())
    return delivery_run


async def process_due_schedules() -> int:
    """Send due MIS schedules; invoked hourly by the application worker."""
    schedules = await db.mis_report_schedules.find(
        {"is_enabled": True, "next_run_at": {"$lte": now_iso()}},
        {"_id": 0},
    ).to_list(500)
    processed = 0
    for schedule in schedules:
        try:
            user_context = {"role": "admin", "organization_id": schedule.get("organization_id"), "email": schedule.get("created_by_email")}
            await send_schedule(schedule, user_context)
            await db.mis_report_schedules.update_one(
                {"id": schedule["id"]},
                {"$set": {"last_run_at": now_iso(), "next_run_at": next_run_at(schedule["frequency"], schedule.get("run_time", "09:00"), schedule.get("run_day"))}},
            )
            processed += 1
        except Exception:
            await db.mis_report_schedules.update_one(
                {"id": schedule["id"]},
                {"$set": {"next_run_at": next_run_at(schedule["frequency"], schedule.get("run_time", "09:00"), schedule.get("run_day"))}},
            )
    return processed


def build_excel(summary: Dict[str, Any]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Emissions Summary"
    sheet.append(["Emissions Summary", summary["unit"]])
    sheet.append(["Total emissions", summary["total_emissions"]])
    sheet.append(["Source records", summary["record_count"]])
    sheet.append([])
    sheet.append(["Scope", "Emissions"])
    for row in summary["scope_breakdown"]:
        sheet.append([row["scope"], row["emissions"]])
    for cell in sheet[1]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="166534")
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 20
    buffer = io.BytesIO(); workbook.save(buffer); return buffer.getvalue()


def build_pdf(summary: Dict[str, Any]) -> bytes:
    buffer = io.BytesIO(); pdf = canvas.Canvas(buffer, pagesize=A4); width, height = A4
    pdf.setFont("Helvetica-Bold", 18); pdf.drawString(48, height - 60, "Emissions Summary")
    pdf.setFont("Helvetica", 11); pdf.drawString(48, height - 88, f"Total emissions: {summary['total_emissions']:,.2f} {summary['unit']}")
    pdf.drawString(48, height - 106, f"Source records: {summary['record_count']}")
    y = height - 145; pdf.setFont("Helvetica-Bold", 11); pdf.drawString(48, y, "Scope"); pdf.drawRightString(width - 48, y, "Emissions")
    pdf.setFont("Helvetica", 10)
    for row in summary["scope_breakdown"]:
        y -= 20; pdf.drawString(48, y, row["scope"]); pdf.drawRightString(width - 48, y, f"{row['emissions']:,.2f}")
    pdf.save(); return buffer.getvalue()


def build_executive_excel(report: Dict[str, Any]) -> bytes:
    workbook = Workbook(); workbook.remove(workbook.active)
    context = report.get("reporting_context") or {}
    context_rows = []
    if context:
        context_rows = [
            ["Reporting Period", context["reporting_period"]["label"]], ["Comparison Period", context["comparison_period"]["label"]],
            ["Current FY/CY YTD", f"{context['ytd_period']['start_date']} to {context['ytd_period']['end_date']}"], ["Previous FY/CY YTD", f"{context['previous_ytd_period']['start_date']} to {context['previous_ytd_period']['end_date']}"],
            ["Reporting Calendar", context["reporting_calendar"]["label"]], [],
        ]
    def sheet(name, rows):
        ws = workbook.create_sheet(name); ws.append(["SustainRepo ESG MIS Report"]); ws.append([])
        for row in context_rows: ws.append(row)
        for row in rows: ws.append(row)
        for cell in ws[1]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="166534")
        ws.column_dimensions["A"].width = 34; ws.column_dimensions["B"].width = 22; ws.column_dimensions["C"].width = 22
    sheet("Executive Summary", [["KPI", "Current", "Previous", "Change %", "FY/CY YTD", "Previous FY/CY YTD", "YTD Change %"]] + [[k["label"], k["value"], k["previous"], k["change_pct"], k.get("ytd"), k.get("previous_ytd"), k.get("ytd_change_pct")] for k in report["kpis"]])
    sheet("Emissions Overview", [["Scope", "tCO2e"]] + [[r["scope"], r["emissions"]] for r in report["current"]["scope_breakdown"]] + [[]] + [["Category", "tCO2e"]] + [[r["category"], r["emissions"]] for r in report["current"]["category_breakdown"]])
    sheet("Facility Performance", [["Facility", "Current", "Previous", "Change %"]] + [[r["facility"], r["current"], r["previous"], r["change_pct"]] for r in report.get("facility_comparisons", [])])
    sheet("Operations", [["Energy", report["energy"].get("total"), "MWh"], ["Renewable share", report["energy"].get("renewable_pct"), "%"], ["Water recycled", report["water"].get("recycled"), "KL"], ["Waste recovered", report["waste"].get("recovered"), ""], ["LTIFR", report["operational_kpis"].get("ltifr"), ""], ["Account payable days", report["operational_kpis"].get("account_payable_days"), "days"]])
    sheet("Suppliers & Targets", [["Supplier", "Overall Score", "ESG Score", "GHG Score"]] + [[r.get("company_name") or r.get("supplier_name") or "Supplier", r.get("overall_score"), r.get("esg_score"), r.get("ghg_score")] for r in report.get("supplier_scores", [])] + [[]] + [["Target", "Target", "Actual", "Direction", "Status"]] + [[r.get("name"), r.get("target_value"), r.get("actual_value"), r.get("target_direction"), r.get("status")] for r in report.get("targets", [])])
    buffer = io.BytesIO(); workbook.save(buffer); return buffer.getvalue()


def build_executive_pdf(report: Dict[str, Any], organization_name: str, generated_by: str) -> bytes:
    from modules.mis_reports.pdf_builder import build_beautiful_executive_pdf
    return build_beautiful_executive_pdf(report, organization_name, generated_by)


# ─── Executive Summary v2: 13-Month Average Insight Engine ────────────────────


def _compute_avg_with_count(values: list) -> tuple:
    """Average of non-None values, returning (avg, count) or (None, 0)."""
    valid = [v for v in values if v is not None]
    if not valid:
        return None, 0
    return round(sum(valid) / len(valid), 4), len(valid)


def _generate_insight(current_val, avg, direction: str = "neutral", months_count: int = 13) -> dict:
    """Factual, data-driven insight comparing current month to historical average."""
    if current_val is None:
        return {"text": "No data available", "variance_pct": None, "color": "grey"}
    if avg is None:
        return {"text": "No meaningful historical average available", "variance_pct": None, "color": "grey"}
    if avg == 0:
        if current_val == 0:
            return {"text": "No activity recorded", "variance_pct": None, "color": "grey"}
        return {"text": "No meaningful historical average available", "variance_pct": None, "color": "grey"}

    # Special case: current is zero
    if current_val == 0:
        avg_fmt = f"{avg:,.2f}"
        color = "green" if direction == "decrease" else ("red" if direction == "increase" else "grey")
        return {"text": f"Decreased to 0 from a {months_count}-month average of {avg_fmt}", "variance_pct": -100.0, "color": color}

    variance = round(((current_val - avg) / avg) * 100, 1)
    avg_fmt = f"{avg:,.2f}"

    if abs(variance) < 2:
        return {"text": f"Remained broadly in line with the previous {months_count}-month average of {avg_fmt}", "variance_pct": variance, "color": "grey"}

    word = "Increased" if variance > 0 else "Decreased"
    text = f"{word} {abs(variance):,.1f}% compared with the previous {months_count}-month average of {avg_fmt}"

    if direction == "decrease":
        color = "green" if variance < 0 else "red"
    elif direction == "increase":
        color = "green" if variance > 0 else "red"
    else:
        color = "amber" if abs(variance) > 20 else "grey"

    return {"text": text, "variance_pct": variance, "color": color}


async def build_executive_summary_data(report: Dict[str, Any], filters: Dict[str, Any], current_user: dict) -> Dict[str, Any]:
    """Build structured Page 2 executive summary data with 13-month average insights.

    Uses bulk queries for emissions, reuses existing 12-month trend data for resources
    and operational metrics (supplementing with 2 additional months), and computes
    intensity metrics from production/revenue denominators.
    """
    ctx = report.get("reporting_context")
    if ctx:
        current_label = ctx["reporting_period"]["label"]
        previous_label = ctx["comparison_period"]["label"]
        current_start = datetime.fromisoformat(ctx["reporting_period"]["start_date"]).date().replace(day=1)
    else:
        # Derive from filters when no schedule context is provided
        current_start = datetime.strptime(filters["reporting_period_start"], "%Y-%m").date().replace(day=1)
        current_label = current_start.strftime("%B %Y")
        prev = current_start - relativedelta(months=1)
        previous_label = prev.strftime("%B %Y")

    # 13 completed months immediately preceding the current month (oldest first)
    months_13 = [(current_start - relativedelta(months=i)).strftime("%Y-%m") for i in range(13, 0, -1)]

    org_id = current_user.get("organization_id")
    fac_ids = filters.get("facility_ids") or await organization_facility_ids(current_user)
    all_facs = await organization_facility_ids(current_user)
    fac_filter = None if set(fac_ids) == set(all_facs) else fac_ids

    # ════════════════════════════════════════════════════════════════════════════
    # 1. EMISSIONS BY SCOPE — single bulk query across 13 months
    # ════════════════════════════════════════════════════════════════════════════
    emission_recs = await db.emission_records.find(
        {"facility_id": {"$in": fac_ids}, "scope": {"$in": ALL_SCOPES}, "reporting_period": {"$in": months_13}, **eligible_ghg_record_filter()},
        {"_id": 0, "scope": 1, "reporting_period": 1, "co2e_emissions": 1},
    ).to_list(100000)

    scope_by_month = {m: {s: 0.0 for s in ALL_SCOPES} for m in months_13}
    months_with_emission_data: set = set()
    for rec in emission_recs:
        p, s = rec.get("reporting_period", ""), rec.get("scope", "")
        if p in scope_by_month and s in ALL_SCOPES:
            try:
                scope_by_month[p][s] += float(rec.get("co2e_emissions") or 0)
                months_with_emission_data.add(p)
            except (TypeError, ValueError):
                pass

    scope_avgs: Dict[str, tuple] = {}
    for s in ALL_SCOPES:
        vals = [scope_by_month[m][s] for m in months_13 if m in months_with_emission_data]
        scope_avgs[s] = _compute_avg_with_count(vals)

    # ════════════════════════════════════════════════════════════════════════════
    # 2. RESOURCES — reuse 12-month trend data + fetch 2 extra months
    # ════════════════════════════════════════════════════════════════════════════
    resource_trends = report.get("twelve_month_resource_trends", {})
    op_trends = report.get("twelve_month_operational_trends", {})

    def _trend_map(key):
        return {e["period"]: e.get("value") for e in resource_trends.get(key, [])}

    energy_tm = _trend_map("energy")
    renew_pct_tm = _trend_map("renewable_energy")
    w_cons_tm = _trend_map("water_consumption")
    w_with_tm = _trend_map("water_withdrawal")
    w_disc_tm = _trend_map("water_discharge")
    w_rec_tm = _trend_map("water_recycle")
    waste_gen_tm = _trend_map("waste_generated")
    waste_disp_tm = _trend_map("waste_disposal")
    waste_rec_pct_tm = _trend_map("waste_recovery")

    inc_tm = {e["period"]: e.get("value") for e in op_trends.get("incidents", [])}
    ltifr_tm = {e["period"]: e.get("value") for e in op_trends.get("ltifr", [])}
    ap_tm = {e["period"]: e.get("value") for e in op_trends.get("account_payable_days", [])}

    # Determine which months are NOT covered by the 12-month trends
    extra_months = [m for m in months_13 if m not in energy_tm]
    for period in extra_months:
        mf = {**filters, "reporting_period_start": period, "reporting_period_end": period, "strict_period": True}
        snap = await build_resource_snapshot(org_id, fac_filter, mf)
        e, w, ws = snap["energy"], snap["water"], snap["waste"]
        energy_tm[period] = e.get("total")
        renew_pct_tm[period] = e.get("renewable_pct")
        w_cons_tm[period] = w.get("consumption")
        w_with_tm[period] = w.get("withdrawal")
        w_disc_tm[period] = w.get("discharge")
        w_rec_tm[period] = w.get("recycled")
        waste_gen_tm[period] = ws.get("generated")
        waste_disp_tm[period] = ws.get("disposal")
        waste_rec_pct_tm[period] = ws.get("recovery_pct")

        # Operational data for extra months
        scope12 = scope_by_month[period]["scope1"] + scope_by_month[period]["scope2"]
        inc = await db.governance_records.count_documents({
            "org_id": org_id, **_governance_period_filter(period),
            "approval_status": {"$in": ["approved", "not_required", None]},
            "$or": [
                {"subcategory": "Health & Safety Incidents"},
                {"category": {"$regex": "data breach", "$options": "i"}},
                {"subcategory": {"$regex": "data breach", "$options": "i"}},
                {"category": {"$regex": "violation", "$options": "i"}},
                {"subcategory": {"$regex": "violation", "$options": "i"}},
            ],
        })
        ops = await get_operational_kpis(org_id, scope12, e.get("total", 0) or 0, inc, mf)
        inc_tm[period] = inc
        ltifr_tm[period] = ops.get("ltifr")
        ap_tm[period] = ops.get("account_payable_days")

    # ── Build 13-month value arrays for resource averages ──
    energy_vals = [energy_tm.get(m) for m in months_13]
    renew_abs_vals, non_renew_abs_vals = [], []
    for i, m in enumerate(months_13):
        et = energy_vals[i]
        rp = renew_pct_tm.get(m)
        if et is not None and et > 0 and rp is not None:
            renew_abs_vals.append(round(et * rp / 100, 4))
            non_renew_abs_vals.append(round(et * (1 - rp / 100), 4))
        elif et is not None and et == 0:
            renew_abs_vals.append(0.0)
            non_renew_abs_vals.append(0.0)
        else:
            renew_abs_vals.append(None)
            non_renew_abs_vals.append(None)

    w_cons_vals = [w_cons_tm.get(m) for m in months_13]
    w_with_vals = [w_with_tm.get(m) for m in months_13]
    w_disc_vals = [w_disc_tm.get(m) for m in months_13]
    w_rec_vals = [w_rec_tm.get(m) for m in months_13]

    waste_gen_vals = [waste_gen_tm.get(m) for m in months_13]
    waste_disp_vals = [waste_disp_tm.get(m) for m in months_13]
    waste_recyc_vals = []
    for i in range(13):
        g = waste_gen_vals[i]
        rp = waste_rec_pct_tm.get(months_13[i])
        if g is not None and rp is not None:
            rec = g * rp / 100
            waste_recyc_vals.append(round(rec, 4))
        else:
            waste_recyc_vals.append(None)

    resource_avgs = {
        "energy_total": _compute_avg_with_count(energy_vals),
        "energy_renewable": _compute_avg_with_count(renew_abs_vals),
        "energy_non_renewable": _compute_avg_with_count(non_renew_abs_vals),
        "water_consumption": _compute_avg_with_count(w_cons_vals),
        "water_withdrawal": _compute_avg_with_count(w_with_vals),
        "water_discharge": _compute_avg_with_count(w_disc_vals),
        "water_recycle": _compute_avg_with_count(w_rec_vals),
        "waste_generated": _compute_avg_with_count(waste_gen_vals),
        "waste_disposed": _compute_avg_with_count(waste_disp_vals),
        "waste_recycled": _compute_avg_with_count(waste_recyc_vals),
    }

    inc_vals = [inc_tm.get(m) for m in months_13]
    ltifr_vals = [ltifr_tm.get(m) for m in months_13]
    ap_vals = [ap_tm.get(m) for m in months_13]
    op_avgs = {
        "incidents": _compute_avg_with_count(inc_vals),
        "ltifr": _compute_avg_with_count(ltifr_vals),
        "ap_days": _compute_avg_with_count(ap_vals),
    }

    # ════════════════════════════════════════════════════════════════════════════
    # 3. INTENSITY DENOMINATORS — production & revenue
    # ════════════════════════════════════════════════════════════════════════════
    from shared.utils.period_utils import period_variants
    report_year = int(filters["reporting_period_start"][:4])
    production, prod_unit = None, "MT"
    for variant in period_variants(report_year, "FY"):
        rec = await db.production_quantities.find_one(
            {"organization_id": org_id, "facility_id": None, "reporting_period": variant, "is_deleted": {"$ne": True}},
            {"_id": 0, "quantity": 1, "unit": 1},
        )
        if rec and rec.get("quantity"):
            try:
                production = float(rec["quantity"])
            except (TypeError, ValueError):
                pass
            prod_unit = rec.get("unit", "MT")
            break

    revenue, rev_currency = None, "INR"
    fin = await db.organization_financials.find_one({"org_id": org_id}, {"_id": 0, "turnover": 1, "currency": 1})
    if fin:
        try:
            revenue = float(fin.get("turnover") or 0) or None
        except (TypeError, ValueError):
            pass
        rev_currency = fin.get("currency", "INR")

    # Current / previous scope 1+2 for intensity
    current_scopes = {r["scope"]: r["emissions"] for r in report["current"]["scope_breakdown"]}
    previous_scopes = {r["scope"]: r["emissions"] for r in report["previous"]["scope_breakdown"]}
    c_scope12 = current_scopes.get("scope1", 0) + current_scopes.get("scope2", 0)
    p_scope12 = previous_scopes.get("scope1", 0) + previous_scopes.get("scope2", 0)

    ghg_int_prod = round(c_scope12 / production, 6) if production else None
    ghg_int_prod_prev = round(p_scope12 / production, 6) if production else None
    ghg_int_rev = round(c_scope12 / revenue, 6) if revenue else None
    ghg_int_rev_prev = round(p_scope12 / revenue, 6) if revenue else None

    e_total = report.get("energy", {}).get("total", 0) or 0
    pe_total = report.get("previous_resources", {}).get("energy", {}).get("total", 0) or 0
    energy_int_prod = round(e_total / production, 6) if production else None
    energy_int_prod_prev = round(pe_total / production, 6) if production else None
    energy_int_rev = round(e_total / revenue, 6) if revenue else None
    energy_int_rev_prev = round(pe_total / revenue, 6) if revenue else None

    # 13-month intensity averages
    scope12_monthly = [scope_by_month[m]["scope1"] + scope_by_month[m]["scope2"] for m in months_13]
    ghg_int_prod_avg = _compute_avg_with_count([v / production for v in scope12_monthly]) if production else (None, 0)
    ghg_int_rev_avg = _compute_avg_with_count([v / revenue for v in scope12_monthly]) if revenue else (None, 0)
    energy_int_prod_avg = _compute_avg_with_count([(v or 0) / production for v in energy_vals if v is not None]) if production else (None, 0)
    energy_int_rev_avg = _compute_avg_with_count([(v or 0) / revenue for v in energy_vals if v is not None]) if revenue else (None, 0)

    # ════════════════════════════════════════════════════════════════════════════
    # 4. INCIDENT BREAKDOWN (current period)
    # ════════════════════════════════════════════════════════════════════════════
    period_start = filters["reporting_period_start"]
    period_end = filters["reporting_period_end"]
    incident_base = {"org_id": org_id, **_governance_period_filter(period_start), "approval_status": {"$in": ["approved", "not_required", None]}}
    safety_count = await db.governance_records.count_documents({**incident_base, "subcategory": "Health & Safety Incidents"})
    breach_count = await db.governance_records.count_documents({**incident_base, "$or": [{"category": {"$regex": "data breach", "$options": "i"}}, {"subcategory": {"$regex": "data breach", "$options": "i"}}]})
    violation_count = await db.governance_records.count_documents({**incident_base, "$or": [{"category": {"$regex": "violation", "$options": "i"}}, {"subcategory": {"$regex": "violation", "$options": "i"}}]})

    # ════════════════════════════════════════════════════════════════════════════
    # 5. PREVIOUS MONTH OPERATIONAL VALUES (from trend maps)
    # ════════════════════════════════════════════════════════════════════════════
    prev_month = (current_start - relativedelta(months=1)).strftime("%Y-%m")
    prev_ltifr = ltifr_tm.get(prev_month)
    prev_ap = ap_tm.get(prev_month)
    prev_incidents = inc_tm.get(prev_month, 0)

    # ════════════════════════════════════════════════════════════════════════════
    # 6. BUILD STRUCTURED SECTIONS
    # ════════════════════════════════════════════════════════════════════════════
    selected = set(report.get("selected_sections") or [])
    include_all = not selected

    cr = report.get("energy", {})
    pr = report.get("previous_resources", {})
    water_data = report.get("water", {})
    waste_data = report.get("waste", {})
    ops = report.get("operational_kpis", {})

    def _m(name: str, current, previous, unit: str, avg_tuple: tuple, direction: str) -> dict:
        avg_val, count = avg_tuple if isinstance(avg_tuple, tuple) else (avg_tuple, 13)
        ins = _generate_insight(current, avg_val, direction, count if count else 13)
        return {"name": name, "current": current, "previous": previous, "unit": unit, **ins}

    sections = []

    # ── GHG Emissions ──
    if include_all or "ghg" in selected:
        sections.append({
            "key": "ghg", "title": "GHG Emissions", "color": "#0e7490",
            "metrics": [
                _m("Scope 1", current_scopes.get("scope1", 0), previous_scopes.get("scope1", 0), "tCO2e", scope_avgs.get("scope1", (None, 0)), "decrease"),
                _m("Scope 2", current_scopes.get("scope2", 0), previous_scopes.get("scope2", 0), "tCO2e", scope_avgs.get("scope2", (None, 0)), "decrease"),
                _m("Scope 3", current_scopes.get("scope3", 0), previous_scopes.get("scope3", 0), "tCO2e", scope_avgs.get("scope3", (None, 0)), "decrease"),
                _m("Biogenic", current_scopes.get("biogenic", 0), previous_scopes.get("biogenic", 0), "tCO2e", scope_avgs.get("biogenic", (None, 0)), "decrease"),
                _m("Intensity by Production", ghg_int_prod, ghg_int_prod_prev, f"tCO2e/{prod_unit}", ghg_int_prod_avg, "decrease"),
                _m("Intensity by Revenue", ghg_int_rev, ghg_int_rev_prev, f"tCO2e/{rev_currency}", ghg_int_rev_avg, "decrease"),
            ],
        })

    # ── Energy Consumption ──
    if include_all or "energy" in selected or "ghg" in selected:
        sections.append({
            "key": "energy", "title": "Energy Consumption", "color": "#d97706",
            "metrics": [
                _m("Renewable Energy", cr.get("renewable_total", 0), pr.get("energy", {}).get("renewable_total", 0), "MWh", resource_avgs["energy_renewable"], "increase"),
                _m("Non-Renewable Energy", cr.get("non_renewable_total", 0), pr.get("energy", {}).get("non_renewable_total", 0), "MWh", resource_avgs["energy_non_renewable"], "decrease"),
                _m("Intensity by Production", energy_int_prod, energy_int_prod_prev, f"MWh/{prod_unit}", energy_int_prod_avg, "decrease"),
                _m("Intensity by Revenue", energy_int_rev, energy_int_rev_prev, f"MWh/{rev_currency}", energy_int_rev_avg, "decrease"),
            ],
        })

    # ── Water ──
    if include_all or "water" in selected:
        sections.append({
            "key": "water", "title": "Water", "color": "#0284c7",
            "metrics": [
                _m("Consumption", water_data.get("consumption", 0), pr.get("water", {}).get("consumption", 0), "KL", resource_avgs["water_consumption"], "decrease"),
                _m("Withdrawal", water_data.get("withdrawal", 0), pr.get("water", {}).get("withdrawal", 0), "KL", resource_avgs["water_withdrawal"], "decrease"),
                _m("Discharge", water_data.get("discharge", 0), pr.get("water", {}).get("discharge", 0), "KL", resource_avgs["water_discharge"], "neutral"),
                _m("Recycle", water_data.get("recycled", 0), pr.get("water", {}).get("recycled", 0), "KL", resource_avgs["water_recycle"], "increase"),
            ],
        })

    # ── Waste ──
    if include_all or "waste" in selected:
        wg = waste_data.get("generated", 0) or 0
        wr = waste_data.get("recovered", 0) or 0
        wd = max(wg - wr, 0)
        pwg = pr.get("waste", {}).get("generated", 0) or 0
        pwr = pr.get("waste", {}).get("recovered", 0) or 0
        pwd = max(pwg - pwr, 0)
        sections.append({
            "key": "waste", "title": "Waste", "color": "#7e22ce",
            "metrics": [
                _m("Generated", wg, pwg, "kg", resource_avgs["waste_generated"], "decrease"),
                _m("Disposed", wd, pwd, "kg", resource_avgs["waste_disposed"], "decrease"),
                _m("Recycled", wr, pwr, "kg", resource_avgs["waste_recycled"], "increase"),
            ],
        })

    # ── Social & Governance ──
    show_social = include_all or "social" in selected or "governance" in selected
    if show_social:
        total_incidents = inc_tm.get(period_start, 0)
        sections.append({
            "key": "social_governance", "title": "Social & Governance", "color": "#312e81",
            "metrics": [
                _m("LTIFR", ops.get("ltifr"), prev_ltifr, "", op_avgs["ltifr"], "decrease"),
                _m("Account Payable Days", ops.get("account_payable_days"), prev_ap, "days", op_avgs["ap_days"], "neutral"),
                _m("Number of Incidents", total_incidents, prev_incidents, "count", op_avgs["incidents"], "decrease"),
            ],
            "incident_breakdown": {
                "total": total_incidents,
                "safety_incidents": safety_count,
                "data_breaches": breach_count,
                "violations": violation_count,
            },
        })

    return {
        "current_month_label": current_label,
        "previous_month_label": previous_label,
        "sections": sections,
    }



SCOPE3_CANONICAL = [
    "C1 - Purchased Goods and Services",
    "C2 - Capital Goods",
    "C3 - Fuel and Energy Related Activities Not Included in Scope 1 or Scope 2",
    "C4 - Upstream Transportation and Distribution",
    "C5 - Waste Generated in Operations",
    "C6 - Business Travel",
    "C7 - Employee Commuting",
    "C8 - Upstream Leased Assets",
    "C9 - Downstream Transportation and Distribution",
    "C10 - Processing of Sold Products",
    "C11 - Use of Sold Products",
    "C12 - End-of-Life Treatment of Sold Products",
    "C13 - Downstream Leased Assets",
    "C14 - Franchises",
    "C15 - Investments",
]


def _period_to_months(period_str: str, fy_start_month: int = 4) -> List[str]:
    """Parse FY/CY/range period strings into a list of YYYY-MM months they cover."""
    import re as _re
    if not period_str:
        return []
    # CY format: "CY2025", "CY 2025"
    cy = _re.match(r"CY\s*(\d{4})", period_str, _re.I)
    if cy:
        y = int(cy.group(1))
        return [f"{y}-{m:02d}" for m in range(1, 13)]
    # Range: "2025-01 to 2025-12", "2025-04 to 2025-03"
    rng = _re.match(r"(\d{4}-\d{2})\s*to\s*(\d{4}-\d{2})", period_str, _re.I)
    if rng:
        s = datetime.strptime(rng.group(1), "%Y-%m").date()
        e = datetime.strptime(rng.group(2), "%Y-%m").date()
        if e < s:
            e = e.replace(year=e.year + 1)
        out, cur = [], s
        while cur <= e:
            out.append(cur.strftime("%Y-%m"))
            cur += relativedelta(months=1)
        return out
    # FY format: "FY 2025-2026", "FY 2025-26", "2025-2026", "2025-26"
    fy = _re.match(r"(?:FY\s*)?(\d{4})-(\d{2,4})$", period_str, _re.I)
    if fy:
        start_year = int(fy.group(1))
        out = []
        for i in range(12):
            m = fy_start_month + i
            y = start_year
            if m > 12:
                m -= 12
                y += 1
            out.append(f"{y}-{m:02d}")
        return out
    return []


async def build_emissions_deep_data(report: Dict[str, Any], filters: Dict[str, Any], current_user: dict) -> Dict[str, Any]:
    """Fetch granular per-scope, per-category emissions for 12-13 months via a single aggregation."""
    ctx = report.get("reporting_context")
    if ctx:
        current_start = datetime.fromisoformat(ctx["reporting_period"]["start_date"]).date().replace(day=1)
    else:
        current_start = datetime.strptime(filters["reporting_period_start"], "%Y-%m").date()

    current_month = current_start.strftime("%Y-%m")
    months = [(current_start - relativedelta(months=i)).strftime("%Y-%m") for i in range(12, -1, -1)]
    months_set = set(months)
    fac_ids = filters.get("facility_ids") or await organization_facility_ids(current_user)
    org_id = current_user.get("organization_id")

    # ── 1. Monthly records (exact YYYY-MM match) ──
    pipeline = [
        {"$match": {"facility_id": {"$in": fac_ids}, "scope": {"$in": ALL_SCOPES}, "reporting_period": {"$in": months}, **eligible_ghg_record_filter()}},
        {"$addFields": {"_ev": {"$toDouble": {"$ifNull": ["$co2e_emissions", 0]}}}},
        {"$group": {"_id": {"p": "$reporting_period", "s": "$scope", "c": {"$ifNull": ["$category", "Uncategorized"]}}, "v": {"$sum": "$_ev"}}},
    ]
    raw = await db.emission_records.aggregate(pipeline).to_list(100000)

    tree: Dict[str, Dict[str, Dict[str, float]]] = {}
    for r in raw:
        p, s, c = r["_id"]["p"], r["_id"]["s"], r["_id"]["c"]
        tree.setdefault(p, {}).setdefault(s, {})[c] = round(r["v"], 4)

    # ── 2. FY / CY / range annual records — distribute across months ──
    org = await db.organizations.find_one({"id": org_id}, {"_id": 0, "financial_year_start_month": 1})
    fy_month = int((org or {}).get("financial_year_start_month") or 4)

    from shared.utils.period_utils import period_variants as _pv
    annual_candidates: List[str] = []
    for y in {int(m[:4]) for m in months}:
        annual_candidates.extend(_pv(y, "FY"))
        annual_candidates.extend(_pv(y - 1, "FY"))
        annual_candidates.extend([f"CY{y}", f"CY {y}"])

    # Also catch range-style periods
    annual_recs = await db.emission_records.find(
        {"facility_id": {"$in": fac_ids}, "scope": {"$in": ALL_SCOPES},
         "$or": [{"reporting_period": {"$in": annual_candidates}},
                 {"reporting_period": {"$regex": r"\d{4}-\d{2}\s*to\s*\d{4}-\d{2}"}}],
         **eligible_ghg_record_filter()},
        {"_id": 0, "scope": 1, "category": 1, "co2e_emissions": 1, "reporting_period": 1},
    ).to_list(10000)

    for rec in annual_recs:
        covered = _period_to_months(rec.get("reporting_period", ""), fy_month)
        if not covered:
            continue
        try:
            val = float(rec.get("co2e_emissions") or 0)
        except (TypeError, ValueError):
            continue
        if val == 0:
            continue
        share = round(val / len(covered), 6)
        scope = rec.get("scope", "")
        cat = rec.get("category") or "Uncategorized"
        for cm in covered:
            if cm in months_set:
                tree.setdefault(cm, {}).setdefault(scope, {})[cat] = round(tree.get(cm, {}).get(scope, {}).get(cat, 0) + share, 4)

    # ── 3. Build output ──
    def _block(sk: str, canonical_cats: Optional[List[str]] = None) -> dict:
        cats: set = set()
        for m in months:
            cats.update(tree.get(m, {}).get(sk, {}).keys())
        if canonical_cats:
            cats.update(canonical_cats)
        trend = []
        for m in months:
            sc = tree.get(m, {}).get(sk)
            trend.append({"period": m, "value": round(sum(sc.values()), 2) if sc else None})
        cc = tree.get(current_month, {}).get(sk, {})
        ct = sum(cc.values()) if cc else 0
        comp = sorted(
            [{"category": c, "value": round(cc.get(c, 0), 2), "pct": round(cc.get(c, 0) / ct * 100, 1) if ct else 0} for c in cats],
            key=lambda x: -x["value"],
        )
        ct_trends = {}
        for c in sorted(cats):
            ct_trends[c] = [{"period": m, "value": round(tree.get(m, {}).get(sk, {}).get(c, 0), 2) if tree.get(m, {}).get(sk) else None} for m in months]
        return {"trend": trend, "current_value": round(ct, 2), "composition": comp, "category_trends": ct_trends}

    total_trend = []
    for m in months:
        md = tree.get(m)
        total_trend.append({"period": m, "value": round(sum(sum(cs.values()) for cs in md.values()), 2) if md else None})

    cd = tree.get(current_month, {})
    tc = sum(sum(cs.values()) for cs in cd.values())
    scope_comp = [
        {"scope": s, "label": s.replace("scope", "Scope ").title(), "value": round(sum(cd.get(s, {}).values()), 2), "pct": round(sum(cd.get(s, {}).values()) / tc * 100, 1) if tc else 0}
        for s in ALL_SCOPES
    ]

    return {
        "months": months, "current_month": current_month,
        "current_month_label": current_start.strftime("%B %Y"),
        "total": {"trend": total_trend, "current_value": round(tc, 2), "composition": scope_comp},
        "scope1": _block("scope1"),
        "scope2": _block("scope2"),
        "scope3": _block("scope3", canonical_cats=SCOPE3_CANONICAL),
        "biogenic": _block("biogenic"),
    }



WATER_SOURCE_FIELDS = {
    "water_withdrawal_through_ground_water": "Groundwater",
    "water_withdrawal_through_surface_water": "Surface Water",
    "water_withdrawal_through_third_party_water": "Third-Party Water",
    "water_withdrawal_through_seawater_desalinated_water": "Seawater / Desalinated",
}


def _record_to_month(rp: dict) -> str | None:
    """Extract YYYY-MM from an environment_record reporting_period dict."""
    if not rp:
        return None
    if rp.get("month"):
        y = rp.get("year")
        m = rp.get("month")
        if not y:
            return None
        try:
            month_num = int(m)
        except (TypeError, ValueError):
            try:
                month_num = datetime.strptime(str(m), "%B").month
            except (TypeError, ValueError):
                return None
        return f"{y}-{month_num:02d}"
    if rp.get("date"):
        return str(rp["date"])[:7]
    return None


async def _build_resources_deep(report: dict, current_user: dict) -> dict:
    """Build structured energy/water/waste deep data from existing trend + current-month snapshots."""
    ctx = report.get("reporting_context")
    rt = report.get("twelve_month_resource_trends", {})
    if not rt:
        return {}

    def _tmap(key):
        return {e["period"]: e.get("value") for e in rt.get(key, [])}

    months = [e["period"] for e in rt.get("energy", [])]
    if not months:
        return {}

    current_month = months[-1]
    if ctx:
        cm_label = ctx["reporting_period"]["label"]
    else:
        cm_label = datetime.strptime(current_month, "%Y-%m").strftime("%B %Y")

    # Energy
    energy_total_m = _tmap("energy")
    renew_total_m = _tmap("energy_renewable_total")
    nonrenew_total_m = _tmap("energy_non_renewable_total")
    cr = report.get("energy", {})

    energy = {
        "total_trend": [{"period": m, "value": energy_total_m.get(m)} for m in months],
        "renewable_trend": [{"period": m, "value": renew_total_m.get(m)} for m in months],
        "non_renewable_trend": [{"period": m, "value": nonrenew_total_m.get(m)} for m in months],
        "current": cr,
        "composition": [
            {"category": "Renewable Energy", "value": cr.get("renewable_total", 0) or 0,
             "pct": cr.get("renewable_pct", 0) or 0},
            {"category": "Non-Renewable Energy",
             "value": cr.get("non_renewable_total", 0) or 0,
             "pct": round(100 - (cr.get("renewable_pct", 0) or 0), 1)},
        ],
    }

    # Water
    w_cons_m = _tmap("water_consumption")
    w_with_m = _tmap("water_withdrawal")
    w_disc_m = _tmap("water_discharge")
    w_rec_m = _tmap("water_recycle")
    cw = report.get("water", {})

    water = {
        "consumption_trend": [{"period": m, "value": w_cons_m.get(m)} for m in months],
        "withdrawal_trend": [{"period": m, "value": w_with_m.get(m)} for m in months],
        "discharge_trend": [{"period": m, "value": w_disc_m.get(m)} for m in months],
        "recycle_trend": [{"period": m, "value": w_rec_m.get(m)} for m in months],
        "current": cw,
    }

    # Waste
    gen_m = _tmap("waste_generated")
    rec_m = _tmap("waste_recovered")
    disp_m = _tmap("waste_disposal")
    haz_gen_m = _tmap("waste_haz_generated")
    haz_rec_m = _tmap("waste_haz_recovered")
    haz_disp_m = _tmap("waste_haz_disposed")
    nhaz_gen_m = _tmap("waste_nonhaz_generated")
    nhaz_rec_m = _tmap("waste_nonhaz_recovered")
    nhaz_disp_m = _tmap("waste_nonhaz_disposed")
    wst = report.get("waste", {})

    waste = {
        "generated_trend": [{"period": m, "value": gen_m.get(m)} for m in months],
        "recovered_trend": [{"period": m, "value": rec_m.get(m)} for m in months],
        "disposed_trend": [{"period": m, "value": disp_m.get(m)} for m in months],
        "haz_generated_trend": [{"period": m, "value": haz_gen_m.get(m)} for m in months],
        "haz_recovered_trend": [{"period": m, "value": haz_rec_m.get(m)} for m in months],
        "haz_disposed_trend": [{"period": m, "value": haz_disp_m.get(m)} for m in months],
        "nonhaz_generated_trend": [{"period": m, "value": nhaz_gen_m.get(m)} for m in months],
        "nonhaz_recovered_trend": [{"period": m, "value": nhaz_rec_m.get(m)} for m in months],
        "nonhaz_disposed_trend": [{"period": m, "value": nhaz_disp_m.get(m)} for m in months],
        "current": {
            "generated": wst.get("generated", 0) or 0,
            "recovered": wst.get("recovered", 0) or 0,
            "disposed": wst.get("disposal", 0) or 0,
            "haz_generated": wst.get("hazardous_generated", 0) or 0,
            "haz_recovered": wst.get("hazardous_recovered", 0) or 0,
            "haz_disposed": wst.get("hazardous_disposed", 0) or 0,
            "nonhaz_generated": wst.get("non_hazardous_generated", 0) or 0,
            "nonhaz_recovered": wst.get("non_hazardous_recovered", 0) or 0,
            "nonhaz_disposed": wst.get("non_hazardous_disposed", 0) or 0,
        },
    }

    # ── Water withdrawal sources (single bulk query) ──
    org_id = current_user.get("organization_id")
    years_in_window = list({int(m[:4]) for m in months})
    source_recs = await db.environment_records.find(
        {"org_id": org_id, "category": "Water", "subcategory": "Withdrawal",
         "approval_status": {"$in": ["approved", "not_required", None]},
         "reporting_period.year": {"$in": years_in_window}},
        {"_id": 0, "field_values": 1, "reporting_period": 1},
    ).to_list(10000)

    source_by_month: Dict[str, Dict[str, float]] = {m: {} for m in months}
    for rec in source_recs:
        rm = _record_to_month(rec.get("reporting_period"))
        if rm not in source_by_month:
            continue
        fv = rec.get("field_values") or {}
        for field_key, label in WATER_SOURCE_FIELDS.items():
            val = to_kilolitres(fv.get(field_key), fv.get("unit"))
            if val:
                source_by_month[rm][label] = round(source_by_month[rm].get(label, 0) + val, 2)

    all_sources = sorted({s for mv in source_by_month.values() for s in mv})
    cur_sources = source_by_month.get(current_month, {})
    cur_total_src = sum(cur_sources.values()) or 0
    water["source_composition"] = [
        {"category": s, "value": round(cur_sources.get(s, 0), 2),
         "pct": round(cur_sources.get(s, 0) / cur_total_src * 100, 1) if cur_total_src else 0}
        for s in (all_sources or list(WATER_SOURCE_FIELDS.values()))
    ]
    water["source_trends"] = {
        s: [{"period": m, "value": source_by_month[m].get(s)} for m in months]
        for s in (all_sources or list(WATER_SOURCE_FIELDS.values()))
    }

    return {
        "months": months, "current_month": current_month,
        "current_month_label": cm_label,
        "energy": energy, "water": water, "waste": waste,
    }



async def build_facility_deep_data(report: Dict[str, Any], filters: Dict[str, Any], current_user: dict) -> Dict[str, Any]:
    """Per-facility emissions breakdown for ALL org facilities, including zeros."""
    ctx = report.get("reporting_context")
    if ctx:
        current_start = datetime.fromisoformat(ctx["reporting_period"]["start_date"]).date().replace(day=1)
    else:
        current_start = datetime.strptime(filters["reporting_period_start"], "%Y-%m").date()

    current_month = current_start.strftime("%Y-%m")
    prev_month = (current_start - relativedelta(months=1)).strftime("%Y-%m")
    months = [(current_start - relativedelta(months=i)).strftime("%Y-%m") for i in range(12, -1, -1)]

    org_id = current_user.get("organization_id")

    # ── All org facilities (master list — never truncated) ──
    fac_docs = await db.facilities.find(
        {"organization_id": org_id, "is_deleted": {"$ne": True}},
        {"_id": 0, "id": 1, "name": 1},
    ).to_list(500)
    fac_map = {f["id"]: f.get("name", f["id"][:8]) for f in fac_docs}
    all_fac_ids = list(fac_map.keys())

    if not all_fac_ids:
        return {}

    # ── Single aggregation: (period, scope, category, facility) ──
    pipeline = [
        {"$match": {"facility_id": {"$in": all_fac_ids}, "scope": {"$in": ALL_SCOPES}, "reporting_period": {"$in": months}, **eligible_ghg_record_filter()}},
        {"$addFields": {"_ev": {"$toDouble": {"$ifNull": ["$co2e_emissions", 0]}}}},
        {"$group": {"_id": {"p": "$reporting_period", "s": "$scope", "c": {"$ifNull": ["$category", "Uncategorized"]}, "f": "$facility_id"}, "v": {"$sum": "$_ev"}}},
    ]
    raw = await db.emission_records.aggregate(pipeline).to_list(200000)

    # tree: {facility: {period: {scope: {category: value}}}}
    tree: Dict[str, Dict[str, Dict[str, Dict[str, float]]]] = {}
    for r in raw:
        f, p, s, c = r["_id"]["f"], r["_id"]["p"], r["_id"]["s"], r["_id"]["c"]
        tree.setdefault(f, {}).setdefault(p, {}).setdefault(s, {})[c] = round(r["v"], 4)

    # ── Also handle FY/CY annual records ──
    from shared.utils.period_utils import period_variants as _pv
    org_doc = await db.organizations.find_one({"id": org_id}, {"_id": 0, "financial_year_start_month": 1})
    fy_month = int((org_doc or {}).get("financial_year_start_month") or 4)
    annual_candidates: List[str] = []
    for y in {int(m[:4]) for m in months}:
        annual_candidates.extend(_pv(y, "FY"))
        annual_candidates.extend(_pv(y - 1, "FY"))
        annual_candidates.extend([f"CY{y}", f"CY {y}"])
    annual_recs = await db.emission_records.find(
        {"facility_id": {"$in": all_fac_ids}, "scope": {"$in": ALL_SCOPES},
         "$or": [{"reporting_period": {"$in": annual_candidates}},
                 {"reporting_period": {"$regex": r"\d{4}-\d{2}\s*to\s*\d{4}-\d{2}"}}],
         **eligible_ghg_record_filter()},
        {"_id": 0, "facility_id": 1, "scope": 1, "category": 1, "co2e_emissions": 1, "reporting_period": 1},
    ).to_list(50000)
    months_set = set(months)
    for rec in annual_recs:
        covered = _period_to_months(rec.get("reporting_period", ""), fy_month)
        if not covered:
            continue
        try:
            val = float(rec.get("co2e_emissions") or 0)
        except (TypeError, ValueError):
            continue
        if val == 0:
            continue
        share = round(val / len(covered), 6)
        f_id = rec.get("facility_id", "")
        s = rec.get("scope", "")
        c = rec.get("category") or "Uncategorized"
        for cm in covered:
            if cm in months_set:
                tree.setdefault(f_id, {}).setdefault(cm, {}).setdefault(s, {})[c] = round(tree.get(f_id, {}).get(cm, {}).get(s, {}).get(c, 0) + share, 4)

    # ── Build per-facility blocks ──
    facilities = []
    for fac_id in all_fac_ids:
        fac_name = fac_map[fac_id]
        ft = tree.get(fac_id, {})

        # Monthly trend
        trend = []
        for m in months:
            md = ft.get(m)
            trend.append({"period": m, "value": round(sum(sum(cs.values()) for cs in md.values()), 2) if md else 0})

        # Current / previous month totals
        def _month_total(m):
            md = ft.get(m, {})
            return round(sum(sum(cs.values()) for cs in md.values()), 2) if md else 0

        cur_total = _month_total(current_month)
        prev_total = _month_total(prev_month)
        if prev_total:
            change_pct = round((cur_total - prev_total) / prev_total * 100, 1)
        elif cur_total == 0:
            change_pct = 0
        else:
            change_pct = None  # N/A

        # Scope breakdown (current month)
        cur_data = ft.get(current_month, {})
        scope_breakdown = []
        for s in ALL_SCOPES:
            sv = round(sum(cur_data.get(s, {}).values()), 2)
            pct = round(sv / cur_total * 100, 1) if cur_total else 0
            scope_breakdown.append({"scope": s, "label": s.replace("scope", "Scope ").title(), "value": sv, "pct": pct})

        # Category breakdowns per scope
        def _scope_cats(sk, canonical=None):
            cats_in_data = set()
            for m in months:
                cats_in_data.update(ft.get(m, {}).get(sk, {}).keys())
            if canonical:
                cats_in_data.update(canonical)
            cc = cur_data.get(sk, {})
            total_scope = sum(cc.values()) if cc else 0
            return sorted(
                [{"category": c, "value": round(cc.get(c, 0), 2), "pct": round(cc.get(c, 0) / total_scope * 100, 1) if total_scope else 0} for c in cats_in_data],
                key=lambda x: -x["value"],
            )

        facilities.append({
            "id": fac_id, "name": fac_name,
            "current_total": cur_total, "previous_total": prev_total, "change_pct": change_pct,
            "scope_breakdown": scope_breakdown,
            "scope1_categories": _scope_cats("scope1"),
            "scope2_categories": _scope_cats("scope2"),
            "scope3_categories": _scope_cats("scope3", canonical=SCOPE3_CANONICAL),
            "biogenic_categories": _scope_cats("biogenic"),
            "monthly_trend": trend,
        })

    # Sort by current_total descending (zero-emission facilities at bottom)
    facilities.sort(key=lambda f: -f["current_total"])

    return {
        "facilities": facilities,
        "months": months,
        "current_month": current_month,
        "current_month_label": current_start.strftime("%B %Y"),
        "previous_month_label": (current_start - relativedelta(months=1)).strftime("%B %Y"),
    }



# ─── Monthly Target History Builder ──────────────────────────────────────────

# Maps target category/kpi to the correct trend key in emissions_deep or resources_deep
_CATEGORY_TREND_MAP = {
    "ghg": ("emissions_deep", "total", "trend"),
    "scope 1": ("emissions_deep", "scope1", "trend"),
    "scope 2": ("emissions_deep", "scope2", "trend"),
    "scope 3": ("emissions_deep", "scope3", "trend"),
    "energy": ("resources_deep", "energy", "total_trend"),
    "water": ("resources_deep", "water", "consumption_trend"),
    "waste": ("resources_deep", "waste", "generated_trend"),
}


def _resolve_actual_trend(report: dict, category: str) -> Dict[str, float]:
    """Given a target category, find the matching monthly actual trend and return {period: value}."""
    cat_lower = (category or "").lower().strip()
    for key, (deep_key, section, trend_key) in _CATEGORY_TREND_MAP.items():
        if key in cat_lower:
            deep = report.get(deep_key, {})
            sec = deep.get(section, {})
            trend = sec.get(trend_key, [])
            return {d["period"]: d.get("value") for d in trend if d.get("value") is not None}
    return {}


def _attach_monthly_target_history(report: dict) -> None:
    """For each monthly target, build a monthly_history array pairing target values with actuals."""
    targets = report.get("targets", [])
    for t in targets:
        if t.get("tracking_mode") != "monthly":
            continue
        tv_map = t.get("tracking_values") or {}
        if not tv_map:
            continue
        actual_map = _resolve_actual_trend(report, t.get("category", ""))
        # Build history for months where a target is configured, within the 12-month window
        months_in_window = set()
        for deep_key in ("emissions_deep", "resources_deep"):
            for m in report.get(deep_key, {}).get("months", []):
                months_in_window.add(m)
        history = []
        for period in sorted(tv_map.keys()):
            if months_in_window and period not in months_in_window:
                continue
            history.append({
                "period": period,
                "target": tv_map[period],
                "actual": actual_map.get(period),
            })
        t["monthly_history"] = history
