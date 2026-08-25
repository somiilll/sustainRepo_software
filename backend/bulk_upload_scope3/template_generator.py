"""
Excel Template Generator for Scope 3 Bulk Upload
Creates a professionally formatted Excel workbook with all 15 category sheets
"""
import io
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.worksheet.protection import SheetProtection

from .models import CATEGORY_COLUMNS, ACTIVITY_TYPES, CalculationMethod
from .ghg_config_resolver import ResolvedGhgCapabilities
import re


def sanitize_for_excel(value):
    """Remove illegal characters that cannot be used in Excel worksheets"""
    if value is None:
        return ""
    if not isinstance(value, str):
        value = str(value)
    # Remove illegal XML characters (control characters except tab, newline, carriage return)
    # Excel/openpyxl cannot handle these
    illegal_chars_pattern = re.compile(
        r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]'
    )
    return illegal_chars_pattern.sub('', value)


# Style definitions
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
MANDATORY_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
OPTIONAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
SUPPLIER_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


class TemplateGenerator:
    """Generates Excel template for Scope 3 bulk upload"""
    
    def __init__(self, facilities: List[Dict], activities_by_category: Dict[str, List[Dict]], 
                 units_by_category: Dict[str, List[str]], organization_name: str = "",
                 capabilities: Optional[ResolvedGhgCapabilities] = None):
        """
        Initialize template generator with dynamic data
        
        Args:
            facilities: List of facility dicts with 'id' and 'name'
            activities_by_category: Dict mapping category code to list of activities
            units_by_category: Dict mapping category code to list of allowed units
            organization_name: Organization name for template header
            capabilities: Resolved org GHG capabilities (controls which sheets/categories appear)
        """
        self.facilities = facilities
        self.activities_by_category = activities_by_category
        self.units_by_category = units_by_category
        self.organization_name = organization_name
        self.capabilities = capabilities or ResolvedGhgCapabilities()
        self.workbook = Workbook()
        
    def generate(self) -> io.BytesIO:
        """Generate the complete template workbook"""
        # Remove default sheet
        if "Sheet" in self.workbook.sheetnames:
            del self.workbook["Sheet"]
        
        # Create helper sheets first (hidden)
        self._create_helper_sheets()
        
        # Create Scope 1 and Scope 2 sheets only if the org has access
        if self.capabilities.scope1_enabled:
            self._create_scope12_sheet("Scope1")
        if self.capabilities.scope2_enabled:
            self._create_scope12_sheet("Scope2")
        
        # Create category sheets (Scope 3) — only enabled categories
        if self.capabilities.scope3_enabled:
            for category_code in ["C1", "C2", "C3", "C4", "C5", "C6", "C7", "C8", "C9", "C10", "C11", "C12", "C13", "C14", "C15"]:
                if self.capabilities.is_scope3_sheet_enabled(category_code):
                    self._create_category_sheet(category_code)
        
        # Create instructions sheet
        self._create_instructions_sheet()
        
        # Move instructions to first position
        self.workbook.move_sheet("Instructions", offset=-len(self.workbook.sheetnames)+1)
        
        # Save to BytesIO
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output
    
    def _create_helper_sheets(self):
        """Create hidden helper sheets for dropdowns"""
        # Facilities helper sheet
        ws_facilities = self.workbook.create_sheet("_Facilities")
        ws_facilities.append(["Facility Name"])
        for facility in self.facilities:
            ws_facilities.append([facility.get("name", "")])
        ws_facilities.sheet_state = 'hidden'
        
        # Methods helper sheet
        ws_methods = self.workbook.create_sheet("_Methods")
        ws_methods.append(["Method Key", "Method Display"])
        ws_methods.append(["activity_basis", "Activity Based"])
        ws_methods.append(["spend_basis", "Spend Based"])
        ws_methods.append(["supplier_basis", "Supplier Based"])
        ws_methods.sheet_state = 'hidden'
        
        # Activity types for C6 and C7
        ws_activity_types = self.workbook.create_sheet("_ActivityTypes")
        ws_activity_types.append(["Category", "Key", "Display Name"])
        for cat, types in ACTIVITY_TYPES.items():
            for at in types:
                ws_activity_types.append([cat, at["key"], at["name"]])
        ws_activity_types.sheet_state = 'hidden'
        
        # Create Reference Data sheet with activities grouped by method for each category
        self._create_reference_data_sheet()
        
        # Activities by category and method (keep for backward compatibility)
        ws_activities = self.workbook.create_sheet("_Activities")
        ws_activities.append(["Category", "Method", "Activity Name", "Activity ID", "Activity Type"])
        for cat_code, activities in self.activities_by_category.items():
            for act in activities:
                ws_activities.append([
                    sanitize_for_excel(cat_code),
                    sanitize_for_excel(act.get("method", "")),
                    sanitize_for_excel(act.get("activity", act.get("name", ""))),
                    sanitize_for_excel(act.get("id", "")),
                    sanitize_for_excel(act.get("activity_type", ""))
                ])
        ws_activities.sheet_state = 'hidden'
        
        # Units by category
        ws_units = self.workbook.create_sheet("_Units")
        ws_units.append(["Category", "Unit"])
        for cat_code, units in self.units_by_category.items():
            for unit in units:
                ws_units.append([sanitize_for_excel(cat_code), sanitize_for_excel(unit)])
        ws_units.sheet_state = 'hidden'
    
    def _create_reference_data_sheet(self):
        """Create Reference Data sheet with activities grouped by method for INDIRECT dropdowns"""
        ws = self.workbook.create_sheet("_RefData")
        
        # Track column positions for each category's methods
        col_offset = 1
        
        # Categories that use the method-based approach
        method_categories = ["C1", "C2", "C3", "C4", "C5", "C6", "C7", "C8", "C9", "C10", "C11", "C12", "C13", "C14"]
        
        for category_code in method_categories:
            activities = self.activities_by_category.get(category_code, [])
            if not activities:
                continue
            
            # Group activities by method
            activity_by_method = {
                "activity_basis": [],
                "spend_basis": [],
            }
            
            for act in activities:
                method = act.get("method", "activity_basis")
                activity_name = sanitize_for_excel(act.get("activity", act.get("fuel_name", act.get("name", ""))))
                if activity_name and method in activity_by_method:
                    if activity_name not in activity_by_method[method]:
                        activity_by_method[method].append(activity_name)
            
            # supplier_basis uses combined list of activity_basis + spend_basis
            activity_by_method["supplier_basis"] = list(set(
                activity_by_method["activity_basis"] + activity_by_method["spend_basis"]
            ))
            
            # Write activities to columns and create named ranges
            for method_key in ["activity_basis", "spend_basis", "supplier_basis"]:
                method_activities = sorted(activity_by_method.get(method_key, []))
                if not method_activities:
                    continue
                
                # Header row
                col_letter = get_column_letter(col_offset)
                ws.cell(row=1, column=col_offset, value=f"{category_code}_{method_key}")
                
                # Write activities starting from row 2
                for row_idx, activity_name in enumerate(method_activities, start=2):
                    ws.cell(row=row_idx, column=col_offset, value=activity_name)
                
                # Create named range for this category-method combination
                # Named range format: C1_activity_basis, C1_spend_basis, C1_supplier_basis, etc.
                range_name = f"{category_code}_{method_key}"
                end_row = len(method_activities) + 1  # +1 because we start from row 2
                range_ref = f"'_RefData'!${col_letter}$2:${col_letter}${end_row}"
                
                # Add to workbook's defined names
                from openpyxl.workbook.defined_name import DefinedName
                defn = DefinedName(range_name, attr_text=range_ref)
                self.workbook.defined_names.add(defn)
                
                col_offset += 1
        
        ws.sheet_state = 'hidden'
    
    def _create_scope12_sheet(self, scope_code: str):
        """Create a Scope 1 or Scope 2 sheet with all formatting and validations"""
        config = CATEGORY_COLUMNS.get(scope_code)
        if not config:
            return
        
        ws = self.workbook.create_sheet(config["sheet_name"])
        columns = config["columns"]
        
        # Set column widths and create headers
        for col_idx, col_config in enumerate(columns, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(20, len(col_config["name"]) + 5)
            
            # Header cell
            header_cell = ws.cell(row=1, column=col_idx, value=col_config["name"])
            header_cell.fill = HEADER_FILL
            header_cell.font = HEADER_FONT
            header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            header_cell.border = THIN_BORDER
            
            # Add comment with instructions
            comment_text = self._get_scope12_column_comment(col_config, scope_code)
            if comment_text:
                header_cell.comment = Comment(comment_text, "SustainRepo")
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Set row height for header
        ws.row_dimensions[1].height = 30
        
        # Add data validations for Scope 1/2
        self._add_scope12_data_validations(ws, columns, scope_code, config)
        
        # Format data rows (first 1000 rows)
        for row in range(2, 1002):
            for col_idx, col_config in enumerate(columns, start=1):
                cell = ws.cell(row=row, column=col_idx)
                cell.border = THIN_BORDER
                cell.protection = Protection(locked=False)
                
                # Color code based on field type
                if col_config["mandatory"]:
                    cell.fill = MANDATORY_FILL
                else:
                    cell.fill = OPTIONAL_FILL
        
        # Protect sheet structure but allow cell editing
        ws.protection = SheetProtection(
            sheet=True,
            objects=True,
            scenarios=True,
            formatCells=False,
            formatColumns=False,
            formatRows=False,
            insertColumns=False,
            insertRows=True,
            insertHyperlinks=False,
            deleteColumns=False,
            deleteRows=True,
            selectLockedCells=False,
            selectUnlockedCells=False,
            password=None
        )
    
    def _get_scope12_column_comment(self, col_config: Dict, scope_code: str) -> str:
        """Generate comment text for Scope 1/2 column headers"""
        key = col_config["key"]
        mandatory = "(Mandatory) " if col_config.get("mandatory") else "(Optional) "
        
        comments = {
            "facility_name": f"{mandatory}Select from dropdown list of facilities",
            "reporting_month": f"{mandatory}Format: MMM-YYYY (e.g., Jan-2025). Fill either this OR Reporting Year, not both.",
            "reporting_year": "(Optional) Format: FY YYYY-YYYY or CY YYYY. Fill either this OR Reporting Month, not both.",
            "category": f"{mandatory}Select from dropdown",
            "process_type": "(Optional) Required for Process Emissions. Select: Venting, N2O from Overall Combustion, CH4 from Overall Combustion",
            "fuel_gas": f"{mandatory}Select fuel/gas from dropdown. Must match exact fuel name from database.",
            "energy_used": f"{mandatory}Select energy type from dropdown. Must match exact name from database.",
            "qty": f"{mandatory}Numeric value for quantity consumed",
            "qty_energy": f"{mandatory}Numeric value for energy quantity consumed",
            "unit_qty": f"{mandatory}Unit of measurement for quantity",
            "co2_gwp_fugitives": "(Optional) Global Warming Potential for fugitive emissions. Only applicable for Fugitive Emissions category.",
            "cv": "(Optional) Override calorific value. If provided, Unit of Calorific Value is required.",
            "cv_unit": "(Conditional) Required if Calorific Value is provided",
            "density": "(Optional) Override density value. If provided, Unit of Density is required.",
            "density_unit": "(Conditional) Required if Density is provided",
            "ef_quantity": "(Optional) Override emission factor in kgCO2/kg",
            "ef_quantity_electricity_co2": "(Optional) Override emission factor for electricity/energy",
            "ef_unit": "(Conditional) Required if Emission Factor is provided",
            "process_name": "(Optional) Name of the process or activity",
            "process_description": "(Optional) Description of the process",
            "record_source": "(Optional) Source of information for this data entry",
            "responsible_person": "(Optional) Name of person responsible for this data",
            "responsible_designation": "(Optional) Designation of responsible person",
            "responsible_contact": "(Optional) Contact details of responsible person",
            "notes": "(Optional) Additional notes or comments",
        }
        
        return comments.get(key, "")
    
    def _add_scope12_data_validations(self, ws, columns: List[Dict], scope_code: str, config: Dict):
        """Add data validations for Scope 1/2 sheets"""
        from .models import SCOPE1_CATEGORIES, SCOPE2_CATEGORIES
        
        for col_idx, col_config in enumerate(columns, start=1):
            col_letter = get_column_letter(col_idx)
            key = col_config["key"]
            col_type = col_config.get("type", "text")
            
            # Facility dropdown
            if key == "facility_name" and self.facilities:
                facility_count = len(self.facilities)
                dv = DataValidation(
                    type="list",
                    formula1=f"'_Facilities'!$A$2:$A${facility_count + 1}",
                    allow_blank=not col_config["mandatory"]
                )
                dv.error = "Please select a valid facility"
                dv.errorTitle = "Invalid Facility"
                ws.add_data_validation(dv)
                dv.add(f"{col_letter}2:{col_letter}1001")
            
            # Category dropdown — filtered by org capabilities
            elif key == "category":
                if scope_code == "Scope1":
                    from .models import SCOPE1_CATEGORIES
                    enabled_keys = self.capabilities.enabled_scope1_categories()
                    categories = [
                        c["name"] for c in SCOPE1_CATEGORIES
                        if c["key"] in enabled_keys
                    ]
                else:
                    categories = [c["name"] for c in SCOPE2_CATEGORIES]
                dv = DataValidation(
                    type="list",
                    formula1='"' + ','.join(categories) + '"',
                    allow_blank=not col_config["mandatory"]
                )
                dv.error = "Please select a valid category"
                dv.errorTitle = "Invalid Category"
                ws.add_data_validation(dv)
                dv.add(f"{col_letter}2:{col_letter}1001")
            
            # Process Type dropdown (Scope 1 only)
            elif key == "process_type" and scope_code == "Scope1":
                # Build list from standard process types, filtered by org config
                all_types = [
                    ("venting", "Venting"),
                    ("n2o_overall_combustion", "N2O from Overall Combustion"),
                    ("ch4_overall_combustion", "CH4 from Overall Combustion"),
                ]
                allowed = [
                    label for key_val, label in all_types
                    if self.capabilities.is_process_type_allowed(key_val)
                ]
                if allowed:
                    dv = DataValidation(
                        type="list",
                        formula1='"' + ','.join(allowed) + '"',
                        allow_blank=True,
                    )
                    dv.error = "Please select a valid process type"
                    dv.errorTitle = "Invalid Process Type"
                    ws.add_data_validation(dv)
                    dv.add(f"{col_letter}2:{col_letter}1001")
            
            # Fuel/Energy dropdown - use activities from fuel_database
            elif key in ["fuel_gas", "energy_used"]:
                # Get fuels/energy from activities_by_category (should be passed with scope12 fuels)
                fuel_list_key = "Scope1_fuels" if scope_code == "Scope1" else "Scope2_fuels"
                fuels = self.activities_by_category.get(fuel_list_key, [])
                
                if fuels:
                    # Create a named range for fuels
                    fuel_names = sorted(set([
                        sanitize_for_excel(f.get("fuel_name", f.get("name", "")))
                        for f in fuels if f.get("fuel_name") or f.get("name")
                    ]))
                    
                    if fuel_names:
                        # Add to helper sheet if not already there
                        ws_ref = "_Scope12Fuels"
                        if ws_ref not in self.workbook.sheetnames:
                            ws_fuels = self.workbook.create_sheet(ws_ref)
                            ws_fuels.sheet_state = 'hidden'
                        else:
                            ws_fuels = self.workbook[ws_ref]
                        
                        # Find the next available column
                        col_offset = 1
                        while ws_fuels.cell(row=1, column=col_offset).value:
                            col_offset += 1
                        
                        # Write header and fuel names
                        fuel_col_letter = get_column_letter(col_offset)
                        ws_fuels.cell(row=1, column=col_offset, value=f"{scope_code}_{key}")
                        for row_idx, fuel_name in enumerate(fuel_names, start=2):
                            ws_fuels.cell(row=row_idx, column=col_offset, value=fuel_name)
                        
                        # Create data validation using the list
                        end_row = len(fuel_names) + 1
                        dv = DataValidation(
                            type="list",
                            formula1=f"'{ws_ref}'!${fuel_col_letter}$2:${fuel_col_letter}${end_row}",
                            allow_blank=not col_config["mandatory"]
                        )
                        dv.error = f"Please select a valid {'fuel/gas' if key == 'fuel_gas' else 'energy type'}"
                        dv.errorTitle = "Invalid Selection"
                        ws.add_data_validation(dv)
                        dv.add(f"{col_letter}2:{col_letter}1001")
            
            # Unit dropdown
            elif key == "unit_qty":
                units = self.units_by_category.get(scope_code, [])
                if units:
                    dv = DataValidation(
                        type="list",
                        formula1='"' + ','.join(units[:250]) + '"',  # Excel limit
                        allow_blank=not col_config["mandatory"]
                    )
                    dv.error = "Please select a valid unit"
                    dv.errorTitle = "Invalid Unit"
                    ws.add_data_validation(dv)
                    dv.add(f"{col_letter}2:{col_letter}1001")
            
            # Number validation
            elif col_type == "number":
                dv = DataValidation(
                    type="decimal",
                    operator="greaterThanOrEqual",
                    formula1="0",
                    allow_blank=True
                )
                dv.error = "Please enter a valid positive number"
                dv.errorTitle = "Invalid Number"
                ws.add_data_validation(dv)
                dv.add(f"{col_letter}2:{col_letter}1001")
    
    def _create_category_sheet(self, category_code: str):
        """Create a single category sheet with all formatting and validations"""
        config = CATEGORY_COLUMNS.get(category_code)
        if not config:
            return
        
        ws = self.workbook.create_sheet(config["sheet_name"])
        columns = config["columns"]
        
        # Set column widths and create headers
        for col_idx, col_config in enumerate(columns, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(20, len(col_config["name"]) + 5)
            
            # Header cell
            header_cell = ws.cell(row=1, column=col_idx, value=col_config["name"])
            header_cell.fill = HEADER_FILL
            header_cell.font = HEADER_FONT
            header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            header_cell.border = THIN_BORDER
            
            # Add comment with instructions
            comment_text = self._get_column_comment(col_config, category_code)
            if comment_text:
                header_cell.comment = Comment(comment_text, "SustainRepo")
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Set row height for header
        ws.row_dimensions[1].height = 30
        
        # Add data validations
        self._add_data_validations(ws, columns, category_code, config)
        
        # Format data rows (first 1000 rows)
        for row in range(2, 1002):
            for col_idx, col_config in enumerate(columns, start=1):
                cell = ws.cell(row=row, column=col_idx)
                cell.border = THIN_BORDER
                cell.protection = Protection(locked=False)
                
                # Color code based on field type
                if col_config["mandatory"]:
                    cell.fill = MANDATORY_FILL
                elif "supplier" in col_config["key"].lower():
                    cell.fill = SUPPLIER_FILL
                else:
                    cell.fill = OPTIONAL_FILL
        
        # Protect sheet structure but allow cell editing
        ws.protection = SheetProtection(
            sheet=True,
            objects=True,
            scenarios=True,
            formatCells=False,
            formatColumns=False,
            formatRows=False,
            insertColumns=False,
            insertRows=True,
            insertHyperlinks=False,
            deleteColumns=False,
            deleteRows=True,
            selectLockedCells=False,
            sort=True,
            autoFilter=True,
            pivotTables=False,
            selectUnlockedCells=False
        )
    
    def _get_column_comment(self, col_config: Dict, category_code: str) -> str:
        """Generate helpful comment for a column"""
        key = col_config["key"]
        comments = {
            "facility_name": "Select the facility name from the dropdown.\nThis must match exactly with your registered facilities.",
            "reporting_month": "Enter in format: Jan-2025, Feb-2025, etc.\nFormat: MMM-YYYY",
            "calculation_method": "Select calculation method:\n- Activity Based: Use standard emission factors\n- Spend Based: Calculate from monetary spend\n- Supplier Based: Use supplier-provided emission factors",
            "activity": "Select the activity from dropdown.\nDropdown shows activities available for selected method.",
            "activity_type": "Select the type of activity (e.g., Air Travel, Car Travel).\nThis filters the available activities.",
            "quantity_used": "Enter the quantity value.\nRequired for Activity Based method.",
            "unit_quantity": "Select the unit of measurement.\nMust be compatible with selected activity.",
            "spent_amount": "Enter amount spent in INR.\nRequired for Spend Based method.",
            "distance_travelled": "Enter distance in kilometers.\nRequired for transportation activities.",
            "quantity_goods": "Enter quantity of goods transported.\nRequired for freight activities.",
            "unit_goods": "Select unit for goods quantity (t, kg, g).",
            "supplier_quantity": "Enter supplier-provided quantity.\nRequired for Supplier Based method.",
            "supplier_unit": "Enter unit of supplier quantity.\nRequired for Supplier Based method.",
            "supplier_ef": "Enter supplier emission factor value.\nRequired for Supplier Based method.",
            "supplier_ef_unit": "Enter emission factor unit (e.g., kgCO2e/unit).\nRequired for Supplier Based method.",
            "supplier_name": "Optional: Enter supplier name for traceability.",
            "supplier_code": "Optional: Enter supplier code/ID.",
            "employee_name": "Enter employee name.\nRequired for C7 Employee Commuting.",
            "employee_id": "Optional: Enter employee ID.",
            "department": "Optional: Enter employee department.",
            "sub_category": "Select sub-category from dropdown.\nFilters available activities.",
            "passengers": "Enter number of passengers.",
            "rooms": "Enter number of hotel rooms.",
            "nights": "Enter number of nights stayed.",
            "working_days": "Enter number of working days.",
            "working_hours": "Enter working hours per day.",
            "responsible_person": "Optional: Name of person responsible for this data.",
            "responsible_designation": "Optional: Designation of responsible person.",
            "responsible_contact": "Optional: Contact details of responsible person.",
            "inflation_rate": "Optional: Inflation rate for spend adjustment.",
            "purchase_power_value": "Optional: Purchase power parity value.",
            "asset_name": "Name or identifier of the leased asset, franchise, or investment.\nRequired for C8, C13, C14, C15.",
            "process_name": "Optional: Name of the process or activity generating emissions.",
            "process_description": "Optional: Description of the process or activity.",
        }
        return comments.get(key, "")
    
    def _add_data_validations(self, ws, columns: List[Dict], category_code: str, config: Dict):
        """Add data validations to the worksheet"""
        # Find column indices
        col_indices = {col["key"]: idx + 1 for idx, col in enumerate(columns)}
        
        # Facility dropdown
        if "facility_name" in col_indices:
            facility_count = len(self.facilities) + 1
            dv_facility = DataValidation(
                type="list",
                formula1=f"'_Facilities'!$A$2:$A${facility_count}",
                allow_blank=False,
                showErrorMessage=True,
                showDropDown=False,  # False = show the dropdown arrow (openpyxl quirk)
                errorTitle="Invalid Facility",
                error="Please select a facility from the dropdown list."
            )
            col_letter = get_column_letter(col_indices["facility_name"])
            dv_facility.add(f"{col_letter}2:{col_letter}1001")
            ws.add_data_validation(dv_facility)
        
        # Calculation method dropdown - use INTERNAL KEYS for INDIRECT to work
        if "calculation_method" in col_indices:
            methods = config["supported_methods"]
            # Use internal keys (activity_basis, spend_basis, supplier_basis) for INDIRECT lookup
            method_list = ",".join([m.value for m in methods])
            dv_method = DataValidation(
                type="list",
                formula1=f'"{method_list}"',
                allow_blank=False,
                showErrorMessage=True,
                showDropDown=False,  # False = show the dropdown arrow (openpyxl quirk)
                errorTitle="Invalid Method",
                error=f"Please select a valid calculation method: {method_list}"
            )
            col_letter = get_column_letter(col_indices["calculation_method"])
            dv_method.add(f"{col_letter}2:{col_letter}1001")
            ws.add_data_validation(dv_method)
        
        # Activity type dropdown for C6 and C7 with DISPLAY LABELS
        if config.get("has_activity_type") and "activity_type" in col_indices:
            activity_types = ACTIVITY_TYPES.get(category_code, [])
            # Use display names instead of keys
            at_list = ",".join([at["name"] for at in activity_types])
            dv_at = DataValidation(
                type="list",
                formula1=f'"{at_list}"',
                allow_blank=False,
                showErrorMessage=True,
                showDropDown=False,  # False = show the dropdown arrow (openpyxl quirk)
                errorTitle="Invalid Activity Type",
                error="Please select a valid activity type"
            )
            col_letter = get_column_letter(col_indices["activity_type"])
            dv_at.add(f"{col_letter}2:{col_letter}1001")
            ws.add_data_validation(dv_at)
        
        # Sub-category dropdown for categories with subcategories
        if config.get("has_subcategory") and "sub_category" in col_indices:
            # Get unique subcategories from activities
            subcategories = set()
            for act in self.activities_by_category.get(category_code, []):
                if act.get("sub_category"):
                    subcategories.add(act["sub_category"])
            if subcategories:
                subcat_list = ",".join(sorted(subcategories))
                dv_subcat = DataValidation(
                    type="list",
                    formula1=f'"{subcat_list}"',
                    allow_blank=False,
                    showErrorMessage=True,
                    showDropDown=False,  # False = show the dropdown arrow (openpyxl quirk)
                    errorTitle="Invalid Sub-Category",
                    error="Please select a valid sub-category."
                )
                col_letter = get_column_letter(col_indices["sub_category"])
                dv_subcat.add(f"{col_letter}2:{col_letter}1001")
                ws.add_data_validation(dv_subcat)
        
        # Activity dropdown - use INDIRECT with helper column for stable dependent dropdowns
        if "activity" in col_indices and category_code != "C15":
            method_col_letter = get_column_letter(col_indices.get("calculation_method", 1))
            activity_col_letter = get_column_letter(col_indices["activity"])
            
            # Use a hidden helper column (column Z) to build the named range reference
            # This is more stable than concatenation inside INDIRECT for Data Validation
            helper_col = 26  # Column Z
            helper_col_letter = get_column_letter(helper_col)
            
            # Add header for helper column (hidden)
            ws.cell(row=1, column=helper_col, value="_RangeHelper")
            
            # Add helper formula to each row: ="C1_"&$C2 (category + method selection)
            # This creates the named range reference string
            for row in range(2, 502):  # Limit to 500 data rows for performance
                ws.cell(row=row, column=helper_col, value=f'="{category_code}_"&${method_col_letter}{row}')
            
            # Hide the helper column
            ws.column_dimensions[helper_col_letter].hidden = True
            
            # Create row-wise validations for proper INDIRECT evaluation
            # Each row gets its own validation that references its helper cell
            for row in range(2, 502):  # Match helper column range
                dv_activity = DataValidation(
                    type="list",
                    formula1=f'=INDIRECT(${helper_col_letter}{row})',
                    allow_blank=False,
                    showErrorMessage=True,
                    showDropDown=False,  # False = show the dropdown arrow (openpyxl quirk)
                    errorTitle="Invalid Activity",
                    error="Please select a valid activity from the dropdown."
                )
                dv_activity.add(f"{activity_col_letter}{row}")
                ws.add_data_validation(dv_activity)
        
        # Unit dropdowns
        if "unit_quantity" in col_indices:
            units = self.units_by_category.get(category_code, ["t", "kg", "g"])
            unit_list = ",".join(units[:20])  # Limit units
            dv_unit = DataValidation(
                type="list",
                formula1=f'"{unit_list}"',
                allow_blank=True,
                showErrorMessage=True,
                showDropDown=False,  # False = show the dropdown arrow (openpyxl quirk)
                errorTitle="Invalid Unit",
                error=f"Please select a valid unit: {unit_list}"
            )
            col_letter = get_column_letter(col_indices["unit_quantity"])
            dv_unit.add(f"{col_letter}2:{col_letter}1001")
            ws.add_data_validation(dv_unit)
        
        if "unit_goods" in col_indices:
            dv_unit_goods = DataValidation(
                type="list",
                formula1='"t,kg,g"',
                allow_blank=True,
                showDropDown=False  # False = show the dropdown arrow (openpyxl quirk)
            )
            col_letter = get_column_letter(col_indices["unit_goods"])
            dv_unit_goods.add(f"{col_letter}2:{col_letter}1001")
            ws.add_data_validation(dv_unit_goods)
        
        # Numeric validations
        numeric_columns = ["quantity_used", "spent_amount", "distance_travelled", "quantity_goods",
                          "supplier_quantity", "supplier_ef", "passengers", "rooms", "nights",
                          "working_days", "working_hours", "inflation_rate", "purchase_power_value"]
        
        for num_col in numeric_columns:
            if num_col in col_indices:
                dv_num = DataValidation(
                    type="decimal",
                    operator="greaterThanOrEqual",
                    formula1="0",
                    allow_blank=True,
                    showErrorMessage=True,
                    errorTitle="Invalid Number",
                    error="Please enter a valid positive number."
                )
                col_letter = get_column_letter(col_indices[num_col])
                dv_num.add(f"{col_letter}2:{col_letter}1001")
                ws.add_data_validation(dv_num)
    
    def _create_instructions_sheet(self):
        """Create instructions sheet with guidance"""
        ws = self.workbook.create_sheet("Instructions")
        
        # Title
        ws.merge_cells('A1:G1')
        title_cell = ws.cell(row=1, column=1, value="Scope 3 Bulk Upload Template - Instructions")
        title_cell.font = Font(bold=True, size=16, color="1F4E79")
        title_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 30
        
        # Organization name
        ws.merge_cells('A2:G2')
        org_cell = ws.cell(row=2, column=1, value=f"Organization: {self.organization_name}")
        org_cell.font = Font(size=12, italic=True)
        
        instructions = [
            "",
            "GENERAL INSTRUCTIONS:",
            "1. Each sheet corresponds to a Scope 3 category (C1-C15)",
            "2. Do NOT rename, delete, or rearrange sheets",
            "3. Yellow cells are MANDATORY fields",
            "4. Green cells are OPTIONAL fields",
            "5. Blue cells are for SUPPLIER-BASED calculations only",
            "",
            "CALCULATION METHODS:",
            "• activity_basis - Use standard emission factors from database",
            "• spend_basis - Calculate emissions from monetary spend (INR)",
            "• supplier_basis - Use supplier-provided emission factors",
            "",
            "REPORTING MONTH FORMAT:",
            "• Use format: Jan-2025, Feb-2025, Mar-2025, etc.",
            "",
            "COLOR CODING:",
            "• Yellow (FFF2CC) - Mandatory field",
            "• Green (E2EFDA) - Optional field",
            "• Blue (DDEBF7) - Supplier-based fields",
            "",
            "ACTIVITY SELECTION:",
            "• For activity_basis and spend_basis: Select from dropdown",
            "• For supplier_basis: You can enter custom activity names",
            "",
            "C7 - EMPLOYEE COMMUTING:",
            "• Enter one row per employee per month per activity",
            "• System will aggregate emissions by activity and month",
            "• Employee Name is required for C7",
            "",
            "C15 - INVESTMENTS:",
            "• Only supplier_basis method is supported",
            "• Custom activities are allowed",
            "",
            "VALIDATION:",
            "• System will validate all fields during upload",
            "• Invalid rows will be flagged with specific error messages",
            "• You can choose to save valid rows while fixing errors",
            "",
            "TIPS:",
            "• Use the dropdown menus where available",
            "• Check the column headers for tooltips",
            "• Ensure numeric fields contain only numbers",
            "• Remove any extra spaces from text fields",
        ]
        
        for idx, text in enumerate(instructions, start=3):
            cell = ws.cell(row=idx, column=1, value=text)
            if text.endswith(":"):
                cell.font = Font(bold=True, size=11)
            elif text.startswith("•"):
                cell.font = Font(size=10)
            ws.column_dimensions['A'].width = 80
        
        # Protect instructions sheet
        ws.protection = SheetProtection(sheet=True, objects=True, scenarios=True)


async def generate_scope3_template(db, organization_id: str,
                                   capabilities: Optional[ResolvedGhgCapabilities] = None) -> io.BytesIO:
    """
    Generate Scope 3 bulk upload template with dynamic data from database
    
    Args:
        db: Database connection
        organization_id: Organization ID to fetch facilities for
        capabilities: Pre-resolved org GHG capabilities (optional; resolved here if None)
        
    Returns:
        BytesIO containing the Excel workbook
    """
    if capabilities is None:
        from .ghg_config_resolver import resolve_ghg_capabilities as _resolve
        capabilities = await _resolve(db, organization_id)
    
    # Fetch facilities for the organization
    facilities = await db.facilities.find(
        {"organization_id": organization_id, "is_active": {"$ne": False}},
        {"_id": 0, "id": 1, "name": 1}
    ).to_list(1000)
    
    # Fetch organization name
    org = await db.organizations.find_one(
        {"id": organization_id},
        {"_id": 0, "name": 1}
    )
    org_name = org.get("name", "") if org else ""
    
    # Fetch activities from scope3_ef collection grouped by category and method
    activities_by_category = {}
    scope3_ef_data = await db.scope3_ef.find(
        {},
        {"_id": 0, "id": 1, "category": 1, "method": 1, "activity": 1, 
         "activity_type": 1, "sub_category": 1, "allowed_units": 1, "default_unit": 1}
    ).to_list(10000)
    
    for ef in scope3_ef_data:
        # Extract category code (C1, C2, etc.)
        category_name = ef.get("category", "")
        category_code = category_name.split(" - ")[0].strip() if " - " in category_name else category_name[:3]
        
        if category_code not in activities_by_category:
            activities_by_category[category_code] = []
        
        activities_by_category[category_code].append({
            "id": ef.get("id"),
            "activity": ef.get("activity"),
            "method": ef.get("method"),
            "activity_type": ef.get("activity_type"),
            "sub_category": ef.get("sub_category"),
            "allowed_units": ef.get("allowed_units", []),
            "default_unit": ef.get("default_unit")
        })
    
    # Get unique units by category
    units_by_category = {}
    for cat_code, activities in activities_by_category.items():
        units = set()
        for act in activities:
            if act.get("allowed_units"):
                units.update(act["allowed_units"])
            if act.get("default_unit"):
                units.add(act["default_unit"])
        # Add common units if none found
        if not units:
            units = {"t", "kg", "g", "L", "kWh"}
        units_by_category[cat_code] = sorted(list(units))
    
    # Fetch fuels for Scope 1 and Scope 2 from fuel_database
    fuel_database_data = await db.fuel_database.find(
        {},
        {"_id": 0, "id": 1, "fuel_name": 1, "fuel_type": 1, "subcategory": 1, "allowed_units": 1, "default_unit": 1}
    ).to_list(10000)
    
    # Separate fuels for Scope 1 and Scope 2
    scope1_fuels = []
    scope2_fuels = []
    
    for fuel in fuel_database_data:
        fuel_type = (fuel.get("fuel_type") or "").lower()
        subcategory = (fuel.get("subcategory") or "").lower()
        
        # Scope 2 fuels: electricity, heat, steam
        if 'electricity' in fuel_type or 'heat' in fuel_type or 'steam' in fuel_type or \
           'electricity' in subcategory or 'energy' in subcategory:
            scope2_fuels.append(fuel)
        else:
            # Everything else is Scope 1 fuel
            scope1_fuels.append(fuel)
    
    activities_by_category["Scope1_fuels"] = scope1_fuels
    activities_by_category["Scope2_fuels"] = scope2_fuels
    
    # Add units for Scope 1/2
    scope1_units = set()
    scope2_units = set()
    
    for fuel in scope1_fuels:
        if fuel.get("allowed_units"):
            scope1_units.update(fuel["allowed_units"])
        if fuel.get("default_unit"):
            scope1_units.add(fuel["default_unit"])
    
    for fuel in scope2_fuels:
        if fuel.get("allowed_units"):
            scope2_units.update(fuel["allowed_units"])
        if fuel.get("default_unit"):
            scope2_units.add(fuel["default_unit"])
    
    # Add common units
    if not scope1_units:
        scope1_units = {"t", "kg", "g", "L", "kL", "m3"}
    if not scope2_units:
        scope2_units = {"kWh", "MWh", "GWh", "GJ", "MJ"}
    
    units_by_category["Scope1"] = sorted(list(scope1_units))
    units_by_category["Scope2"] = sorted(list(scope2_units))
    
    # Generate template
    generator = TemplateGenerator(
        facilities=facilities,
        activities_by_category=activities_by_category,
        units_by_category=units_by_category,
        organization_name=org_name,
        capabilities=capabilities,
    )
    
    return generator.generate()
