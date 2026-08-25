"""
Upload Processor for Scope 3 Bulk Upload
Main orchestrator for processing uploaded Excel files
"""
import io
import uuid
import logging
from typing import Dict, List, Optional, Tuple, Any
from datetime import datetime, timezone
from openpyxl import load_workbook

from ..models import (
    ValidationError, ErrorSeverity, RowResult, UploadSummary, UploadStatus,
    BulkUploadJob, CATEGORY_COLUMNS
)
from ..ghg_config_resolver import ResolvedGhgCapabilities, resolve_ghg_capabilities
from .row_processor import RowProcessor
from .scope12_processor import Scope12RowProcessor
from modules.entitlements.dependencies import (
    assert_period_row_batch_limit,
    partition_records_by_period_row_limit,
)

logger = logging.getLogger(__name__)

# ── Upload limits ────────────────────────────────────────────────────────
MAX_ROWS_PER_SHEET = 5000
MAX_TOTAL_ROWS = 25000


# Sheets to ignore (instruction sheets, reference sheets, etc.)
IGNORED_SHEET_PATTERNS = [
    "instruction", "instructions", "reference", "ref", "help", "guide",
    "dropdown", "dropdowns", "lookup", "lookups", "list", "lists",
    "readme", "read me", "info", "about"
]


class UploadProcessor:
    """Main processor for bulk upload files"""
    
    def __init__(self, db, organization_id: str, user_id: str,
                 user_email: str = "", user_name: str = "",
                 capabilities: Optional[ResolvedGhgCapabilities] = None):
        self.db = db
        self.organization_id = organization_id
        self.user_id = user_id
        self.user_email = user_email
        self.user_name = user_name
        self.capabilities = capabilities
        self.row_processor = RowProcessor(db, organization_id, user_id, user_email, user_name)
        self.scope12_processor = Scope12RowProcessor(
            db, organization_id, user_id, user_email, user_name,
            capabilities=capabilities,
        )
    
    def _should_skip_sheet(self, sheet_name: str) -> bool:
        """Check if a sheet should be skipped (instruction/reference sheets)"""
        sheet_lower = sheet_name.lower().strip()
        
        # Skip hidden sheets (starting with _)
        if sheet_lower.startswith("_"):
            return True
        
        # Skip sheets matching ignored patterns
        for pattern in IGNORED_SHEET_PATTERNS:
            if pattern in sheet_lower:
                return True
        
        # Check if it's a Scope 1 or Scope 2 sheet first
        if sheet_lower in ['scope1', 'scope 1', 'scope2', 'scope 2']:
            return False
        
        # Skip sheets that don't start with C1-C15
        # This handles any sheets after C15 like "Instructions", etc.
        is_category_sheet = any(sheet_lower.startswith(f"c{i}") for i in range(1, 16))
        if not is_category_sheet:
            # Check if it matches the configured sheet names
            for code, config in CATEGORY_COLUMNS.items():
                if config["sheet_name"].lower() == sheet_lower:
                    return False
                # Check aliases
                for alias in config.get("sheet_name_aliases", []):
                    if alias.lower() == sheet_lower:
                        return False
            # Not a recognized category sheet
            return True
        
        return False
    
    @staticmethod
    def _build_preview(valid_records: list) -> dict:
        """Build a dry-run preview summary from validated records.

        Returns a dict with counts by scope, category, fuel type (standard vs
        custom), and methodology — giving users a clear picture of what will be
        saved before they confirm.
        """
        by_scope: Dict[str, int] = {}
        by_category: Dict[str, int] = {}
        standard_fuel_count = 0
        custom_fuel_count = 0
        by_methodology: Dict[str, int] = {}
        total_co2e = 0.0

        for rec in valid_records:
            scope = rec.get("scope", "unknown")
            by_scope[scope] = by_scope.get(scope, 0) + 1

            cat = rec.get("category", "Unknown")
            by_category[cat] = by_category.get(cat, 0) + 1

            if rec.get("is_custom_fuel"):
                custom_fuel_count += 1
                meth = rec.get("calculation_methodology", "unknown")
                by_methodology[meth] = by_methodology.get(meth, 0) + 1
            else:
                standard_fuel_count += 1

            total_co2e += rec.get("co2e_emissions", 0) or 0

        return {
            "total_valid_records": len(valid_records),
            "standard_fuel_records": standard_fuel_count,
            "custom_fuel_records": custom_fuel_count,
            "by_scope": by_scope,
            "by_category": by_category,
            "by_methodology": by_methodology if by_methodology else None,
            "total_co2e_tco2e": round(total_co2e, 6),
        }

    async def _apply_period_row_limits(
        self,
        results: List[RowResult],
        records: List[Dict],
    ) -> Tuple[List[Dict], List[ValidationError]]:
        """Apply the canonical organization period allowance to validated rows."""
        accepted, rejected = await partition_records_by_period_row_limit(
            self.organization_id,
            "ghg",
            "emission_records",
            records,
            database=self.db,
        )
        quota_errors: List[ValidationError] = []
        result_by_emission_id = {
            result.emission_id: result
            for result in results
            if result.emission_id
        }
        used_fallback_results = set()

        for violation in rejected:
            record = violation["record"]
            result = result_by_emission_id.get(record.get("id"))
            if result is None:
                for candidate in results:
                    candidate_period = (candidate.row_data or {}).get("reporting_period")
                    candidate_key = (candidate.sheet, candidate.row)
                    if (
                        candidate.success
                        and candidate_key not in used_fallback_results
                        and candidate_period == violation["reporting_period"]
                    ):
                        result = candidate
                        used_fallback_results.add(candidate_key)
                        break

            error = ValidationError(
                sheet=result.sheet if result else str(record.get("scope") or "GHG"),
                row=result.row if result else 0,
                column="Reporting Month/Year",
                error_type="PERIOD_ROW_LIMIT_EXCEEDED",
                message=violation["message"],
                suggestion="Remove excess rows or ask your administrator to increase the GHG monthly row allowance.",
                severity=ErrorSeverity.ERROR,
            )
            quota_errors.append(error)
            if result is not None:
                result.success = False
                result.co2e = None
                result.errors.append(error)
            else:
                results.append(RowResult(
                    sheet=error.sheet,
                    row=error.row,
                    success=False,
                    emission_id=record.get("id"),
                    errors=[error],
                    row_data={"reporting_period": violation["reporting_period"]},
                ))

        return accepted, quota_errors
    
    async def process_upload(self, file_content: bytes, filename: str,
                              validate_only: bool = True) -> UploadSummary:
        """
        Process an uploaded Excel file.
        
        Args:
            file_content: Raw bytes of the Excel file
            filename: Original filename
            validate_only: If True, only validate without saving (default: True)
            
        Returns:
            UploadSummary with results (records not saved when validate_only=True)
        """
        job_id = str(uuid.uuid4())
        now = datetime.now(timezone.utc)
        
        # Create job record
        job = BulkUploadJob(
            id=job_id,
            organization_id=self.organization_id,
            uploaded_by=self.user_id,
            uploaded_at=now,
            filename=filename,
            status=UploadStatus.PROCESSING,
            allow_partial_success=not validate_only
        )
        
        await self.db.bulk_upload_jobs.insert_one(job.dict())
        
        try:
            # ── Resolve org GHG capabilities (single source of truth) ────
            if self.capabilities is None:
                self.capabilities = await resolve_ghg_capabilities(
                    self.db, self.organization_id
                )
                # Propagate to child processors
                self.scope12_processor.capabilities = self.capabilities

            # Load workbook
            workbook = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            
            # ── Row count pre-check ──────────────────────────────────────
            total_row_estimate = 0
            for _sn in workbook.sheetnames:
                if self._should_skip_sheet(_sn):
                    continue
                _ws = workbook[_sn]
                sheet_rows = (_ws.max_row or 1) - 1  # minus header
                if sheet_rows > MAX_ROWS_PER_SHEET:
                    raise ValueError(
                        f"Sheet '{_sn}' has ~{sheet_rows} data rows, "
                        f"which exceeds the limit of {MAX_ROWS_PER_SHEET} rows per sheet. "
                        f"Please split the data into smaller files."
                    )
                total_row_estimate += max(sheet_rows, 0)
            if total_row_estimate > MAX_TOTAL_ROWS:
                raise ValueError(
                    f"Workbook contains ~{total_row_estimate} total data rows across all sheets, "
                    f"which exceeds the limit of {MAX_TOTAL_ROWS}. "
                    f"Please split the data into smaller files."
                )
            
            all_results: List[RowResult] = []
            all_errors: List[ValidationError] = []
            all_warnings: List[ValidationError] = []
            all_emission_records: List[Dict] = []
            categories_processed = set()
            total_co2e = 0.0
            skipped_sheets = []
            
            # Process each sheet
            for sheet_name in workbook.sheetnames:
                # Skip helper and instruction sheets
                if self._should_skip_sheet(sheet_name):
                    skipped_sheets.append(sheet_name)
                    continue
                
                # Detect category from sheet name
                category_code = self._detect_category(sheet_name)
                if not category_code:
                    all_warnings.append(ValidationError(
                        sheet=sheet_name,
                        row=0,
                        column=None,
                        error_type="UNKNOWN_SHEET",
                        message=f"Sheet '{sheet_name}' does not match any recognized category",
                        suggestion="Use sheet names like 'Scope1', 'Scope2', 'C1-PurchasedGoods', etc.",
                        severity=ErrorSeverity.WARNING
                    ))
                    continue
                
                categories_processed.add(category_code)
                
                # ── Enforce org-level access ─────────────────────────────
                caps = self.capabilities
                sheet_blocked = False
                
                if category_code == "Scope1" and not caps.scope1_enabled:
                    sheet_blocked = True
                    block_msg = "Scope 1 is not enabled for your organization"
                elif category_code == "Scope2" and not caps.scope2_enabled:
                    sheet_blocked = True
                    block_msg = "Scope 2 is not enabled for your organization"
                elif category_code not in ("Scope1", "Scope2"):
                    # Scope 3 sheet
                    if not caps.is_scope3_sheet_enabled(category_code):
                        sheet_blocked = True
                        if not caps.scope3_enabled:
                            block_msg = "Scope 3 is not enabled for your organization"
                        else:
                            block_msg = f"Category {category_code} is disabled for your organization"
                
                if sheet_blocked:
                    all_errors.append(ValidationError(
                        sheet=sheet_name,
                        row=0,
                        column=None,
                        error_type="DISABLED_SCOPE_OR_CATEGORY",
                        message=block_msg,
                        suggestion="Contact your administrator to enable this scope/category, or remove this sheet from the workbook",
                        severity=ErrorSeverity.ERROR,
                    ))
                    continue
                
                # Process sheet
                sheet = workbook[sheet_name]
                sheet_results, sheet_records = await self._process_sheet(
                    sheet, category_code, sheet_name, job_id
                )
                
                all_results.extend(sheet_results)
                all_emission_records.extend(sheet_records)
                
                for result in sheet_results:
                    all_errors.extend(result.errors)
                    all_warnings.extend(result.warnings)
                    if result.co2e:
                        total_co2e += result.co2e
            
            workbook.close()
            
            valid_records = [r for r in all_emission_records if r is not None]
            valid_records, quota_errors = await self._apply_period_row_limits(
                all_results,
                valid_records,
            )
            all_errors.extend(quota_errors)

            # Counts reflect emission records that can actually be saved.
            success_count = len(valid_records)
            error_count = sum(1 for r in all_results if not r.success)
            total_co2e = sum((record.get("co2e_emissions", 0) or 0) for record in valid_records)
            
            if error_count == 0 and success_count > 0:
                status = UploadStatus.COMPLETED
            elif success_count == 0:
                status = UploadStatus.FAILED
            else:
                status = UploadStatus.PARTIAL_SUCCESS
            
            # Handle saving based on validate_only flag
            created_ids = []
            preview = self._build_preview(valid_records) if valid_records else None
            
            if validate_only and valid_records:
                # Store pending records for later save (with 24h expiry)
                expires_at = datetime.now(timezone.utc).replace(
                    hour=0, minute=0, second=0, microsecond=0
                )
                expires_at = expires_at.replace(day=expires_at.day + 1) if expires_at.day < 28 else expires_at
                from datetime import timedelta
                expires_at = datetime.now(timezone.utc) + timedelta(hours=24)
                for record in valid_records:
                    record["job_id"] = job_id
                    record["expires_at"] = expires_at
                await self.db.bulk_upload_pending_records.insert_many(valid_records)
            elif not validate_only and status in [UploadStatus.COMPLETED, UploadStatus.PARTIAL_SUCCESS] and valid_records:
                # Save immediately to emission_records collection
                await assert_period_row_batch_limit(
                    self.organization_id,
                    "ghg",
                    "emission_records",
                    valid_records,
                    database=self.db,
                )
                await self.db.emission_records.insert_many(valid_records)
                created_ids = [r["id"] for r in valid_records]
                
                # Create emission_history entries for version tracking
                now = datetime.now(timezone.utc)
                history_entries = []
                for record in valid_records:
                    history_entries.append({
                        "id": str(uuid.uuid4()),
                        "emission_id": record["id"],
                        "scope": record.get("scope", "scope3"),
                        "category": record.get("category", ""),
                        "reporting_month": record.get("reporting_period"),
                        "changed_by": self.user_id,
                        "changed_by_email": self.user_email,
                        "changed_by_name": self.user_name,
                        "changed_at": now.isoformat(),
                        "version": 1,
                        "field_changes": [],
                        "changes_summary": "Initial creation via bulk upload",
                        "changes": {
                            "action": "created",
                            "old_values": None,
                            "new_values": {
                                "facility_id": record.get("facility_id"),
                                "reporting_period": record.get("reporting_period"),
                                "category": record.get("category"),
                                "co2e_emissions": record.get("co2e_emissions"),
                                "total_emissions": record.get("total_emissions"),
                            }
                        }
                    })
                if history_entries:
                    await self.db.emission_history.insert_many(history_entries)
                    logger.info(f"[BULK_UPLOAD] Created {len(history_entries)} history entries")
            
            # Update job record (include serialized warnings for later download)
            warning_docs = [
                {
                    "sheet": w.sheet,
                    "row": w.row,
                    "column": w.column,
                    "error_type": w.error_type,
                    "message": w.message,
                    "suggestion": w.suggestion,
                    "severity": w.severity.value,
                }
                for w in all_warnings
            ]
            await self.db.bulk_upload_jobs.update_one(
                {"id": job_id},
                {"$set": {
                    "status": status.value,
                    "total_rows": len(all_results),
                    "success_count": success_count,
                    "error_count": error_count,
                    "warning_count": len(all_warnings),
                    "categories_processed": list(categories_processed),
                    "total_emissions_tco2e": total_co2e,
                    "created_emission_ids": created_ids,
                    "validate_only": validate_only,
                    "skipped_sheets": skipped_sheets,
                    "preview": preview,
                    "warnings": warning_docs,
                }}
            )
            
            # Store errors for later retrieval
            if all_errors:
                error_docs = [
                    {
                        "job_id": job_id,
                        "sheet": e.sheet,
                        "row": e.row,
                        "column": e.column,
                        "error_type": e.error_type,
                        "message": e.message,
                        "suggestion": e.suggestion,
                        "severity": e.severity.value
                    }
                    for e in all_errors
                ]
                await self.db.bulk_upload_errors.insert_many(error_docs)
            
            return UploadSummary(
                job_id=job_id,
                status=status,
                total_rows=len(all_results),
                success_count=success_count,
                error_count=error_count,
                warning_count=len(all_warnings),
                categories_processed=list(categories_processed),
                total_emissions_tco2e=total_co2e,
                errors=all_errors[:100],  # Limit errors in response
                warnings=all_warnings[:50],
                results=all_results[:200],
                created_emission_ids=created_ids,
                preview=preview,
            )
            
        except Exception as e:
            import traceback
            error_traceback = traceback.format_exc()
            print(f"Upload processing error: {e}")
            print(f"Traceback:\n{error_traceback}")
            
            # Update job as failed
            await self.db.bulk_upload_jobs.update_one(
                {"id": job_id},
                {"$set": {
                    "status": UploadStatus.FAILED.value,
                    "error_message": str(e),
                    "error_traceback": error_traceback
                }}
            )
            
            return UploadSummary(
                job_id=job_id,
                status=UploadStatus.FAILED,
                errors=[ValidationError(
                    sheet="System",
                    row=0,
                    error_type="SYSTEM_ERROR",
                    message=f"Failed to process file: {str(e)}",
                    severity=ErrorSeverity.ERROR
                )]
            )
    
    async def save_valid_rows(self, job_id: str) -> Dict:
        """
        Save valid rows from a previously validated upload job.
        
        Args:
            job_id: The job ID from validation
            
        Returns:
            Dict with save results
        """
        # Get job
        job = await self.db.bulk_upload_jobs.find_one(
            {"id": job_id, "organization_id": self.organization_id},
            {"_id": 0}
        )
        
        if not job:
            return {"success": False, "error": "Job not found"}
        
        if job.get("created_emission_ids"):
            return {"success": False, "error": "Records already saved for this job"}
        
        # Get valid results from the job
        # Note: We need to re-process or store pending records during validation
        # For now, return that we need to re-upload
        return {
            "success": False, 
            "error": "Please re-upload the file with save option enabled",
            "job_id": job_id
        }
    
    def _detect_category(self, sheet_name: str) -> Optional[str]:
        """Detect category code from sheet name"""
        sheet_lower = sheet_name.lower().strip()
        
        # Direct match with config sheet names
        for code, config in CATEGORY_COLUMNS.items():
            if config["sheet_name"].lower() == sheet_lower:
                return code
            # Check aliases
            for alias in config.get("sheet_name_aliases", []):
                if alias.lower() == sheet_lower:
                    return code
        
        # Pattern matching (C1, C2, etc.)
        for code in CATEGORY_COLUMNS.keys():
            if sheet_lower.startswith(code.lower()):
                return code
        
        return None
    
    async def _process_sheet(self, sheet, category_code: str, 
                              sheet_name: str, job_id: str) -> Tuple[List[RowResult], List[Dict]]:
        """
        Process a single sheet
        
        Returns:
            Tuple of (row_results, emission_records)
        """
        config = CATEGORY_COLUMNS.get(category_code, {})
        columns = config.get("columns", [])
        
        # Build column mapping from header row
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        col_mapping = {}
        
        for col_idx, header_value in enumerate(header_row):
            if header_value:
                # Clean header - remove newlines and extra spaces
                header_clean = str(header_value).strip().replace('\n', ' ').replace('  ', ' ')
                header_lower = header_clean.lower()
                
                # Find matching column config
                for col_config in columns:
                    col_name_lower = col_config["name"].lower()
                    # Check exact match
                    if col_name_lower == header_lower:
                        col_mapping[col_idx] = col_config["key"]
                        break
                    # Check aliases
                    for alias in col_config.get("aliases", []):
                        alias_clean = alias.replace('\n', ' ').replace('  ', ' ').lower()
                        if alias_clean == header_lower:
                            col_mapping[col_idx] = col_config["key"]
                            break
                    else:
                        continue
                    break
        
        # Process data rows
        results: List[RowResult] = []
        emission_records: List[Dict] = []
        existing_keys = set()
        
        # Check if this is a Scope 1 or Scope 2 sheet
        is_scope12 = category_code in ["Scope1", "Scope2"]
        
        # For C7, collect all rows first for aggregation
        if category_code == "C7":
            c7_rows = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if row_idx - 1 > MAX_ROWS_PER_SHEET:
                    break
                # Skip completely empty rows (e.g., blank spacing rows after header)
                if not any(cell is not None and cell != '' for cell in row):
                    continue
                row_data = self._row_to_dict(row, col_mapping)
                if any(v for v in row_data.values() if v):  # Skip rows with no mapped data
                    c7_rows.append((row_idx, row_data))
            
            if c7_rows:
                results, emission_records = await self.row_processor.process_c7_rows(
                    c7_rows, category_code, job_id
                )
            
            return results, emission_records
        
        # Regular processing for other categories (including Scope 1 and Scope 2)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row_idx - 1 > MAX_ROWS_PER_SHEET:
                break
            # Skip completely empty rows (e.g., blank spacing rows after header)
            if not any(cell is not None and cell != '' for cell in row):
                continue
            
            row_data = self._row_to_dict(row, col_mapping)
            
            # Skip rows with no mapped data
            if not any(v for v in row_data.values() if v):
                continue
            
            # Use appropriate processor based on sheet type
            if is_scope12:
                if category_code == "Scope1":
                    result, emission_record = await self.scope12_processor.process_scope1_row(
                        row_data, row_idx, existing_keys, job_id
                    )
                else:  # Scope2
                    result, emission_record = await self.scope12_processor.process_scope2_row(
                        row_data, row_idx, existing_keys, job_id
                    )
                
                results.append(result)
                if result.success and emission_record:
                    emission_records.append(emission_record)
            else:
                result = await self.row_processor.process_row(
                    row_data, category_code, row_idx, existing_keys, job_id
                )
                
                if isinstance(result, tuple):
                    row_result, emission_record = result
                    results.append(row_result)
                    if row_result.success:
                        emission_records.append(emission_record)
                else:
                    results.append(result)
        
        return results, emission_records
    
    async def _get_facility(self, facility_name: str) -> Optional[Dict]:
        """Get facility by name"""
        if not facility_name:
            return None
        return await self.db.facilities.find_one(
            {"name": facility_name, "organization_id": self.organization_id},
            {"_id": 0}
        )
    
    def _row_to_dict(self, row: tuple, col_mapping: Dict[int, str]) -> Dict[str, Any]:
        """Convert a row tuple to a dictionary using column mapping"""
        result = {}
        for col_idx, key in col_mapping.items():
            if col_idx < len(row):
                value = row[col_idx]
                # Clean string values
                if isinstance(value, str):
                    value = value.strip()
                result[key] = value
        return result
