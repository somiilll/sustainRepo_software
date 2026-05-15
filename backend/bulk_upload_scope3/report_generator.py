"""
Error Report Generator for Scope 3 Bulk Upload
Generates downloadable Excel reports with validation errors
"""
import io
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .models import ValidationError, ErrorSeverity, UploadSummary


class ReportGenerator:
    """Generates Excel reports for bulk upload results"""
    
    @staticmethod
    def generate_error_report(summary: UploadSummary) -> io.BytesIO:
        """
        Generate an Excel error report
        
        Args:
            summary: Upload summary with errors
            
        Returns:
            BytesIO containing the Excel workbook
        """
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Style definitions
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Summary content
        summary_data = [
            ["Bulk Upload Summary Report"],
            [""],
            ["Job ID", summary.job_id],
            ["Status", summary.status.value],
            [""],
            ["Total Rows", summary.total_rows],
            ["Successful Rows", summary.success_count],
            ["Failed Rows", summary.error_count],
            ["Warnings", summary.warning_count],
            [""],
            ["Total Emissions (tCO2e)", f"{summary.total_emissions_tco2e:.4f}"],
            ["Categories Processed", ", ".join(summary.categories_processed)],
        ]
        
        for row_idx, row_data in enumerate(summary_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                elif col_idx == 1 and row_idx > 2:
                    cell.font = Font(bold=True)
        
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 50
        
        # Color code status
        status_cell = ws_summary.cell(row=4, column=2)
        if summary.status.value == "completed":
            status_cell.fill = success_fill
        elif summary.status.value == "failed":
            status_cell.fill = error_fill
        else:
            status_cell.fill = warning_fill
        
        # Errors sheet
        if summary.errors:
            ws_errors = wb.create_sheet("Errors")
            
            # Headers
            error_headers = ["Sheet", "Row", "Column", "Error Type", "Message", "Suggestion"]
            for col_idx, header in enumerate(error_headers, start=1):
                cell = ws_errors.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Error rows
            for row_idx, error in enumerate(summary.errors, start=2):
                values = [
                    error.sheet,
                    error.row,
                    error.column,
                    error.error_type,
                    error.message,
                    error.suggestion
                ]
                for col_idx, value in enumerate(values, start=1):
                    cell = ws_errors.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    if error.severity == ErrorSeverity.ERROR:
                        cell.fill = error_fill
                    else:
                        cell.fill = warning_fill
            
            # Set column widths
            ws_errors.column_dimensions['A'].width = 30
            ws_errors.column_dimensions['B'].width = 8
            ws_errors.column_dimensions['C'].width = 20
            ws_errors.column_dimensions['D'].width = 25
            ws_errors.column_dimensions['E'].width = 50
            ws_errors.column_dimensions['F'].width = 50
            
            ws_errors.freeze_panes = "A2"
        
        # Warnings sheet
        if summary.warnings:
            ws_warnings = wb.create_sheet("Warnings")
            
            warning_headers = ["Sheet", "Row", "Column", "Warning Type", "Message", "Suggestion"]
            for col_idx, header in enumerate(warning_headers, start=1):
                cell = ws_warnings.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            
            for row_idx, warning in enumerate(summary.warnings, start=2):
                values = [
                    warning.sheet,
                    warning.row,
                    warning.column,
                    warning.error_type,
                    warning.message,
                    warning.suggestion
                ]
                for col_idx, value in enumerate(values, start=1):
                    cell = ws_warnings.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.fill = warning_fill
            
            ws_warnings.column_dimensions['A'].width = 30
            ws_warnings.column_dimensions['B'].width = 8
            ws_warnings.column_dimensions['C'].width = 20
            ws_warnings.column_dimensions['D'].width = 25
            ws_warnings.column_dimensions['E'].width = 50
            ws_warnings.column_dimensions['F'].width = 50
            
            ws_warnings.freeze_panes = "A2"
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def generate_results_report(summary: UploadSummary, emission_records: List[Dict]) -> io.BytesIO:
        """
        Generate an Excel report with processed results
        
        Args:
            summary: Upload summary
            emission_records: List of created emission records
            
        Returns:
            BytesIO containing the Excel workbook
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Processed Emissions"
        
        # Style definitions
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            "Emission ID", "Category", "Facility", "Reporting Period",
            "Calculation Method", "Activity", "CO2e (tCO2e)", "Status"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for row_idx, record in enumerate(emission_records, start=2):
            values = [
                record.get("id", ""),
                record.get("category", ""),
                record.get("facility_name", ""),
                record.get("reporting_period", ""),
                record.get("calculation_method_scope3", ""),
                record.get("scope3_activity", ""),
                record.get("co2e_emissions", 0),
                "Saved"
            ]
            
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
        
        # Set column widths
        widths = [40, 35, 25, 15, 18, 40, 15, 10]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        
        ws.freeze_panes = "A2"
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
