"""
Bulk Upload Module for GHG Emissions Data
Supports Excel-based bulk upload with validation, error handling, and audit trail.
Aligned with Greenhouse Gas Protocol and ISO 14064.
"""

import io
import uuid
import re
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple
from fastapi import APIRouter, HTTPException, UploadFile, File, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from rapidfuzz import fuzz, process


def normalize_string(s: str) -> str:
    """Normalize string for case-insensitive matching."""
    if not s:
        return ""
    # Convert to lowercase, strip whitespace, replace multiple spaces
    normalized = s.lower().strip()
    normalized = re.sub(r'\s+', ' ', normalized)
    return normalized


def find_best_match(value: str, candidates: List[str], threshold: int = 80) -> Tuple[Optional[str], int]:
    """
    Find the best fuzzy match for a value in candidates.
    Returns (matched_value, score) or (None, 0) if no good match.
    """
    if not value or not candidates:
        return None, 0
    
    normalized_value = normalize_string(value)
    normalized_candidates = {normalize_string(c): c for c in candidates}
    
    # First try exact match (normalized)
    if normalized_value in normalized_candidates:
        return normalized_candidates[normalized_value], 100
    
    # Fuzzy match
    result = process.extractOne(
        normalized_value, 
        list(normalized_candidates.keys()),
        scorer=fuzz.ratio
    )
    
    if result and result[1] >= threshold:
        return normalized_candidates[result[0]], result[1]
    
    return None, 0


def get_suggestions(value: str, candidates: List[str], limit: int = 3) -> List[str]:
    """Get top suggestions for a misspelled value."""
    if not value or not candidates:
        return []
    
    normalized_value = normalize_string(value)
    normalized_candidates = {normalize_string(c): c for c in candidates}
    
    results = process.extract(
        normalized_value,
        list(normalized_candidates.keys()),
        scorer=fuzz.ratio,
        limit=limit
    )
    
    return [normalized_candidates[r[0]] for r in results if r[1] >= 50]


def create_bulk_upload_router(db, get_current_user, get_admin_user):
    """Create the bulk upload router with database access."""
    
    router = APIRouter(tags=["Bulk Upload"])
    
    # Style definitions
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    ERROR_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    VALID_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    async def get_reference_data(org_id: str) -> Dict[str, Any]:
        """Fetch all reference data needed for validation."""
        
        # Get facilities for this organization
        facilities = await db.facilities.find(
            {"organization_id": org_id, "is_active": {"$ne": False}},
            {"_id": 0, "id": 1, "name": 1}
        ).to_list(1000)
        
        # Get scopes
        scopes = await db.scopes.find(
            {"is_active": {"$ne": False}},
            {"_id": 0, "id": 1, "name": 1, "code": 1}
        ).to_list(100)
        
        # Get categories
        categories = await db.emission_categories.find(
            {"is_active": {"$ne": False}},
            {"_id": 0, "id": 1, "name": 1, "scope_id": 1}
        ).to_list(500)
        
        # Get units
        units = await db.units.find(
            {"is_active": {"$ne": False}},
            {"_id": 0, "symbol": 1, "name": 1, "unit_type": 1}
        ).to_list(500)
        
        # Get fuel database (for Scope 1/2)
        fuels = await db.fuel_database.find(
            {"is_active": {"$ne": False}},
            {"_id": 0, "id": 1, "fuel_name": 1, "category": 1, "scope": 1}
        ).to_list(1000)
        
        # Get Scope 3 EF entries
        scope3_ef = await db.scope3_ef.find(
            {},
            {"_id": 0, "id": 1, "activity": 1, "method": 1, "category": 1, "scope": 1, 
             "emission_factor": 1, "unit": 1, "allowed_units": 1}
        ).to_list(1000)
        
        return {
            "facilities": facilities,
            "scopes": scopes,
            "categories": categories,
            "units": units,
            "fuels": fuels,
            "scope3_ef": scope3_ef,
            "facility_names": [f["name"] for f in facilities],
            "scope_names": [s["name"] for s in scopes],
            "unit_symbols": [u["symbol"] for u in units],
            "methods": ["spend_basis", "activity_basis"],
        }
    
    @router.get("/bulk-upload/template")
    async def download_template(current_user: dict = Depends(get_admin_user)):
        """Generate and download Excel template with dropdowns and reference data."""
        
        org_id = current_user.get("organization_id")
        ref_data = await get_reference_data(org_id)
        
        wb = Workbook()
        
        # ========== MAIN DATA SHEET ==========
        ws = wb.active
        ws.title = "Emissions Data"
        
        # Define columns
        columns = [
            ("facility", "Facility Name *", True),
            ("reporting_month", "Reporting Month (YYYY-MM) *", True),
            ("scope", "Scope *", True),
            ("category", "Category *", True),
            ("activity", "Activity/Fuel *", True),
            ("method", "Method *", True),
            ("quantity", "Quantity *", True),
            ("quantity_unit", "Quantity Unit *", True),
            ("emission_factor", "Emission Factor (optional)", False),
            ("ef_unit", "EF Unit (if EF provided)", False),
            ("evidence_reference", "Evidence Reference", False),
            ("notes", "Notes", False),
        ]
        
        # Write headers
        for col_idx, (key, label, required) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            ws.column_dimensions[cell.column_letter].width = 18
        
        # Lock header row
        ws.row_dimensions[1].height = 30
        
        # Add data validations (dropdowns)
        # Facility dropdown
        if ref_data["facility_names"]:
            facility_dv = DataValidation(
                type="list",
                formula1=f'"{"、".join(ref_data["facility_names"][:200])}"' if len(ref_data["facility_names"]) <= 200 else None,
                allow_blank=False,
                showDropDown=False
            )
            facility_dv.error = "Please select a valid facility"
            facility_dv.errorTitle = "Invalid Facility"
            ws.add_data_validation(facility_dv)
            facility_dv.add("A2:A1000")
        
        # Scope dropdown
        scope_dv = DataValidation(
            type="list",
            formula1='"Scope 1,Scope 2,Scope 3,Biogenic"',
            allow_blank=False
        )
        scope_dv.error = "Please select Scope 1, Scope 2, Scope 3, or Biogenic"
        ws.add_data_validation(scope_dv)
        scope_dv.add("C2:C1000")
        
        # Method dropdown
        method_dv = DataValidation(
            type="list",
            formula1='"spend_basis,activity_basis"',
            allow_blank=False
        )
        method_dv.error = "Please select spend_basis or activity_basis"
        ws.add_data_validation(method_dv)
        method_dv.add("F2:F1000")
        
        # Add example rows
        example_data = [
            ["Main Office", "2024-01", "Scope 1", "Stationary Combustion", "Diesel", "activity_basis", 1000, "L", "", "", "Invoice #123", "January fuel consumption"],
            ["Warehouse", "2024-01", "Scope 3", "Purchased Goods", "Business Travel", "spend_basis", 50000, "INR", "", "", "Expense Report", "Q1 travel expenses"],
        ]
        
        for row_idx, row_data in enumerate(example_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
        
        # ========== REFERENCE SHEET ==========
        ref_ws = wb.create_sheet("Reference Data")
        
        # Facilities
        ref_ws.cell(row=1, column=1, value="Valid Facilities").font = Font(bold=True)
        for idx, facility in enumerate(ref_data["facility_names"], 2):
            ref_ws.cell(row=idx, column=1, value=facility)
        
        # Scopes
        ref_ws.cell(row=1, column=3, value="Valid Scopes").font = Font(bold=True)
        for idx, scope in enumerate(ref_data["scope_names"], 2):
            ref_ws.cell(row=idx, column=3, value=scope)
        
        # Methods
        ref_ws.cell(row=1, column=5, value="Valid Methods").font = Font(bold=True)
        ref_ws.cell(row=2, column=5, value="spend_basis")
        ref_ws.cell(row=3, column=5, value="activity_basis")
        
        # Units
        ref_ws.cell(row=1, column=7, value="Valid Units").font = Font(bold=True)
        for idx, unit in enumerate(ref_data["unit_symbols"], 2):
            ref_ws.cell(row=idx, column=7, value=unit)
        
        # Scope-Category mapping
        ref_ws.cell(row=1, column=9, value="Scope").font = Font(bold=True)
        ref_ws.cell(row=1, column=10, value="Category").font = Font(bold=True)
        row_idx = 2
        for scope in ref_data["scopes"]:
            scope_cats = [c for c in ref_data["categories"] if c["scope_id"] == scope["id"]]
            for cat in scope_cats:
                ref_ws.cell(row=row_idx, column=9, value=scope["name"])
                ref_ws.cell(row=row_idx, column=10, value=cat["name"])
                row_idx += 1
        
        # Activities (Scope 3)
        ref_ws.cell(row=1, column=12, value="Scope 3 Activities").font = Font(bold=True)
        activities = list(set([ef["activity"] for ef in ref_data["scope3_ef"] if ef.get("activity")]))
        for idx, activity in enumerate(sorted(activities), 2):
            ref_ws.cell(row=idx, column=12, value=activity)
        
        # Fuels (Scope 1/2)
        ref_ws.cell(row=1, column=14, value="Fuels (Scope 1/2)").font = Font(bold=True)
        fuel_names = list(set([f["fuel_name"] for f in ref_data["fuels"] if f.get("fuel_name")]))
        for idx, fuel in enumerate(sorted(fuel_names), 2):
            ref_ws.cell(row=idx, column=14, value=fuel)
        
        # Adjust column widths
        for col in ['A', 'C', 'E', 'G', 'I', 'J', 'L', 'N']:
            ref_ws.column_dimensions[col].width = 25
        
        # ========== INSTRUCTIONS SHEET ==========
        instr_ws = wb.create_sheet("Instructions")
        instructions = [
            ("GHG Emissions Bulk Upload - Instructions", True),
            ("", False),
            ("Required Fields:", True),
            ("• facility - Must match exactly with your organization's facilities", False),
            ("• reporting_month - Format: YYYY-MM (e.g., 2024-01)", False),
            ("• scope - Scope 1, Scope 2, Scope 3, or Biogenic", False),
            ("• category - Must be valid for the selected scope (see Reference Data)", False),
            ("• activity - For Scope 1/2: Fuel name. For Scope 3: Activity type", False),
            ("• method - spend_basis (for monetary amounts) or activity_basis (for quantities)", False),
            ("• quantity - Numeric value", False),
            ("• quantity_unit - Must match the method:", False),
            ("  - spend_basis → Currency units (INR, USD, EUR)", False),
            ("  - activity_basis → Physical units (kg, L, kWh, km)", False),
            ("", False),
            ("Optional Fields:", True),
            ("• emission_factor - Override the system emission factor", False),
            ("• ef_unit - Required if emission_factor is provided", False),
            ("• evidence_reference - Document/invoice reference", False),
            ("• notes - Additional notes", False),
            ("", False),
            ("Tips:", True),
            ("• Check the 'Reference Data' sheet for valid values", False),
            ("• The system will auto-fetch emission factors if not provided", False),
            ("• Values are matched case-insensitively (Steel = steel = STEEL)", False),
            ("• Remove example rows before uploading", False),
        ]
        
        for idx, (text, is_bold) in enumerate(instructions, 1):
            cell = instr_ws.cell(row=idx, column=1, value=text)
            if is_bold:
                cell.font = Font(bold=True, size=12)
        
        instr_ws.column_dimensions['A'].width = 80
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"GHG_Emissions_Template_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    @router.post("/bulk-upload/validate")
    async def validate_upload(
        file: UploadFile = File(...),
        current_user: dict = Depends(get_admin_user)
    ):
        """Parse and validate uploaded Excel file."""
        
        if not file.filename.endswith('.xlsx'):
            raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
        
        org_id = current_user.get("organization_id")
        upload_id = str(uuid.uuid4())
        
        # Read file
        contents = await file.read()
        
        try:
            wb = load_workbook(io.BytesIO(contents))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")
        
        ws = wb.active
        ref_data = await get_reference_data(org_id)
        
        # Build lookup maps
        facility_map = {normalize_string(f["name"]): f for f in ref_data["facilities"]}
        scope_map = {normalize_string(s["name"]): s for s in ref_data["scopes"]}
        fuel_map = {normalize_string(f["fuel_name"]): f for f in ref_data["fuels"]}
        
        # Scope 3 activities
        scope3_activities = list(set([ef["activity"] for ef in ref_data["scope3_ef"] if ef.get("activity")]))
        
        # Parse headers
        headers = [cell.value for cell in ws[1]]
        expected_headers = ["facility", "reporting_month", "scope", "category", "activity", 
                          "method", "quantity", "quantity_unit", "emission_factor", 
                          "ef_unit", "evidence_reference", "notes"]
        
        # Map column indices
        col_map = {}
        for idx, header in enumerate(headers):
            if header:
                # Try to match header to expected columns
                normalized_header = normalize_string(header.replace("*", "").strip())
                for exp in expected_headers:
                    if exp in normalized_header or normalized_header in exp:
                        col_map[exp] = idx
                        break
        
        rows_result = []
        valid_count = 0
        invalid_count = 0
        
        # Process each row (skip header and example rows that start with example data)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            # Skip empty rows
            if not any(row):
                continue
            
            # Skip example rows (light blue background rows)
            if row_idx <= 3 and row[0] in ["Main Office", "Warehouse"]:
                continue
            
            row_data = {
                "facility": row[col_map.get("facility", 0)] if "facility" in col_map else row[0],
                "reporting_month": row[col_map.get("reporting_month", 1)] if "reporting_month" in col_map else row[1],
                "scope": row[col_map.get("scope", 2)] if "scope" in col_map else row[2],
                "category": row[col_map.get("category", 3)] if "category" in col_map else row[3],
                "activity": row[col_map.get("activity", 4)] if "activity" in col_map else row[4],
                "method": row[col_map.get("method", 5)] if "method" in col_map else row[5],
                "quantity": row[col_map.get("quantity", 6)] if "quantity" in col_map else row[6],
                "quantity_unit": row[col_map.get("quantity_unit", 7)] if "quantity_unit" in col_map else row[7],
                "emission_factor": row[col_map.get("emission_factor", 8)] if "emission_factor" in col_map else row[8],
                "ef_unit": row[col_map.get("ef_unit", 9)] if "ef_unit" in col_map else row[9],
                "evidence_reference": row[col_map.get("evidence_reference", 10)] if "evidence_reference" in col_map else row[10],
                "notes": row[col_map.get("notes", 11)] if "notes" in col_map else row[11],
            }
            
            errors = []
            matched_data = {}
            
            # ========== LAYER 1: Schema Validation ==========
            
            # Required fields
            required_fields = ["facility", "reporting_month", "scope", "category", "activity", "method", "quantity", "quantity_unit"]
            for field in required_fields:
                if not row_data.get(field):
                    errors.append({
                        "column": field,
                        "message": f"Required field '{field}' is missing",
                        "suggestion": "Please provide a value"
                    })
            
            # Quantity must be numeric
            if row_data.get("quantity"):
                try:
                    matched_data["quantity"] = float(row_data["quantity"])
                except (ValueError, TypeError):
                    errors.append({
                        "column": "quantity",
                        "message": "Quantity must be a number",
                        "suggestion": f"Got '{row_data['quantity']}' - please enter a numeric value"
                    })
            
            # Reporting month format
            if row_data.get("reporting_month"):
                month_str = str(row_data["reporting_month"])
                if not re.match(r'^\d{4}-\d{2}$', month_str):
                    # Try to parse other formats
                    errors.append({
                        "column": "reporting_month",
                        "message": "Invalid date format",
                        "suggestion": "Use YYYY-MM format (e.g., 2024-01)"
                    })
                else:
                    matched_data["reporting_month"] = month_str
            
            # ========== LAYER 2: Referential Validation ==========
            
            # Facility validation
            if row_data.get("facility"):
                facility_match, score = find_best_match(str(row_data["facility"]), ref_data["facility_names"])
                if facility_match:
                    matched_data["facility"] = facility_match
                    matched_data["facility_id"] = facility_map[normalize_string(facility_match)]["id"]
                else:
                    suggestions = get_suggestions(str(row_data["facility"]), ref_data["facility_names"])
                    errors.append({
                        "column": "facility",
                        "message": f"Facility '{row_data['facility']}' not found",
                        "suggestion": f"Did you mean: {', '.join(suggestions)}" if suggestions else "Check Reference Data sheet"
                    })
            
            # Scope validation
            if row_data.get("scope"):
                scope_match, _ = find_best_match(str(row_data["scope"]), ref_data["scope_names"])
                if scope_match:
                    matched_data["scope"] = scope_match
                    matched_data["scope_id"] = scope_map[normalize_string(scope_match)]["id"]
                else:
                    errors.append({
                        "column": "scope",
                        "message": f"Invalid scope '{row_data['scope']}'",
                        "suggestion": "Use: Scope 1, Scope 2, Scope 3, or Biogenic"
                    })
            
            # Category validation (must belong to scope)
            if row_data.get("category") and matched_data.get("scope_id"):
                scope_categories = [c["name"] for c in ref_data["categories"] if c["scope_id"] == matched_data["scope_id"]]
                cat_match, _ = find_best_match(str(row_data["category"]), scope_categories)
                if cat_match:
                    matched_data["category"] = cat_match
                    matched_data["category_id"] = next(
                        (c["id"] for c in ref_data["categories"] 
                         if normalize_string(c["name"]) == normalize_string(cat_match) 
                         and c["scope_id"] == matched_data["scope_id"]),
                        None
                    )
                else:
                    suggestions = get_suggestions(str(row_data["category"]), scope_categories)
                    errors.append({
                        "column": "category",
                        "message": f"Category '{row_data['category']}' not valid for {matched_data.get('scope', row_data.get('scope'))}",
                        "suggestion": f"Valid categories: {', '.join(scope_categories[:5])}" + ("..." if len(scope_categories) > 5 else "")
                    })
            
            # Method validation
            if row_data.get("method"):
                method_str = normalize_string(str(row_data["method"]))
                if method_str in ["spend_basis", "spend", "spend-basis", "spendbasis"]:
                    matched_data["method"] = "spend_basis"
                elif method_str in ["activity_basis", "activity", "activity-basis", "activitybasis"]:
                    matched_data["method"] = "activity_basis"
                else:
                    errors.append({
                        "column": "method",
                        "message": f"Invalid method '{row_data['method']}'",
                        "suggestion": "Use 'spend_basis' or 'activity_basis'"
                    })
            
            # Activity validation (different for Scope 3 vs Scope 1/2)
            if row_data.get("activity") and matched_data.get("scope"):
                is_scope3 = "3" in matched_data["scope"]
                
                if is_scope3:
                    # Match against Scope 3 activities
                    activity_match, _ = find_best_match(str(row_data["activity"]), scope3_activities)
                    if activity_match:
                        matched_data["activity"] = activity_match
                    else:
                        suggestions = get_suggestions(str(row_data["activity"]), scope3_activities)
                        errors.append({
                            "column": "activity",
                            "message": f"Activity '{row_data['activity']}' not found in Scope 3 EF table",
                            "suggestion": f"Did you mean: {', '.join(suggestions)}" if suggestions else "Check Reference Data sheet"
                        })
                else:
                    # Match against fuels
                    fuel_names = [f["fuel_name"] for f in ref_data["fuels"]]
                    fuel_match, _ = find_best_match(str(row_data["activity"]), fuel_names)
                    if fuel_match:
                        matched_data["activity"] = fuel_match
                        matched_data["fuel_id"] = fuel_map[normalize_string(fuel_match)]["id"]
                    else:
                        suggestions = get_suggestions(str(row_data["activity"]), fuel_names)
                        errors.append({
                            "column": "activity",
                            "message": f"Fuel '{row_data['activity']}' not found in Fuel Database",
                            "suggestion": f"Did you mean: {', '.join(suggestions)}" if suggestions else "Check Reference Data sheet"
                        })
            
            # ========== LAYER 3: Calculation Validation ==========
            
            # Unit compatibility with method
            if row_data.get("quantity_unit") and matched_data.get("method"):
                unit_str = str(row_data["quantity_unit"]).strip()
                currency_units = ["INR", "USD", "EUR", "GBP", "JPY", "AUD", "CAD"]
                
                if matched_data["method"] == "spend_basis":
                    # Should be currency
                    if unit_str.upper() not in currency_units:
                        errors.append({
                            "column": "quantity_unit",
                            "message": f"Unit '{unit_str}' is not valid for spend_basis method",
                            "suggestion": f"Use a currency unit: {', '.join(currency_units)}"
                        })
                    else:
                        matched_data["quantity_unit"] = unit_str.upper()
                else:
                    # Should be physical unit
                    unit_match, _ = find_best_match(unit_str, ref_data["unit_symbols"])
                    if unit_match:
                        matched_data["quantity_unit"] = unit_match
                    elif unit_str.upper() in currency_units:
                        errors.append({
                            "column": "quantity_unit",
                            "message": f"Currency unit '{unit_str}' not valid for activity_basis method",
                            "suggestion": "Use a physical unit (kg, L, kWh, km, etc.)"
                        })
                    else:
                        suggestions = get_suggestions(unit_str, ref_data["unit_symbols"])
                        errors.append({
                            "column": "quantity_unit",
                            "message": f"Unit '{unit_str}' not found",
                            "suggestion": f"Did you mean: {', '.join(suggestions)}" if suggestions else "Check Reference Data sheet"
                        })
            
            # Emission factor validation (if provided)
            if row_data.get("emission_factor"):
                try:
                    matched_data["emission_factor"] = float(row_data["emission_factor"])
                    if not row_data.get("ef_unit"):
                        errors.append({
                            "column": "ef_unit",
                            "message": "EF unit is required when emission factor is provided",
                            "suggestion": "Specify the emission factor unit (e.g., kgCO2e/kg)"
                        })
                    else:
                        matched_data["ef_unit"] = str(row_data["ef_unit"]).strip()
                except (ValueError, TypeError):
                    errors.append({
                        "column": "emission_factor",
                        "message": "Emission factor must be a number",
                        "suggestion": f"Got '{row_data['emission_factor']}'"
                    })
            
            # Copy optional fields
            if row_data.get("evidence_reference"):
                matched_data["evidence_reference"] = str(row_data["evidence_reference"])
            if row_data.get("notes"):
                matched_data["notes"] = str(row_data["notes"])
            
            # Determine status
            status = "valid" if not errors else "invalid"
            if status == "valid":
                valid_count += 1
            else:
                invalid_count += 1
            
            rows_result.append({
                "row_number": row_idx,
                "status": status,
                "original_data": {k: str(v) if v is not None else "" for k, v in row_data.items()},
                "matched_data": matched_data,
                "errors": errors
            })
        
        # Store upload session in database
        upload_session = {
            "id": upload_id,
            "organization_id": org_id,
            "uploaded_by": current_user.get("id"),
            "uploaded_by_email": current_user.get("email"),
            "filename": file.filename,
            "total_rows": len(rows_result),
            "valid_rows": valid_count,
            "invalid_rows": invalid_count,
            "rows": rows_result,
            "status": "pending",
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        
        await db.bulk_upload_sessions.insert_one(upload_session)
        
        return {
            "upload_id": upload_id,
            "summary": {
                "total_rows": len(rows_result),
                "valid_rows": valid_count,
                "invalid_rows": invalid_count
            },
            "rows": rows_result
        }
    
    @router.post("/bulk-upload/{upload_id}/save")
    async def save_valid_rows(
        upload_id: str,
        save_mode: str = "valid_only",  # valid_only, all_or_nothing
        current_user: dict = Depends(get_admin_user)
    ):
        """Save valid rows from upload session."""
        
        org_id = current_user.get("organization_id")
        
        # Get upload session
        session = await db.bulk_upload_sessions.find_one(
            {"id": upload_id, "organization_id": org_id},
            {"_id": 0}
        )
        
        if not session:
            raise HTTPException(status_code=404, detail="Upload session not found")
        
        if session["status"] == "completed":
            raise HTTPException(status_code=400, detail="Upload already processed")
        
        valid_rows = [r for r in session["rows"] if r["status"] == "valid"]
        
        if save_mode == "all_or_nothing" and session["invalid_rows"] > 0:
            raise HTTPException(
                status_code=400, 
                detail=f"Cannot save: {session['invalid_rows']} rows have errors. Fix errors or use 'valid_only' mode."
            )
        
        if not valid_rows:
            raise HTTPException(status_code=400, detail="No valid rows to save")
        
        # Process and save each valid row
        saved_count = 0
        saved_ids = []
        
        for row in valid_rows:
            data = row["matched_data"]
            
            # Create emission entry
            emission_entry = {
                "id": str(uuid.uuid4()),
                "organization_id": org_id,
                "facility_id": data.get("facility_id"),
                "facility_name": data.get("facility"),
                "reporting_month": data.get("reporting_month"),
                "scope": data.get("scope"),
                "scope_id": data.get("scope_id"),
                "category": data.get("category"),
                "category_id": data.get("category_id"),
                "activity": data.get("activity"),
                "fuel_id": data.get("fuel_id"),
                "method": data.get("method"),
                "quantity": data.get("quantity"),
                "quantity_unit": data.get("quantity_unit"),
                "emission_factor": data.get("emission_factor"),
                "ef_unit": data.get("ef_unit"),
                "evidence_reference": data.get("evidence_reference"),
                "notes": data.get("notes"),
                "source": "bulk_upload",
                "upload_id": upload_id,
                "created_by": current_user.get("id"),
                "created_at": datetime.now(timezone.utc).isoformat(),
                "status": "draft",  # Can be changed to calculate emissions
            }
            
            await db.emissions.insert_one(emission_entry)
            saved_ids.append(emission_entry["id"])
            saved_count += 1
        
        # Update session status
        await db.bulk_upload_sessions.update_one(
            {"id": upload_id},
            {
                "$set": {
                    "status": "completed",
                    "saved_count": saved_count,
                    "saved_ids": saved_ids,
                    "completed_at": datetime.now(timezone.utc).isoformat()
                }
            }
        )
        
        return {
            "message": f"Successfully saved {saved_count} emission entries",
            "saved_count": saved_count,
            "skipped_count": session["invalid_rows"],
            "saved_ids": saved_ids
        }
    
    @router.get("/bulk-upload/{upload_id}/error-report")
    async def download_error_report(
        upload_id: str,
        current_user: dict = Depends(get_admin_user)
    ):
        """Generate and download error report Excel."""
        
        org_id = current_user.get("organization_id")
        
        session = await db.bulk_upload_sessions.find_one(
            {"id": upload_id, "organization_id": org_id},
            {"_id": 0}
        )
        
        if not session:
            raise HTTPException(status_code=404, detail="Upload session not found")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Error Report"
        
        # Headers
        headers = ["Row #", "Status", "Facility", "Reporting Month", "Scope", "Category", 
                   "Activity", "Method", "Quantity", "Unit", "Error Message", "Suggestion"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
        
        # Data rows
        for row_idx, row in enumerate(session["rows"], 2):
            orig = row["original_data"]
            
            ws.cell(row=row_idx, column=1, value=row["row_number"])
            ws.cell(row=row_idx, column=2, value=row["status"].upper())
            ws.cell(row=row_idx, column=3, value=orig.get("facility", ""))
            ws.cell(row=row_idx, column=4, value=orig.get("reporting_month", ""))
            ws.cell(row=row_idx, column=5, value=orig.get("scope", ""))
            ws.cell(row=row_idx, column=6, value=orig.get("category", ""))
            ws.cell(row=row_idx, column=7, value=orig.get("activity", ""))
            ws.cell(row=row_idx, column=8, value=orig.get("method", ""))
            ws.cell(row=row_idx, column=9, value=orig.get("quantity", ""))
            ws.cell(row=row_idx, column=10, value=orig.get("quantity_unit", ""))
            
            # Combine errors
            if row["errors"]:
                error_msgs = "; ".join([e["message"] for e in row["errors"]])
                suggestions = "; ".join([e["suggestion"] for e in row["errors"] if e.get("suggestion")])
                ws.cell(row=row_idx, column=11, value=error_msgs)
                ws.cell(row=row_idx, column=12, value=suggestions)
            
            # Highlight invalid rows
            fill = ERROR_FILL if row["status"] == "invalid" else VALID_FILL
            for col_idx in range(1, 13):
                ws.cell(row=row_idx, column=col_idx).fill = fill
                ws.cell(row=row_idx, column=col_idx).border = THIN_BORDER
        
        # Adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"Error_Report_{upload_id[:8]}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    @router.get("/bulk-upload/sessions")
    async def list_upload_sessions(
        current_user: dict = Depends(get_admin_user)
    ):
        """List recent upload sessions."""
        
        org_id = current_user.get("organization_id")
        
        sessions = await db.bulk_upload_sessions.find(
            {"organization_id": org_id},
            {"_id": 0, "rows": 0}  # Exclude rows for performance
        ).sort("created_at", -1).limit(20).to_list(20)
        
        return sessions
    
    return router
