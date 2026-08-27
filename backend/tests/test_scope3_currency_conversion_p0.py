"""P0 Scope 3 spend-basis currency conversion regression tests."""

import json
import os
import uuid

import pytest
import requests
from openpyxl import load_workbook
from pymongo import MongoClient
from io import BytesIO


def _load_env(path: str) -> None:
    try:
        with open(path, "r", encoding="utf-8") as file:
            for line in file:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))
    except FileNotFoundError:
        pass


_load_env("/app/frontend/.env")
_load_env("/app/backend/.env")

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL")
if not BASE_URL:
    raise RuntimeError("REACT_APP_BACKEND_URL is required")
BASE_URL = BASE_URL.rstrip("/")
API = f"{BASE_URL}/api"

SUPER_ADMIN_EMAIL = "superadmin@ecotrack.com"
ADMIN_EMAIL = "goyalsomil2001@gmail.com"
PASSWORD = "TestUser123!"

PPP_FORMULA_ID = "6a3c49f2-3cd0-4a6e-ab9a-8ec2f4e1eecb"
STANDARD_FORMULA_ID = "8a9150c2-ea89-4f53-9f85-2a62f64d1028"


# Auth and shared sessions
@pytest.fixture(scope="module")
def super_admin_headers():
    response = requests.post(
        f"{API}/auth/login",
        json={"email": SUPER_ADMIN_EMAIL, "password": PASSWORD},
        timeout=30,
    )
    if response.status_code != 200:
        pytest.skip(f"Super admin login failed: {response.status_code} {response.text[:300]}")
    token = response.json().get("access_token") or response.json().get("token")
    if not token:
        pytest.skip("Super admin token missing")
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


@pytest.fixture(scope="module")
def admin_headers():
    response = requests.post(
        f"{API}/auth/login",
        json={"email": ADMIN_EMAIL, "password": PASSWORD},
        timeout=30,
    )
    if response.status_code != 200:
        pytest.skip(f"Admin login failed: {response.status_code} {response.text[:300]}")
    token = response.json().get("access_token") or response.json().get("token")
    if not token:
        pytest.skip("Admin token missing")
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


# DB-backed helpers for deterministic category selection + cleanup
@pytest.fixture(scope="module")
def db_client():
    mongo_url = os.environ.get("MONGO_URL")
    db_name = os.environ.get("DB_NAME")
    if not mongo_url or not db_name:
        pytest.skip("MONGO_URL / DB_NAME unavailable")
    client = MongoClient(mongo_url)
    db = client[db_name]
    yield db
    client.close()


@pytest.fixture(scope="module")
def spend_basis_category_id(db_client):
    trees = db_client.ce_decision_trees.find({"is_active": True}, {"_id": 0})
    for tree_doc in trees:
        tree_json = json.dumps(tree_doc.get("tree", {}))
        if PPP_FORMULA_ID in tree_json and STANDARD_FORMULA_ID in tree_json:
            category_id = tree_doc.get("category_id")
            if category_id:
                return category_id
    pytest.skip("No active Scope 3 decision tree found with both PPP and standard spend formulas")


@pytest.fixture()
def temp_standard_rates(super_admin_headers):
    created_ids = []

    yearly_payload = {
        "source_currency": "INR",
        "target_currency": "USD",
        "year_applicable": 2025,
        "month_applicable": None,
        "conversion_method": "standard",
        "exchange_rate": 80.0,
        "source": "PYTEST_TEMP_YEARLY",
        "notes": "temporary yearly rate for precedence check",
        "is_active": True,
    }
    monthly_payload = {
        "source_currency": "INR",
        "target_currency": "USD",
        "year_applicable": 2025,
        "month_applicable": 5,
        "conversion_method": "standard",
        "exchange_rate": 90.0,
        "source": "PYTEST_TEMP_MONTHLY",
        "notes": "temporary monthly rate for precedence check",
        "is_active": True,
    }

    for payload in [yearly_payload, monthly_payload]:
        response = requests.post(
            f"{API}/super-admin/currency-conversion",
            json=payload,
            headers=super_admin_headers,
            timeout=30,
        )
        if response.status_code != 200:
            pytest.skip(f"Unable to seed temporary standard rates: {response.status_code} {response.text[:300]}")
        body = response.json()
        config = body.get("config") or {}
        assert config.get("conversion_method") == "standard"
        assert config.get("source_currency") == "INR"
        created_ids.append(config.get("id"))

    yield {
        "yearly": yearly_payload,
        "monthly": monthly_payload,
        "created_ids": created_ids,
    }

    for config_id in created_ids:
        if not config_id:
            continue
        requests.delete(
            f"{API}/super-admin/currency-conversion/{config_id}",
            headers=super_admin_headers,
            timeout=30,
        )


# Super-admin currency conversion endpoints
def test_super_admin_currency_conversions_loads(super_admin_headers):
    response = requests.get(f"{API}/super-admin/currency-conversions", headers=super_admin_headers, timeout=30)
    assert response.status_code == 200, response.text[:300]
    data = response.json()
    assert isinstance(data, list)
    if data:
        sample = data[0]
        assert "year_applicable" in sample
        assert "conversion_method" in sample or "purchase_parity" in sample


def test_super_admin_create_duplicate_identity_includes_method_and_month(super_admin_headers, temp_standard_rates):
    duplicate_monthly = dict(temp_standard_rates["monthly"])
    duplicate_monthly["source"] = "PYTEST_TEMP_DUPLICATE"
    duplicate_response = requests.post(
        f"{API}/super-admin/currency-conversion",
        json=duplicate_monthly,
        headers=super_admin_headers,
        timeout=30,
    )
    assert duplicate_response.status_code == 400, duplicate_response.text[:300]
    detail = duplicate_response.json().get("detail", "")
    assert "already exists" in detail.lower()
    assert "standard" in detail.lower()


# Scope 3 bulk template contract for spend-basis columns
def test_scope3_bulk_template_has_spend_currency_conversion_columns(admin_headers):
    response = requests.get(
        f"{API}/bulk-upload/scope3/template/download",
        headers=admin_headers,
        timeout=60,
    )
    assert response.status_code == 200, response.text[:300]
    workbook = load_workbook(BytesIO(response.content))

    c1_sheet_name = next((name for name in workbook.sheetnames if name.strip().upper().startswith("C1")), None)
    assert c1_sheet_name, f"No C1 worksheet found. Available sheets: {workbook.sheetnames}"
    c1_sheet = workbook[c1_sheet_name]
    headers = [cell.value for cell in c1_sheet[1] if cell.value]

    assert "Spent Currency" in headers
    assert "Currency Conversion Method" in headers
    assert "Exchange Rate (Override)" in headers


# Calc engine formula resolution checks for spend-basis method branching
def test_execute_by_category_defaults_to_ppp_formula_when_method_missing(admin_headers, spend_basis_category_id):
    payload = {
        "category_id": spend_basis_category_id,
        "decision_inputs": {
            "calculation_method_scope3": "spend_basis",
        },
        "inputs": {
            "spent_value": {"value": 1000.0, "unit": "INR"},
        },
        "context": {
            "reporting_period": "2025-05",
        },
        "user_overrides": {
            "emission_factor": {"value": 1.0, "unit": "", "source_name": "pytest"},
            "ppp": {"value": 1.0, "unit": "", "source_name": "pytest"},
            "inflation_rate": {"value": 1.0, "unit": "", "source_name": "pytest"},
        },
        "dry_run": True,
    }
    response = requests.post(f"{API}/calc-engine/execute-by-category", json=payload, headers=admin_headers, timeout=60)
    assert response.status_code == 200, response.text[:500]
    data = response.json()
    resolved = (data.get("resolved_formula") or {}).get("id")
    assert resolved == PPP_FORMULA_ID
    decision_path = json.dumps(data.get("decision_path", []))
    assert "ppp_inflation" in decision_path


def test_execute_by_category_standard_prefers_monthly_rate_over_yearly(
    admin_headers,
    spend_basis_category_id,
    temp_standard_rates,
):
    payload = {
        "category_id": spend_basis_category_id,
        "decision_inputs": {
            "calculation_method_scope3": "spend_basis",
            "spend_currency_conversion_method": "standard",
        },
        "inputs": {
            "spent_value": {"value": 90000.0, "unit": "INR"},
        },
        "context": {
            "reporting_period": "2025-05",
        },
        "user_overrides": {
            "emission_factor": {"value": 1.0, "unit": "", "source_name": "pytest"},
        },
        "dry_run": True,
    }
    response = requests.post(f"{API}/calc-engine/execute-by-category", json=payload, headers=admin_headers, timeout=60)
    assert response.status_code == 200, response.text[:500]
    data = response.json()

    resolved = (data.get("resolved_formula") or {}).get("id")
    assert resolved == STANDARD_FORMULA_ID

    decision_path = json.dumps(data.get("decision_path", []))
    assert "standard" in decision_path

    co2e_output = (((data.get("outputs") or {}).get("co2e") or {}).get("value"))
    assert isinstance(co2e_output, (int, float))
    assert co2e_output == pytest.approx(1.0, rel=1e-6)

    body_text = json.dumps(data)
    assert "90" in body_text
    assert "80" not in body_text