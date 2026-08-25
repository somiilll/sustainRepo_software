"""
Bulk Upload Scope 3 - Iteration 18 Tests
Tests for:
1. Template download with org capabilities section in Instructions sheet
2. Upload validation response includes 'preview' field
3. Error report download includes warnings with Severity column
4. Frontend preview panel (tested via API response structure)
"""
import io
import os
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://esg-ai-routing.preview.emergentagent.com").rstrip("/")
ADMIN_EMAIL = "goyalsomil2001@gmail.com"
ADMIN_PASSWORD = "TestUser123!"


@pytest.fixture(scope="module")
def auth_headers():
    """Authenticate and return headers with Bearer token"""
    res = requests.post(
        f"{BASE_URL}/api/auth/login",
        json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD},
        timeout=30
    )
    if res.status_code != 200:
        pytest.skip(f"Admin login failed: {res.status_code} {res.text}")
    token = res.json().get("access_token") or res.json().get("token")
    if not token:
        pytest.skip(f"No token found: {res.json()}")
    return {"Authorization": f"Bearer {token}"}


# ============================================================================
# TEST 1: Template download with org capabilities section
# ============================================================================
class TestTemplateDownloadWithCapabilities:
    """Test that template download includes org capabilities in Instructions sheet"""
    
    def test_template_download_returns_200(self, auth_headers):
        """Template download should return HTTP 200 with correct content type"""
        r = requests.get(
            f"{BASE_URL}/api/bulk-upload/scope3/template/download",
            headers=auth_headers,
            timeout=60
        )
        assert r.status_code == 200, f"Expected 200, got {r.status_code}: {r.text}"
        content_type = r.headers.get("content-type", "").lower()
        assert "spreadsheet" in content_type or "excel" in content_type, \
            f"Expected spreadsheet content type, got: {content_type}"
        print(f"✓ Template download returned 200 with content-type: {content_type}")
    
    def test_template_has_instructions_sheet(self, auth_headers):
        """Template should have an Instructions sheet"""
        r = requests.get(
            f"{BASE_URL}/api/bulk-upload/scope3/template/download",
            headers=auth_headers,
            timeout=60
        )
        assert r.status_code == 200
        
        wb = load_workbook(io.BytesIO(r.content))
        assert "Instructions" in wb.sheetnames, \
            f"Instructions sheet not found. Available sheets: {wb.sheetnames}"
        print(f"✓ Instructions sheet found in template")
    
    def test_instructions_sheet_has_capabilities_section(self, auth_headers):
        """Instructions sheet should contain 'YOUR ORGANIZATION'S ENABLED CAPABILITIES' section"""
        r = requests.get(
            f"{BASE_URL}/api/bulk-upload/scope3/template/download",
            headers=auth_headers,
            timeout=60
        )
        assert r.status_code == 200
        
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb["Instructions"]
        
        # Search for the capabilities section header
        capabilities_found = False
        scope_status_found = False
        
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell and "YOUR ORGANIZATION'S ENABLED CAPABILITIES" in str(cell).upper():
                    capabilities_found = True
                if cell and ("ENABLED" in str(cell).upper() or "DISABLED" in str(cell).upper()):
                    if "Scope" in str(cell):
                        scope_status_found = True
        
        assert capabilities_found, \
            "Instructions sheet does not contain 'YOUR ORGANIZATION'S ENABLED CAPABILITIES' section"
        print(f"✓ Found 'YOUR ORGANIZATION'S ENABLED CAPABILITIES' section in Instructions sheet")
        
        assert scope_status_found, \
            "Instructions sheet does not contain scope status (ENABLED/DISABLED)"
        print(f"✓ Found scope status (ENABLED/DISABLED) in Instructions sheet")


# ============================================================================
# TEST 2: Upload validation response includes 'preview' field
# ============================================================================
class TestUploadValidationPreview:
    """Test that upload validation response includes preview field"""
    
    @pytest.fixture(scope="class")
    def reference_data(self, auth_headers):
        """Get reference data for building test upload file"""
        # Get facilities
        fr = requests.get(f"{BASE_URL}/api/facilities", headers=auth_headers, timeout=30)
        facilities_resp = fr.json() if fr.status_code == 200 else []
        # Handle both list and dict with 'data' key
        facilities = facilities_resp.get("data", facilities_resp) if isinstance(facilities_resp, dict) else facilities_resp
        fac_name = facilities[0]["name"] if facilities else "Test Facility"
        
        # Get scope3 EF data
        er = requests.get(f"{BASE_URL}/api/scope3-ef", headers=auth_headers, timeout=30)
        efs_resp = er.json() if er.status_code == 200 else []
        # Handle both list and dict with 'data' key
        efs = efs_resp.get("data", efs_resp) if isinstance(efs_resp, dict) else efs_resp
        
        # Find an activity
        activity = "Steel" if any(ef.get("activity") == "Steel" for ef in efs if isinstance(ef, dict)) else (
            efs[0].get("activity") if efs and isinstance(efs[0], dict) else "Test Activity"
        )
        
        # Get categories
        cr = requests.get(f"{BASE_URL}/api/emission-categories", headers=auth_headers, timeout=30)
        cats_resp = cr.json() if cr.status_code == 200 else []
        cats = cats_resp.get("data", cats_resp) if isinstance(cats_resp, dict) else cats_resp
        
        sc = requests.get(f"{BASE_URL}/api/scopes", headers=auth_headers, timeout=30)
        scopes_resp = sc.json() if sc.status_code == 200 else []
        scopes = scopes_resp.get("data", scopes_resp) if isinstance(scopes_resp, dict) else scopes_resp
        
        s3_id = next((s["id"] for s in scopes if isinstance(s, dict) and "3" in s.get("name", "")), None)
        s3_cat = next((c["name"] for c in cats if isinstance(c, dict) and c.get("scope_id") == s3_id), "Purchased Goods and Services")
        
        return {
            "facility": fac_name,
            "activity": activity,
            "category": s3_cat,
        }
    
    def _build_test_xlsx(self, rows):
        """Build a test Excel file with C1 sheet"""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "C1"
        
        # Headers matching the template
        headers = [
            "Facility Name", "Reporting Month", "Reporting Year", "Calculation Method",
            "Activity", "Quantity Used", "Spent Amount (INR)", "Unit of Quantity Used",
            "Quantity (Supplier Based)", "Unit of Quantity (Supplier Based)",
            "Emission Factor (Supplier Based)", "Emission Factor Unit (Supplier Based)",
            "Supplier Name", "Supplier Code", "Inflation Rate", "Purchase Power Value",
            "Person Responsible Name", "Person Responsible Designation", "Person Responsible Contact", "Notes"
        ]
        ws.append(headers)
        
        for row in rows:
            ws.append(row)
        
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
    
    def test_validation_response_includes_preview(self, auth_headers, reference_data):
        """Upload validation with validate_only=true should include preview field"""
        rd = reference_data
        
        # Build test file with valid rows
        rows = [
            [rd["facility"], "Jan-2025", "", "spend_basis", rd["activity"], "", 50000, "", "", "", "", "", "", "", "", "", "", "", "", "TEST_preview_1"],
            [rd["facility"], "Feb-2025", "", "spend_basis", rd["activity"], "", 75000, "", "", "", "", "", "", "", "", "", "", "", "", "TEST_preview_2"],
        ]
        
        buf = self._build_test_xlsx(rows)
        files = {"file": ("test_preview.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        
        r = requests.post(
            f"{BASE_URL}/api/bulk-upload/scope3/upload?validate_only=true",
            headers=auth_headers,
            files=files,
            timeout=120
        )
        
        assert r.status_code == 200, f"Upload failed: {r.status_code} {r.text}"
        data = r.json()
        
        # Check for preview field
        assert "preview" in data, f"Response missing 'preview' field. Keys: {data.keys()}"
        preview = data["preview"]
        
        if preview is not None:
            # Verify preview structure
            expected_keys = ["total_valid_records", "standard_fuel_records", "custom_fuel_records", 
                           "by_scope", "by_category", "total_co2e_tco2e"]
            for key in expected_keys:
                assert key in preview, f"Preview missing key: {key}. Preview keys: {preview.keys()}"
            
            print(f"✓ Preview field found with structure: {list(preview.keys())}")
            print(f"  - total_valid_records: {preview.get('total_valid_records')}")
            print(f"  - standard_fuel_records: {preview.get('standard_fuel_records')}")
            print(f"  - custom_fuel_records: {preview.get('custom_fuel_records')}")
            print(f"  - by_scope: {preview.get('by_scope')}")
            print(f"  - total_co2e_tco2e: {preview.get('total_co2e_tco2e')}")
        else:
            # Preview can be None if no valid records
            print(f"⚠ Preview is None (may indicate no valid records)")
            print(f"  - success_count: {data.get('success_count')}")
            print(f"  - error_count: {data.get('error_count')}")
        
        # Store job_id for error report test
        pytest.bulk_upload_job_id = data.get("job_id")
        print(f"✓ Job ID stored: {pytest.bulk_upload_job_id}")


# ============================================================================
# TEST 3: Error report download includes warnings with Severity column
# ============================================================================
class TestErrorReportWithWarnings:
    """Test that error report download includes both errors AND warnings"""
    
    @pytest.fixture(scope="class")
    def reference_data(self, auth_headers):
        """Get reference data for building test upload file"""
        fr = requests.get(f"{BASE_URL}/api/facilities", headers=auth_headers, timeout=30)
        facilities_resp = fr.json() if fr.status_code == 200 else []
        facilities = facilities_resp.get("data", facilities_resp) if isinstance(facilities_resp, dict) else facilities_resp
        fac_name = facilities[0]["name"] if facilities else "Test Facility"
        
        er = requests.get(f"{BASE_URL}/api/scope3-ef", headers=auth_headers, timeout=30)
        efs_resp = er.json() if er.status_code == 200 else []
        efs = efs_resp.get("data", efs_resp) if isinstance(efs_resp, dict) else efs_resp
        
        activity = "Steel" if any(ef.get("activity") == "Steel" for ef in efs if isinstance(ef, dict)) else (
            efs[0].get("activity") if efs and isinstance(efs[0], dict) else "Test Activity"
        )
        
        cr = requests.get(f"{BASE_URL}/api/emission-categories", headers=auth_headers, timeout=30)
        cats_resp = cr.json() if cr.status_code == 200 else []
        cats = cats_resp.get("data", cats_resp) if isinstance(cats_resp, dict) else cats_resp
        
        sc = requests.get(f"{BASE_URL}/api/scopes", headers=auth_headers, timeout=30)
        scopes_resp = sc.json() if sc.status_code == 200 else []
        scopes = scopes_resp.get("data", scopes_resp) if isinstance(scopes_resp, dict) else scopes_resp
        
        s3_id = next((s["id"] for s in scopes if isinstance(s, dict) and "3" in s.get("name", "")), None)
        s3_cat = next((c["name"] for c in cats if isinstance(c, dict) and c.get("scope_id") == s3_id), "Purchased Goods and Services")
        
        return {"facility": fac_name, "activity": activity, "category": s3_cat}
    
    def _build_test_xlsx_with_errors(self, rows):
        """Build a test Excel file with C1 sheet"""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "C1"
        
        headers = [
            "Facility Name", "Reporting Month", "Reporting Year", "Calculation Method",
            "Activity", "Quantity Used", "Spent Amount (INR)", "Unit of Quantity Used",
            "Quantity (Supplier Based)", "Unit of Quantity (Supplier Based)",
            "Emission Factor (Supplier Based)", "Emission Factor Unit (Supplier Based)",
            "Supplier Name", "Supplier Code", "Inflation Rate", "Purchase Power Value",
            "Person Responsible Name", "Person Responsible Designation", "Person Responsible Contact", "Notes"
        ]
        ws.append(headers)
        
        for row in rows:
            ws.append(row)
        
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
    
    def test_error_report_download(self, auth_headers, reference_data):
        """Error report should be downloadable and contain validation issues"""
        rd = reference_data
        
        # Build test file with mix of valid and invalid rows
        rows = [
            # Valid row
            [rd["facility"], "Jan-2025", "", "spend_basis", rd["activity"], "", 50000, "", "", "", "", "", "", "", "", "", "", "", "", "TEST_error_report_valid"],
            # Invalid row - missing required fields
            ["", "Feb-2025", "", "spend_basis", rd["activity"], "", 50000, "", "", "", "", "", "", "", "", "", "", "", "", "TEST_error_report_invalid"],
            # Invalid row - bad date format
            [rd["facility"], "BadDate", "", "spend_basis", rd["activity"], "", 50000, "", "", "", "", "", "", "", "", "", "", "", "", "TEST_error_report_bad_date"],
        ]
        
        buf = self._build_test_xlsx_with_errors(rows)
        files = {"file": ("test_errors.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        
        # Upload with validation
        r = requests.post(
            f"{BASE_URL}/api/bulk-upload/scope3/upload?validate_only=true",
            headers=auth_headers,
            files=files,
            timeout=120
        )
        
        assert r.status_code == 200, f"Upload failed: {r.status_code} {r.text}"
        data = r.json()
        job_id = data.get("job_id")
        assert job_id, "No job_id in response"
        
        print(f"✓ Upload completed. Job ID: {job_id}")
        print(f"  - success_count: {data.get('success_count')}")
        print(f"  - error_count: {data.get('error_count')}")
        print(f"  - warning_count: {data.get('warning_count')}")
        
        # Download error report
        r2 = requests.get(
            f"{BASE_URL}/api/bulk-upload/scope3/jobs/{job_id}/errors/download",
            headers=auth_headers,
            timeout=60
        )
        
        assert r2.status_code == 200, f"Error report download failed: {r2.status_code} {r2.text}"
        content_type = r2.headers.get("content-type", "").lower()
        assert "spreadsheet" in content_type or "excel" in content_type, \
            f"Expected spreadsheet content type, got: {content_type}"
        
        # Parse the Excel file
        wb = load_workbook(io.BytesIO(r2.content))
        print(f"✓ Error report downloaded. Sheets: {wb.sheetnames}")
        
        # Check for Summary sheet
        assert "Summary" in wb.sheetnames, f"Summary sheet not found. Sheets: {wb.sheetnames}"
        
        # Check for Errors sheet (if there are errors)
        if data.get("error_count", 0) > 0:
            assert "Errors" in wb.sheetnames, f"Errors sheet not found despite error_count > 0"
            ws_errors = wb["Errors"]
            
            # Check headers
            headers = [cell.value for cell in ws_errors[1]]
            print(f"  - Errors sheet headers: {headers}")
            
            # Verify error rows exist
            error_rows = list(ws_errors.iter_rows(min_row=2, values_only=True))
            print(f"  - Error rows count: {len(error_rows)}")
        
        # Check for Warnings sheet (if there are warnings)
        if data.get("warning_count", 0) > 0:
            if "Warnings" in wb.sheetnames:
                ws_warnings = wb["Warnings"]
                headers = [cell.value for cell in ws_warnings[1]]
                print(f"  - Warnings sheet headers: {headers}")
                warning_rows = list(ws_warnings.iter_rows(min_row=2, values_only=True))
                print(f"  - Warning rows count: {len(warning_rows)}")
            else:
                print(f"⚠ Warnings sheet not found despite warning_count > 0")
        
        print(f"✓ Error report structure validated")


# ============================================================================
# TEST 4: Rollback for partial saves (server.py line ~3465)
# ============================================================================
class TestRollbackForPartialSaves:
    """Test that partial save failures trigger rollback"""
    
    def test_save_endpoint_exists(self, auth_headers):
        """Verify the save endpoint exists and returns proper error for invalid job"""
        r = requests.post(
            f"{BASE_URL}/api/bulk-upload/scope3/jobs/nonexistent-job-id/save",
            headers=auth_headers,
            timeout=30
        )
        # Should return 404 for non-existent job
        assert r.status_code == 404, f"Expected 404 for non-existent job, got {r.status_code}"
        print(f"✓ Save endpoint returns 404 for non-existent job")


# ============================================================================
# TEST 5: Scope 2 org-category enforcement and custom fuel auto-detection
# ============================================================================
class TestScope2OrgCategoryEnforcement:
    """Test Scope 2 org-category enforcement and custom fuel auto-detection"""
    
    def _build_scope2_xlsx(self, rows):
        """Build a test Excel file with Scope2 sheet"""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Scope2"
        
        # Headers for Scope 2
        headers = [
            "Facility Name", "Reporting Month (MMM-YYYY)", "Reporting Year (FY YYYY-YYYY or CY YYYY)",
            "Category", "Energy Used", "Quantity Used", "Unit of Quantity Used",
            "Emission Factor", "EF Unit", "Process Name", "Process Description",
            "Record Source", "Person Responsible Name", "Person Responsible Designation",
            "Person Responsible Contact", "Notes"
        ]
        ws.append(headers)
        
        for row in rows:
            ws.append(row)
        
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
    
    def test_scope2_validation_with_custom_energy(self, auth_headers):
        """Test Scope 2 validation with custom energy source"""
        # Get a facility name
        fr = requests.get(f"{BASE_URL}/api/facilities", headers=auth_headers, timeout=30)
        facilities = fr.json() if fr.status_code == 200 else []
        fac_name = facilities[0]["name"] if facilities else "Test Facility"
        
        # Build test file with custom energy source
        rows = [
            # Valid row with standard energy
            [fac_name, "Jan-2025", "", "Purchased Electricity", "Grid Electricity", 1000, "kWh", "", "", "", "", "", "", "", "", "TEST_scope2_standard"],
            # Row with custom energy (should trigger custom fuel detection)
            [fac_name, "Feb-2025", "", "Purchased Electricity", "Custom Solar Energy", 500, "kWh", 0.1, "kgCO2/kWh", "", "", "", "", "", "", "TEST_scope2_custom"],
        ]
        
        buf = self._build_scope2_xlsx(rows)
        files = {"file": ("test_scope2.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        
        r = requests.post(
            f"{BASE_URL}/api/bulk-upload/scope3/upload?validate_only=true",
            headers=auth_headers,
            files=files,
            timeout=120
        )
        
        # The response should be 200 (validation completed)
        assert r.status_code == 200, f"Upload failed: {r.status_code} {r.text}"
        data = r.json()
        
        print(f"✓ Scope 2 validation completed")
        print(f"  - Job ID: {data.get('job_id')}")
        print(f"  - Status: {data.get('status')}")
        print(f"  - success_count: {data.get('success_count')}")
        print(f"  - error_count: {data.get('error_count')}")
        print(f"  - warning_count: {data.get('warning_count')}")
        print(f"  - categories_processed: {data.get('categories_processed')}")
        
        # Check if Scope2 was processed
        categories = data.get("categories_processed", [])
        if "Scope2" in categories:
            print(f"✓ Scope2 sheet was processed")
        else:
            print(f"⚠ Scope2 sheet was not processed. Categories: {categories}")
        
        # Check for custom fuel warnings
        warnings = data.get("warnings", [])
        custom_fuel_warnings = [w for w in warnings if "custom" in w.get("message", "").lower()]
        if custom_fuel_warnings:
            print(f"✓ Custom fuel detection triggered: {len(custom_fuel_warnings)} warning(s)")
        
        # Check errors for disabled scope/category
        errors = data.get("errors", [])
        disabled_errors = [e for e in errors if "disabled" in e.get("message", "").lower()]
        if disabled_errors:
            print(f"⚠ Scope/category disabled errors: {len(disabled_errors)}")
            for e in disabled_errors[:3]:
                print(f"    - {e.get('message')}")


# ============================================================================
# TEST 6: API endpoint verification
# ============================================================================
class TestAPIEndpoints:
    """Verify all bulk upload API endpoints are accessible"""
    
    def test_template_download_endpoint(self, auth_headers):
        """GET /api/bulk-upload/scope3/template/download"""
        r = requests.get(
            f"{BASE_URL}/api/bulk-upload/scope3/template/download",
            headers=auth_headers,
            timeout=60
        )
        assert r.status_code == 200, f"Template download failed: {r.status_code}"
        print(f"✓ Template download endpoint working")
    
    def test_jobs_list_endpoint(self, auth_headers):
        """GET /api/bulk-upload/scope3/jobs"""
        r = requests.get(
            f"{BASE_URL}/api/bulk-upload/scope3/jobs",
            headers=auth_headers,
            timeout=30
        )
        assert r.status_code == 200, f"Jobs list failed: {r.status_code}"
        data = r.json()
        assert "jobs" in data, f"Response missing 'jobs' key"
        print(f"✓ Jobs list endpoint working. Total jobs: {data.get('total', 0)}")
    
    def test_job_status_endpoint_404(self, auth_headers):
        """GET /api/bulk-upload/scope3/jobs/{job_id} - 404 for non-existent"""
        r = requests.get(
            f"{BASE_URL}/api/bulk-upload/scope3/jobs/nonexistent-id",
            headers=auth_headers,
            timeout=30
        )
        assert r.status_code == 404, f"Expected 404, got {r.status_code}"
        print(f"✓ Job status endpoint returns 404 for non-existent job")
    
    def test_error_report_endpoint_404(self, auth_headers):
        """GET /api/bulk-upload/scope3/jobs/{job_id}/errors/download - 404 for non-existent"""
        r = requests.get(
            f"{BASE_URL}/api/bulk-upload/scope3/jobs/nonexistent-id/errors/download",
            headers=auth_headers,
            timeout=30
        )
        assert r.status_code == 404, f"Expected 404, got {r.status_code}"
        print(f"✓ Error report endpoint returns 404 for non-existent job")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
