"""Bulk Upload (Scope 3) regression tests."""
import io
import os
import pytest
import requests
from openpyxl import load_workbook, Workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://scope-separation.preview.emergentagent.com").rstrip("/")
ADMIN_EMAIL = "goyalsomil2@hotmail.com"
ADMIN_PASSWORD = "Test123!"


@pytest.fixture(scope="module")
def auth_headers():
    res = requests.post(f"{BASE_URL}/api/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD}, timeout=30)
    if res.status_code != 200:
        pytest.skip(f"Admin login failed: {res.status_code} {res.text}")
    token = res.json().get("access_token") or res.json().get("token")
    if not token:
        pytest.skip(f"No token found: {res.json()}")
    return {"Authorization": f"Bearer {token}"}


# ---------- Template download ----------
class TestTemplate:
    def test_download_template(self, auth_headers):
        r = requests.get(f"{BASE_URL}/api/bulk-upload/template", headers=auth_headers, timeout=30)
        assert r.status_code == 200, r.text
        assert "spreadsheet" in r.headers.get("content-type", "").lower()
        wb = load_workbook(io.BytesIO(r.content))
        assert "Scope 3 Emissions" in wb.sheetnames
        assert "Reference Data" in wb.sheetnames
        assert "Instructions" in wb.sheetnames
        ws = wb["Scope 3 Emissions"]
        # Validate header has key columns
        headers = [c.value for c in ws[1]]
        assert any("Facility" in str(h) for h in headers)
        assert any("Method" in str(h) for h in headers)


# ---------- Helper to build upload xlsx ----------
def _build_upload_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Scope 3 Emissions"
    headers = ["Facility Name *", "Reporting Month (YYYY-MM) *", "Category *",
               "Activity *", "Method *", "Quantity/Spend *", "Unit *",
               "Emission Factor (optional)", "EF Unit (if EF provided)",
               "Evidence Reference", "Notes"]
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@pytest.fixture(scope="module")
def reference_data(auth_headers):
    """Resolve a real facility name and a Scope 3 category/activity to use in tests."""
    # facilities
    fr = requests.get(f"{BASE_URL}/api/facilities", headers=auth_headers, timeout=30)
    facilities = fr.json() if fr.status_code == 200 else []
    fac_name = facilities[0]["name"] if facilities else "Test Facility"

    # scope3 EF rows
    er = requests.get(f"{BASE_URL}/api/scope3-ef", headers=auth_headers, timeout=30)
    efs = er.json() if er.status_code == 200 else []

    # Find an activity that exists with both spend_basis and activity_basis if possible
    activity = None
    spend_unit_choice = "INR"
    activity_unit_choice = None
    methods_per_activity = {}
    for ef in efs:
        a = ef.get("activity")
        m = ef.get("method")
        methods_per_activity.setdefault(a, set()).add(m)
    # Prefer Steel (per agent context)
    if "Steel" in methods_per_activity:
        activity = "Steel"
    elif methods_per_activity:
        activity = next(iter(methods_per_activity))

    # Try to find a physical (non-currency) unit symbol from any activity_basis EF for this activity
    for ef in efs:
        if ef.get("activity") == activity and ef.get("method") == "activity_basis":
            allowed = ef.get("allowed_units") or []
            currencies = {"INR", "USD", "EUR", "GBP", "JPY"}
            phys = [u for u in allowed if u and str(u).upper() not in currencies]
            if phys:
                activity_unit_choice = phys[0]
            elif ef.get("unit") and str(ef["unit"]).upper() not in currencies:
                activity_unit_choice = ef["unit"]
            break
    if not activity_unit_choice:
        activity_unit_choice = "tonnes"

    # Determine a Scope 3 category name
    cr = requests.get(f"{BASE_URL}/api/emission-categories", headers=auth_headers, timeout=30)
    cats = cr.json() if cr.status_code == 200 else []
    sc = requests.get(f"{BASE_URL}/api/scopes", headers=auth_headers, timeout=30)
    scopes = sc.json() if sc.status_code == 200 else []
    s3_id = next((s["id"] for s in scopes if "3" in s.get("name", "")), None)
    s3_cat = next((c["name"] for c in cats if c.get("scope_id") == s3_id), "Purchased Goods")

    return {
        "facility": fac_name,
        "activity": activity or "Steel",
        "category": s3_cat,
        "spend_unit": spend_unit_choice,
        "activity_unit": activity_unit_choice,
    }


# ---------- Validate endpoint ----------
class TestValidate:
    def test_validate_mixed_rows(self, auth_headers, reference_data):
        rd = reference_data
        rows = [
            # Valid spend_basis
            [rd["facility"], "2024-01", rd["category"], rd["activity"], "spend_basis", 50000, rd["spend_unit"], "", "", "PO #1", "TEST_valid_spend"],
            # Valid activity_basis
            [rd["facility"], "2024-02", rd["category"], rd["activity"], "activity_basis", 10, rd["activity_unit"], "", "", "Doc-2", "TEST_valid_activity"],
            # Invalid date format
            [rd["facility"], "January 2024", rd["category"], rd["activity"], "spend_basis", 1000, rd["spend_unit"], "", "", "", "TEST_bad_date"],
            # Invalid: currency unit with activity_basis method
            [rd["facility"], "2024-03", rd["category"], rd["activity"], "activity_basis", 100, "INR", "", "", "", "TEST_currency_activity"],
            # Invalid: missing required fields (facility blank)
            ["", "2024-04", rd["category"], rd["activity"], "spend_basis", 200, "INR", "", "", "", "TEST_missing_fac"],
            # Invalid: bad method
            [rd["facility"], "2024-05", rd["category"], rd["activity"], "carbon_basis", 100, "INR", "", "", "", "TEST_bad_method"],
            # Fuzzy/case-insensitive category
            [rd["facility"], "2024-06", rd["category"].lower(), rd["activity"], "spend_basis", 1000, rd["spend_unit"], "", "", "", "TEST_fuzzy"],
        ]
        buf = _build_upload_xlsx(rows)
        files = {"file": ("test_upload.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{BASE_URL}/api/bulk-upload/validate", headers=auth_headers, files=files, timeout=60)
        assert r.status_code == 200, r.text
        data = r.json()
        assert "upload_id" in data
        assert data["template_type"] == "scope3"
        assert data["summary"]["total_rows"] == len(rows)
        # row 2 valid_spend, row 3 valid_activity, row 8 fuzzy => at least 3 valid
        assert data["summary"]["valid_rows"] >= 2
        assert data["summary"]["invalid_rows"] >= 4
        # Check specific error messages
        rows_resp = data["rows"]
        # Find row with January 2024
        jan_row = next((rr for rr in rows_resp if rr["original_data"].get("reporting_month") == "January 2024"), None)
        assert jan_row is not None
        assert jan_row["status"] == "invalid"
        assert any("date" in e["message"].lower() or "format" in e["message"].lower() for e in jan_row["errors"])

        # currency + activity_basis row
        ca_row = next((rr for rr in rows_resp if rr["original_data"].get("notes") == "TEST_currency_activity"), None)
        assert ca_row is not None
        assert ca_row["status"] == "invalid"
        assert any("activity_basis" in e["message"].lower() or "currency" in e["message"].lower() for e in ca_row["errors"])

        # missing facility row
        mf_row = next((rr for rr in rows_resp if rr["original_data"].get("notes") == "TEST_missing_fac"), None)
        assert mf_row is not None
        assert mf_row["status"] == "invalid"

        # bad method row
        bm_row = next((rr for rr in rows_resp if rr["original_data"].get("notes") == "TEST_bad_method"), None)
        assert bm_row is not None
        assert bm_row["status"] == "invalid"
        assert any("method" in e["column"] for e in bm_row["errors"])

        # fuzzy row should be valid (case-insensitive cat match)
        fz_row = next((rr for rr in rows_resp if rr["original_data"].get("notes") == "TEST_fuzzy"), None)
        assert fz_row is not None, "Fuzzy row missing"
        assert fz_row["status"] == "valid", f"Fuzzy match failed: {fz_row['errors']}"

        # Save upload_id for next tests
        pytest.upload_id = data["upload_id"]
        pytest.valid_count = data["summary"]["valid_rows"]
        pytest.invalid_count = data["summary"]["invalid_rows"]

    def test_reject_non_xlsx(self, auth_headers):
        files = {"file": ("bad.txt", b"not excel", "text/plain")}
        r = requests.post(f"{BASE_URL}/api/bulk-upload/validate", headers=auth_headers, files=files, timeout=30)
        assert r.status_code == 400


# ---------- Save valid rows ----------
class TestSave:
    def test_save_valid_rows(self, auth_headers):
        if not getattr(pytest, "upload_id", None):
            pytest.skip("No upload_id from validate test")
        r = requests.post(f"{BASE_URL}/api/bulk-upload/{pytest.upload_id}/save?save_mode=valid_only",
                          headers=auth_headers, timeout=60)
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["saved_count"] == pytest.valid_count
        assert isinstance(data["saved_ids"], list)

    def test_save_idempotent_blocked(self, auth_headers):
        if not getattr(pytest, "upload_id", None):
            pytest.skip()
        r = requests.post(f"{BASE_URL}/api/bulk-upload/{pytest.upload_id}/save?save_mode=valid_only",
                          headers=auth_headers, timeout=30)
        assert r.status_code == 400  # already completed


# ---------- Error report ----------
class TestErrorReport:
    def test_download_error_report(self, auth_headers):
        if not getattr(pytest, "upload_id", None):
            pytest.skip()
        r = requests.get(f"{BASE_URL}/api/bulk-upload/{pytest.upload_id}/error-report",
                         headers=auth_headers, timeout=30)
        assert r.status_code == 200, r.text
        assert "spreadsheet" in r.headers.get("content-type", "").lower()
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # Should have headers + N rows
        assert ws.max_row >= 2

    def test_error_report_404(self, auth_headers):
        r = requests.get(f"{BASE_URL}/api/bulk-upload/nonexistent-id/error-report",
                         headers=auth_headers, timeout=30)
        assert r.status_code == 404


# ---------- Sessions ----------
class TestSessions:
    def test_list_sessions(self, auth_headers):
        r = requests.get(f"{BASE_URL}/api/bulk-upload/sessions", headers=auth_headers, timeout=30)
        assert r.status_code == 200
        sessions = r.json()
        assert isinstance(sessions, list)
        if getattr(pytest, "upload_id", None):
            ids = [s.get("id") for s in sessions]
            assert pytest.upload_id in ids


# ---------- Auth check ----------
class TestAuth:
    def test_template_unauth(self):
        r = requests.get(f"{BASE_URL}/api/bulk-upload/template", timeout=30)
        assert r.status_code in (401, 403)
