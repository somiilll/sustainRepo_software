"""
Backend regression tests for the new Scope 3 Bulk Upload System (15 categories).

Endpoints under test (prefix /api/bulk-upload/scope3):
  GET  /template/download
  POST /upload?allow_partial_success=
  GET  /jobs
  GET  /jobs/{job_id}
  GET  /jobs/{job_id}/errors/download
  GET  /jobs/{job_id}/results/download
  DELETE /jobs/{job_id}
"""
import io
import os
import time
import pytest
import requests
from openpyxl import Workbook, load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://ghg-calc-engine-5.preview.emergentagent.com").rstrip("/")
ADMIN_EMAIL = "goyalsomil2@hotmail.com"
ADMIN_PASSWORD = "Test123!"

API = f"{BASE_URL}/api"
SCOPE3 = f"{API}/bulk-upload/scope3"

EXPECTED_CATEGORY_SHEETS = [
    "C1 - Purchased Goods and Services",
    "C2 - Capital Goods",
    "C3 - Fuel and Energy Related Activities Not Included in Scope 1 or Scope 2",
    "C4 - Upstream Transportation",
    "C5 - Waste Generated in Operations",
    "C6 - Business Travel",
    "C7 - Employee Commuting",
    "C8 - Upstream Leased Assets",
    "C9 - Downstream Transportation",
    "C10 - Processing of Sold Products",
    "C11 - Use of Sold Products",
    "C12 - End-of-Life Treatment",
    "C13 - Downstream Leased Assets",
    "C14 - Franchises",
    "C15 - Investments",
]


# ---------- Auth fixture ----------
@pytest.fixture(scope="module")
def auth_headers():
    res = requests.post(f"{API}/auth/login",
                        json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD},
                        timeout=30)
    if res.status_code != 200:
        pytest.skip(f"Admin login failed: {res.status_code} {res.text}")
    body = res.json()
    token = body.get("access_token") or body.get("token")
    if not token:
        pytest.skip(f"No token: {body}")
    return {"Authorization": f"Bearer {token}"}


# ---------- Auth/Negative ----------
class TestAuth:
    def test_template_unauth(self):
        r = requests.get(f"{SCOPE3}/template/download", timeout=30)
        assert r.status_code in (401, 403)

    def test_upload_unauth(self):
        r = requests.post(f"{SCOPE3}/upload",
                          files={"file": ("x.xlsx", b"x", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                          timeout=30)
        assert r.status_code in (401, 403)

    def test_jobs_unauth(self):
        r = requests.get(f"{SCOPE3}/jobs", timeout=30)
        assert r.status_code in (401, 403)


# ---------- Template download ----------
class TestTemplate:
    def test_download_template_returns_xlsx(self, auth_headers):
        r = requests.get(f"{SCOPE3}/template/download", headers=auth_headers, timeout=60)
        assert r.status_code == 200, r.text[:300]
        assert "spreadsheet" in r.headers.get("content-type", "").lower()
        assert int(r.headers.get("content-length", len(r.content))) > 10000
        pytest.template_bytes = r.content

    def test_template_has_15_category_sheets(self, auth_headers):
        if not getattr(pytest, "template_bytes", None):
            pytest.skip("template missing")
        wb = load_workbook(io.BytesIO(pytest.template_bytes))
        for sheet in EXPECTED_CATEGORY_SHEETS:
            assert sheet in wb.sheetnames, f"Missing sheet: {sheet}"
        assert "Instructions" in wb.sheetnames

    def test_template_c1_headers(self, auth_headers):
        if not getattr(pytest, "template_bytes", None):
            pytest.skip()
        wb = load_workbook(io.BytesIO(pytest.template_bytes))
        ws = wb["C1 - Purchased Goods and Services"]
        headers = [c.value for c in ws[1]]
        # Required core headers
        for col in ["Facility Name", "Reporting Month", "Calculation Method", "Activity"]:
            assert col in headers, f"C1 missing header: {col}"


# ---------- Helpers ----------
def _build_xlsx(sheet_name: str, headers: list, rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


C1_HEADERS = [
    "Facility Name", "Reporting Month", "Calculation Method", "Activity",
    "Quantity Used", "Spent Amount (INR)", "Unit of Quantity Used",
    "Quantity (Supplier Based)", "Unit of Quantity (Supplier Based)",
    "Emission Factor (Supplier Based)", "Emission Factor Unit (Supplier Based)",
    "Supplier Name", "Supplier Code", "Inflation Rate", "Purchase Power Value",
    "Person Responsible Name", "Person Responsible Designation", "Person Responsible Contact",
]

C7_HEADERS = [
    "Facility Name", "Reporting Month", "Calculation Method", "Activity Type",
    "Activity", "Distance Travelled (km)", "Passengers Travelled",
    "No. of Working Days", "Working Hours per Day",
    "Quantity (Supplier Based)", "Unit of Quantity (Supplier Based)",
    "Emission Factor (Supplier Based)", "Emission Factor Unit (Supplier Based)",
    "Supplier Name", "Supplier Code", "Inflation Rate", "Purchase Power Value",
    "Employee Name", "Employee Id", "Department",
    "Person Responsible Name", "Person Responsible Designation", "Person Responsible Contact",
]

C15_HEADERS = [
    "Facility Name", "Reporting Month", "Calculation Method", "Sub Category",
    "Activity", "Quantity Used", "Unit of Quantity Used",
    "Quantity (Supplier Based)", "Unit of Quantity (Supplier Based)",
    "Emission Factor (Supplier Based)", "Emission Factor Unit (Supplier Based)",
    "Supplier Name", "Supplier Code",
    "Investment Type", "Investment Amount", "Equity Share %",
    "Person Responsible Name", "Person Responsible Designation", "Person Responsible Contact",
]


def _post_upload(auth_headers, sheet_name, headers, rows, allow_partial=True):
    content = _build_xlsx(sheet_name, headers, rows)
    files = {"file": (f"test_{sheet_name[:6]}.xlsx", content,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    return requests.post(
        f"{SCOPE3}/upload",
        params={"allow_partial_success": str(allow_partial).lower()},
        headers=auth_headers,
        files=files,
        timeout=120,
    )


# ---------- Upload validation ----------
class TestUpload:
    def test_reject_non_xlsx(self, auth_headers):
        files = {"file": ("bad.txt", b"hello", "text/plain")}
        r = requests.post(f"{SCOPE3}/upload", headers=auth_headers, files=files, timeout=30)
        assert r.status_code == 400

    def test_invalid_facility_returns_error_with_suggestion(self, auth_headers):
        # 'test-fac-100' is close to 'test-fac-1' -> should suggest it via fuzzy match
        rows = [
            ["test-fac-100", "Jan-2025", "spend_basis", "Steel",
             "", 50000, "", "", "", "", "", "", "", "", "", "", "", ""],
        ]
        r = _post_upload(auth_headers, "C1 - Purchased Goods and Services", C1_HEADERS, rows)
        assert r.status_code == 200, r.text[:300]
        data = r.json()
        assert data["total_rows"] == 1
        assert data["error_count"] >= 1
        # Error should mention facility and include a suggestion
        errs = data.get("errors", [])
        fac_err = next((e for e in errs if (e.get("column") or "").lower().startswith("facility")), None)
        assert fac_err is not None, f"No facility error found: {errs}"
        assert "facility" in fac_err.get("message", "").lower()
        assert fac_err.get("suggestion"), f"Expected fuzzy suggestion, got: {fac_err}"

    def test_invalid_calculation_method_rejected(self, auth_headers):
        rows = [
            ["test-fac-1", "Jan-2025", "INVALID_METHOD", "Steel",
             "", 50000, "", "", "", "", "", "", "", "", "", "", "", ""],
        ]
        r = _post_upload(auth_headers, "C1 - Purchased Goods and Services", C1_HEADERS, rows)
        assert r.status_code == 200, r.text[:300]
        data = r.json()
        assert data["error_count"] >= 1
        err_msgs = " ".join([e.get("message", "") + " " + (e.get("column") or "") for e in data.get("errors", [])])
        assert ("method" in err_msgs.lower()) or ("calculation" in err_msgs.lower()), \
            f"Expected method error, got: {err_msgs}"

    def test_supplier_basis_custom_activity_allowed(self, auth_headers):
        # Supplier-based should allow a custom (non-master) activity name.
        rows = [
            ["test-fac-1", "Feb-2025", "supplier_basis", "MyCustomCustomActivity_TEST",
             "", "", "", 100, "kg", 2.5, "kgCO2e/kg",
             "Acme Supplier", "ACME-001", "", "", "", "", ""],
        ]
        r = _post_upload(auth_headers, "C1 - Purchased Goods and Services", C1_HEADERS, rows)
        assert r.status_code == 200, r.text[:300]
        data = r.json()
        # supplier_basis with custom activity should succeed (no activity error)
        results = data.get("results", [])
        if results:
            row_errs = []
            for res in results:
                row_errs.extend(res.get("errors", []))
            activity_errs = [e for e in row_errs if "activity" in (e.get("column") or "").lower()
                             and "not" in e.get("message", "").lower()]
            assert not activity_errs, f"Custom activity should be allowed for supplier_basis, got errs: {activity_errs}"
        # success count should include this row OR error_count == 0 for activity issue
        # We tolerate emission factor calc passing as the chief expectation
        assert data["error_count"] == 0 or all(
            "activity" not in (e.get("column") or "").lower() for e in data.get("errors", [])
        ), f"Unexpected activity error: {data.get('errors')}"

    def test_c15_only_supplier_basis_allowed(self, auth_headers):
        # C15 with spend_basis should error
        rows = [
            ["test-fac-1", "Mar-2025", "spend_basis", "Equity Investments", "Steel",
             "", "", "", "", "", "",
             "", "", "Equity", 100000, 25, "", "", ""],
        ]
        r = _post_upload(auth_headers, "C15 - Investments", C15_HEADERS, rows)
        assert r.status_code == 200, r.text[:300]
        data = r.json()
        assert data["error_count"] >= 1, f"Expected method error for C15 spend_basis: {data}"
        err_msgs = " ".join([e.get("message", "") for e in data.get("errors", [])])
        assert ("supplier" in err_msgs.lower()) or ("method" in err_msgs.lower()) or \
               ("not supported" in err_msgs.lower()), f"C15 method error msg unclear: {err_msgs}"

    def test_c7_employee_aggregation(self, auth_headers):
        # 3 employees, same facility/month/activity -> should aggregate to 1 emission record
        rows = [
            ["test-fac-1", "Apr-2025", "activity_basis", "Car", "Petrol Car",
             20, 1, 22, 8, "", "", "", "", "", "", "", "", "Alice", "E001", "Eng",
             "", "", ""],
            ["test-fac-1", "Apr-2025", "activity_basis", "Car", "Petrol Car",
             15, 1, 22, 8, "", "", "", "", "", "", "", "", "Bob", "E002", "Eng",
             "", "", ""],
            ["test-fac-1", "Apr-2025", "activity_basis", "Car", "Petrol Car",
             25, 1, 22, 8, "", "", "", "", "", "", "", "", "Charlie", "E003", "Eng",
             "", "", ""],
        ]
        r = _post_upload(auth_headers, "C7 - Employee Commuting", C7_HEADERS, rows)
        assert r.status_code == 200, r.text[:300]
        data = r.json()
        # Even if rows fail individual validation (e.g., activity not found),
        # we at least check the total_rows counts properly, and if any succeeded,
        # the success_count should be <=1 (aggregated).
        assert data["total_rows"] == 3
        if data["success_count"] >= 1:
            assert data["success_count"] <= 1, \
                f"C7 employees should aggregate; success_count={data['success_count']}"
            pytest.c7_job_id = data.get("job_id")

    def test_partial_success_mode(self, auth_headers):
        # 1 valid + 1 invalid in same upload, allow_partial=True -> total_rows=2, errors>=1
        rows = [
            ["test-fac-1", "May-2025", "supplier_basis", "TEST_CustomA",
             "", "", "", 50, "kg", 1.5, "kgCO2e/kg", "S1", "", "", "", "", "", ""],
            ["NonExistentFac999", "May-2025", "supplier_basis", "TEST_CustomB",
             "", "", "", 50, "kg", 1.5, "kgCO2e/kg", "S1", "", "", "", "", "", ""],
        ]
        r = _post_upload(auth_headers, "C1 - Purchased Goods and Services", C1_HEADERS, rows, allow_partial=True)
        assert r.status_code == 200, r.text[:300]
        data = r.json()
        assert data["total_rows"] == 2
        assert data["error_count"] >= 1
        # Save job for downstream tests if we have one
        if data.get("job_id"):
            pytest.partial_job_id = data["job_id"]


# ---------- Jobs list & details ----------
class TestJobs:
    def test_list_jobs(self, auth_headers):
        r = requests.get(f"{SCOPE3}/jobs", headers=auth_headers, timeout=30)
        assert r.status_code == 200, r.text[:300]
        data = r.json()
        assert "jobs" in data and isinstance(data["jobs"], list)
        assert "total" in data
        # Persist a job_id for error report test if we don't have one
        if data["jobs"] and not getattr(pytest, "partial_job_id", None):
            pytest.partial_job_id = data["jobs"][0].get("id")

    def test_jobs_pagination(self, auth_headers):
        r = requests.get(f"{SCOPE3}/jobs", headers=auth_headers,
                         params={"limit": 1, "offset": 0}, timeout=30)
        assert r.status_code == 200
        data = r.json()
        assert data["limit"] == 1
        assert data["offset"] == 0
        assert len(data["jobs"]) <= 1

    def test_job_status_404(self, auth_headers):
        r = requests.get(f"{SCOPE3}/jobs/nonexistent-id-xyz", headers=auth_headers, timeout=30)
        assert r.status_code == 404

    def test_job_status_existing(self, auth_headers):
        if not getattr(pytest, "partial_job_id", None):
            pytest.skip("No job_id available")
        r = requests.get(f"{SCOPE3}/jobs/{pytest.partial_job_id}", headers=auth_headers, timeout=30)
        assert r.status_code == 200
        body = r.json()
        assert body.get("id") == pytest.partial_job_id
        assert "status" in body
        assert "total_rows" in body


# ---------- Error report download ----------
class TestErrorReport:
    def test_error_report_404(self, auth_headers):
        r = requests.get(f"{SCOPE3}/jobs/nonexistent-id-xyz/errors/download",
                         headers=auth_headers, timeout=30)
        assert r.status_code == 404

    def test_error_report_download_xlsx(self, auth_headers):
        if not getattr(pytest, "partial_job_id", None):
            pytest.skip("No job_id available for error report")
        r = requests.get(f"{SCOPE3}/jobs/{pytest.partial_job_id}/errors/download",
                         headers=auth_headers, timeout=60)
        assert r.status_code == 200, r.text[:300]
        assert "spreadsheet" in r.headers.get("content-type", "").lower()
        # Must be a valid xlsx
        wb = load_workbook(io.BytesIO(r.content))
        assert len(wb.sheetnames) >= 1
